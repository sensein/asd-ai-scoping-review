#!/usr/bin/env python3
"""Deduplicate enriched PRISMA records and write an audit-friendly workbook."""

from __future__ import annotations

import argparse
import difflib
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_INPUT = ROOT / "output" / "metadata_enrichment" / "enriched_records.csv"
DEFAULT_OUTPUT = ROOT / "output" / "deduplication" / "deduplicated_records.xlsx"


def temporary_output_path(output: Path) -> Path:
    if output.suffix:
        return output.with_name(f"{output.stem}.tmp{output.suffix}")
    return output.with_name(f"{output.name}.tmp")


def clean(value: object) -> str:
    return "" if pd.isna(value) else str(value).strip()


def normalize_doi(value: object) -> str:
    doi = clean(value).lower()
    doi = re.sub(r"^https?://(dx\.)?doi\.org/", "", doi)
    doi = re.sub(r"^doi:\s*", "", doi)
    return doi.strip()


def normalize_link(value: object) -> str:
    link = clean(value).lower()
    link = re.sub(r"^https?://", "", link)
    return link.rstrip("/")


def normalize_title(value: object) -> str:
    title = clean(value).lower()
    title = re.sub(r"[\W_]+", " ", title)
    return re.sub(r"\s+", " ", title).strip()


def read_records(path: Path, sheet: str | None) -> pd.DataFrame:
    if path.suffix.lower() in {".xlsx", ".xls"}:
        if sheet:
            return pd.read_excel(path, sheet_name=sheet, dtype=str).fillna("")
        workbook = pd.ExcelFile(path)
        sheet_name = "Enriched_Metadata" if "Enriched_Metadata" in workbook.sheet_names else workbook.sheet_names[0]
        return pd.read_excel(path, sheet_name=sheet_name, dtype=str).fillna("")
    return pd.read_csv(path, dtype=str).fillna("")


def find_near_title_match(title_key: str, retained: list[dict], threshold: float) -> dict | None:
    if len(title_key) < 24:
        return None
    best = None
    best_score = 0.0
    for candidate in retained:
        candidate_key = candidate.get("_title_key", "")
        if not candidate_key or abs(len(candidate_key) - len(title_key)) > 35:
            continue
        score = difflib.SequenceMatcher(None, title_key, candidate_key).ratio()
        if score > best_score:
            best = candidate
            best_score = score
    if best and best_score >= threshold:
        return best
    return None


def deduplicate(df: pd.DataFrame, near_title_threshold: float, use_near_title: bool) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    df = df.copy().fillna("")
    if "record_id" not in df.columns:
        df.insert(0, "record_id", [f"REC{i:05d}" for i in range(1, len(df) + 1)])
    for column in [
        "duplicate_status",
        "duplicate_reason",
        "duplicate_match_type",
        "retained_record_id",
        "retained_record_title",
    ]:
        if column not in df.columns:
            df[column] = ""

    doi_index: dict[str, dict] = {}
    link_index: dict[str, dict] = {}
    title_index: dict[str, dict] = {}
    retained_rows: list[dict] = []
    all_rows: list[dict] = []
    duplicate_log: list[dict] = []

    for _, raw in df.iterrows():
        row = raw.to_dict()
        doi_key = normalize_doi(row.get("doi", ""))
        link_key = normalize_link(row.get("link", ""))
        title_key = normalize_title(row.get("title", ""))

        match = None
        match_type = ""
        if doi_key and doi_key in doi_index:
            match = doi_index[doi_key]
            match_type = "DOI"
        elif link_key and link_key in link_index:
            match = link_index[link_key]
            match_type = "Same article link"
        elif title_key and title_key in title_index:
            match = title_index[title_key]
            match_type = "Exact title"
        elif use_near_title:
            match = find_near_title_match(title_key, retained_rows, near_title_threshold)
            if match:
                match_type = "Near-identical title"

        if match:
            row["duplicate_status"] = "Duplicate"
            row["duplicate_reason"] = f"Matched retained record {match.get('record_id', '')}"
            row["duplicate_match_type"] = match_type
            row["retained_record_id"] = match.get("record_id", "")
            row["retained_record_title"] = match.get("title", "")
            duplicate_log.append(
                {
                    "removed_record_id": row.get("record_id", ""),
                    "removed_record_title": row.get("title", ""),
                    "retained_record_id": match.get("record_id", ""),
                    "retained_record_title": match.get("title", ""),
                    "duplicate_reason": row["duplicate_reason"],
                    "duplicate_match_type": match_type,
                    "removed_doi": row.get("doi", ""),
                    "retained_doi": match.get("doi", ""),
                    "removed_link": row.get("link", ""),
                    "retained_link": match.get("link", ""),
                }
            )
        else:
            row["duplicate_status"] = "Retained"
            row["_title_key"] = title_key
            retained_rows.append(row)
            if doi_key:
                doi_index[doi_key] = row
            if link_key:
                link_index[link_key] = row
            if title_key:
                title_index[title_key] = row
        all_rows.append(row)

    all_df = pd.DataFrame(all_rows).drop(columns=["_title_key"], errors="ignore")
    retained_df = all_df[all_df["duplicate_status"].eq("Retained")].copy()
    log_df = pd.DataFrame(duplicate_log)
    return retained_df, all_df, log_df


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="22577A")
    font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for column_cells in ws.columns:
            letter = get_column_letter(column_cells[0].column)
            max_len = max(len(str(cell.value or "")[:100]) for cell in column_cells)
            ws.column_dimensions[letter].width = max(12, min(max_len + 2, 55))
        ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def write_workbook(output: Path, retained_df: pd.DataFrame, all_df: pd.DataFrame, log_df: pd.DataFrame) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    summary = pd.DataFrame(
        [
            {"metric": "Input records", "value": len(all_df)},
            {"metric": "Retained after deduplication", "value": len(retained_df)},
            {"metric": "Duplicates removed", "value": len(log_df)},
            {"metric": "Duplicate DOI matches", "value": int((log_df.get("duplicate_match_type", pd.Series(dtype=str)) == "DOI").sum())},
            {"metric": "Duplicate link matches", "value": int((log_df.get("duplicate_match_type", pd.Series(dtype=str)) == "Same article link").sum())},
            {"metric": "Duplicate exact-title matches", "value": int((log_df.get("duplicate_match_type", pd.Series(dtype=str)) == "Exact title").sum())},
            {"metric": "Duplicate near-title matches", "value": int((log_df.get("duplicate_match_type", pd.Series(dtype=str)) == "Near-identical title").sum())},
        ]
    )
    temp_output = temporary_output_path(output)
    with pd.ExcelWriter(temp_output, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        retained_df.to_excel(writer, sheet_name="Deduplicated_Retained", index=False)
        all_df.to_excel(writer, sheet_name="All_With_Duplicate_Status", index=False)
        log_df.to_excel(writer, sheet_name="Duplicate_Log", index=False)
    style_workbook(temp_output)
    temp_output.replace(output)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Deduplicate enriched PRISMA metadata records.")
    parser.add_argument("--input", default=str(DEFAULT_INPUT))
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    parser.add_argument("--sheet", default=None)
    parser.add_argument("--near-title-threshold", type=float, default=0.97)
    parser.add_argument("--no-near-title", action="store_true")
    return parser


def main() -> None:
    args = build_parser().parse_args()
    input_path = Path(args.input)
    if not input_path.exists():
        raise SystemExit(f"Input file not found: {input_path}")
    df = read_records(input_path, args.sheet)
    retained_df, all_df, log_df = deduplicate(df, args.near_title_threshold, not args.no_near_title)
    write_workbook(Path(args.output), retained_df, all_df, log_df)
    print(f"Wrote deduplicated workbook: {Path(args.output)}")
    print(f"Input records: {len(all_df)}")
    print(f"Retained after deduplication: {len(retained_df)}")
    print(f"Duplicates removed: {len(log_df)}")


if __name__ == "__main__":
    main()
