#!/usr/bin/env python3
"""Shared utilities for reusable PRISMA pipeline scripts."""

from __future__ import annotations

import html
import difflib
import json
import os
import re
import time
import urllib.error
import urllib.request
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
DATA_ROOT = Path(os.environ.get("ASD_REVIEW_DATA_ROOT", ROOT / "data")).expanduser()
OUTPUT_ROOT = Path(os.environ.get("ASD_REVIEW_OUTPUT_ROOT", ROOT / "output")).expanduser()
if not DATA_ROOT.is_absolute():
    DATA_ROOT = ROOT / DATA_ROOT
if not OUTPUT_ROOT.is_absolute():
    OUTPUT_ROOT = ROOT / OUTPUT_ROOT
DEFAULT_CRITERIA = ROOT / "config" / "review_criteria.json"
EXAMPLE_CRITERIA = Path(__file__).resolve().parent / "review_criteria.example.json"

INCLUDE = "Include"
MAYBE = "Maybe"
EXCLUDE = "Exclude"
DECISION_LABELS = (INCLUDE, MAYBE, EXCLUDE)
MISSING_IDENTIFIER_TOKENS = {
    "", "-", "--", "?", "??", "n/a", "na", "none", "null", "not available", "not found", "unknown",
}


def temporary_output_path(output: Path) -> Path:
    if output.suffix:
        return output.with_name(f"{output.stem}.tmp{output.suffix}")
    return output.with_name(f"{output.name}.tmp")


def clean(value: Any) -> str:
    text = html.unescape(str(value or ""))
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def normalize(value: Any) -> str:
    text = html.unescape(str(value or "")).lower()
    text = re.sub(r"\.pdf$", "", text)
    text = re.sub(r"[_\-/]+", " ", text)
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def normalize_title(value: Any) -> str:
    return normalize(value)


def normalize_doi(value: Any) -> str:
    doi = str(value or "").strip()
    if doi.lower() in MISSING_IDENTIFIER_TOKENS:
        return ""
    doi = re.sub(r"^https?://(dx\.)?doi\.org/", "", doi, flags=re.IGNORECASE)
    doi = re.sub(r"^doi:\s*", "", doi, flags=re.IGNORECASE)
    embedded = re.search(r"10\.\d{4,9}/[^\s\"<>]+", doi, flags=re.IGNORECASE)
    if embedded:
        doi = embedded.group(0)
    doi = doi.strip().rstrip(".,; ").lower()
    return "" if doi in MISSING_IDENTIFIER_TOKENS else doi


def normalize_link(value: Any) -> str:
    link = clean(value).lower()
    if link in MISSING_IDENTIFIER_TOKENS:
        return ""
    link = re.sub(r"^https?://", "", link)
    link = link.rstrip("/")
    return "" if link in MISSING_IDENTIFIER_TOKENS else link


def find_duplicate_keys(
    df: pd.DataFrame,
    *,
    doi_col: str = "doi",
    title_col: str = "title",
    link_col: str = "link",
) -> pd.DataFrame:
    """Return a non-destructive audit of exact duplicate DOI, title, and link keys."""
    normalizers = {
        "DOI": (doi_col, normalize_doi),
        "Exact title": (title_col, normalize_title),
        "Link": (link_col, normalize_link),
    }
    rows: list[dict[str, Any]] = []
    for match_type, (column, normalizer) in normalizers.items():
        if column not in df.columns:
            continue
        normalized = df[column].map(normalizer)
        counts = normalized[normalized.ne("")].value_counts()
        duplicate_values = counts[counts.gt(1)]
        for normalized_value, record_count in duplicate_values.items():
            for row_index in df.index[normalized.eq(normalized_value)]:
                row = df.loc[row_index]
                rows.append(
                    {
                        "duplicate_match_type": match_type,
                        "normalized_value": normalized_value,
                        "record_count": int(record_count),
                        "row_index": row_index,
                        "record_id": row.get("record_id", ""),
                        "title": row.get(title_col, ""),
                        "doi": row.get(doi_col, ""),
                        "link": row.get(link_col, ""),
                    }
                )
    columns = [
        "duplicate_match_type", "normalized_value", "record_count", "row_index", "record_id", "title", "doi", "link",
    ]
    return pd.DataFrame(rows, columns=columns)


def title_similarity(left: Any, right: Any, include_token_jaccard: bool = False) -> float:
    maximum, sequence_score, _ = title_similarity_components(left, right)
    return maximum if include_token_jaccard else sequence_score


def title_similarity_components(left: Any, right: Any) -> tuple[float, float, float]:
    left_norm = normalize_title(left)
    right_norm = normalize_title(right)
    if not left_norm or not right_norm:
        return 0.0, 0.0, 0.0
    if left_norm == right_norm:
        return 1.0, 1.0, 1.0
    sequence_score = difflib.SequenceMatcher(None, left_norm, right_norm).ratio()
    left_tokens = set(left_norm.split())
    right_tokens = set(right_norm.split())
    jaccard = len(left_tokens & right_tokens) / len(left_tokens | right_tokens) if left_tokens and right_tokens else 0.0
    return max(sequence_score, jaccard), sequence_score, jaccard


def parse_decisions(value: str) -> set[str]:
    requested = {item.strip().title() for item in str(value).split(",") if item.strip()}
    unknown = requested - set(DECISION_LABELS)
    if unknown:
        raise ValueError(f"Unknown decision label(s): {sorted(unknown)}; use {DECISION_LABELS}")
    return requested


def read_table(path: Path, sheet: str | int | None = None, preferred_sheet: str | None = None) -> pd.DataFrame:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xls"}:
        if sheet is not None:
            return pd.read_excel(path, sheet_name=sheet, dtype=str).fillna("")
        workbook = pd.ExcelFile(path)
        sheet_name = preferred_sheet if preferred_sheet in workbook.sheet_names else workbook.sheet_names[0]
        return pd.read_excel(path, sheet_name=sheet_name, dtype=str).fillna("")
    if suffix == ".jsonl":
        rows = [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]
        return pd.DataFrame(rows).fillna("")
    return pd.read_csv(path, dtype=str).fillna("")


def excel_safe_df(df: pd.DataFrame) -> pd.DataFrame:
    cleaned = df.copy().fillna("")
    illegal = re.compile(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]")
    for col in cleaned.columns:
        if cleaned[col].dtype == object:
            cleaned[col] = cleaned[col].map(lambda value: illegal.sub(" ", str(value)) if value is not None else "")
    return cleaned


def style_workbook(path: Path, header_color: str = "22577A") -> None:
    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor=header_color)
    font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        if not ws.max_row or not ws.max_column:
            continue
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for column_cells in ws.columns:
            letter = get_column_letter(column_cells[0].column)
            max_len = max(len(str(cell.value or "")[:100]) for cell in column_cells)
            ws.column_dimensions[letter].width = max(12, min(max_len + 2, 60))
    wb.save(path)


def write_workbook(output: Path, sheets: dict[str, pd.DataFrame], header_color: str = "22577A") -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    temp_output = temporary_output_path(output)
    with pd.ExcelWriter(temp_output, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            excel_safe_df(df).to_excel(writer, sheet_name=sheet_name[:31], index=False)
    style_workbook(temp_output, header_color)
    temp_output.replace(output)


def count_frame(series: pd.Series, name: str, blank_label: str = "Blank") -> pd.DataFrame:
    return series.fillna("").replace("", blank_label).value_counts(dropna=False).rename_axis(name).reset_index(name="count")


def load_criteria(path: Path | str | None = None) -> dict[str, Any]:
    criteria_path = Path(path) if path else DEFAULT_CRITERIA
    if not criteria_path.exists():
        raise FileNotFoundError(
            "PRISMA review criteria are required and were not found at "
            f"{criteria_path}. Copy {EXAMPLE_CRITERIA} to {DEFAULT_CRITERIA}, "
            "replace every placeholder, and rerun (or pass --criteria explicitly)."
        )
    criteria = json.loads(criteria_path.read_text(encoding="utf-8"))
    required_groups = ("population", "method", "data_source", "outcome")
    missing_groups = [group for group in required_groups if not terms(criteria, group)]
    placeholder_terms = [
        term
        for group in required_groups
        for term in terms(criteria, group)
        if "example " in term.lower() or "placeholder" in term.lower()
    ]
    if missing_groups or placeholder_terms:
        details = []
        if missing_groups:
            details.append(f"missing required term groups: {missing_groups}")
        if placeholder_terms:
            details.append(f"placeholder terms remain: {placeholder_terms}")
        raise ValueError(f"Unsafe PRISMA criteria file {criteria_path}: " + "; ".join(details))
    criteria["_criteria_path"] = str(criteria_path.resolve())
    return criteria


def load_json_cache(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise ValueError(f"Invalid JSON cache {path}: {exc}") from exc


def save_json_cache(path: Path, cache: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = temporary_output_path(path)
    temporary.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
    temporary.replace(path)


def request_text_cached(
    url: str,
    cache: dict[str, Any],
    *,
    user_agent: str,
    delay_seconds: float,
    timeout: int,
    retries: int,
    accept: str = "application/json",
) -> str | None:
    """Fetch a URL with the pipeline's shared retry and in-memory cache policy."""
    cached = cache.get(url)
    if isinstance(cached, dict) and "_response_text" in cached:
        return cached["_response_text"]
    if isinstance(cached, str):
        return cached
    for attempt in range(retries + 1):
        if delay_seconds:
            time.sleep(delay_seconds)
        request = urllib.request.Request(url, headers={"User-Agent": user_agent, "Accept": accept})
        try:
            with urllib.request.urlopen(request, timeout=timeout) as response:
                text = response.read().decode("utf-8", errors="replace")
            cache[url] = {"_response_text": text}
            return text
        except (urllib.error.HTTPError, urllib.error.URLError, TimeoutError):
            if attempt >= retries:
                return None
            time.sleep(min(2**attempt, 8))
    return None


def request_json_cached(
    url: str,
    cache: dict[str, Any],
    **kwargs: Any,
) -> dict[str, Any] | None:
    text = request_text_cached(url, cache, **kwargs)
    if not text:
        return None
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return None


def apply_year_overrides(criteria: dict[str, Any], start_year: int | None, end_year: int | None) -> dict[str, Any]:
    updated = json.loads(json.dumps(criteria))
    date_range = dict(updated.get("date_range") or {})
    if start_year is not None:
        date_range["start_year"] = start_year
    if end_year is not None:
        date_range["end_year"] = end_year
    if date_range:
        updated["date_range"] = date_range
    return updated


def group_config(criteria: dict[str, Any], group: str) -> dict[str, Any]:
    groups = criteria.get("term_groups") or {}
    value = groups.get(group, {})
    if isinstance(value, list):
        return {"terms": value}
    return value if isinstance(value, dict) else {}


def terms(criteria: dict[str, Any], group: str) -> list[str]:
    value = group_config(criteria, group).get("terms", [])
    return [str(item) for item in value if str(item).strip()]


def group_label(criteria: dict[str, Any], group: str) -> str:
    return str(group_config(criteria, group).get("label") or group.replace("_", " "))


def group_reason(criteria: dict[str, Any], group: str, fallback: str) -> str:
    return str(group_config(criteria, group).get("missing_reason") or group_config(criteria, group).get("reason") or fallback)


def has_any(text: str, term_list: list[str]) -> bool:
    padded = f" {normalize(text)} "
    return any(f" {normalize(term)} " in padded for term in term_list if normalize(term))


def matched_terms(text: str, term_list: list[str], limit: int = 8) -> list[str]:
    padded = f" {normalize(text)} "
    hits = [term for term in term_list if normalize(term) and f" {normalize(term)} " in padded]
    return hits[:limit]


def has_group(text: str, criteria: dict[str, Any], group: str) -> bool:
    return has_any(text, terms(criteria, group))


def matched_group_terms(text: str, criteria: dict[str, Any], group: str, limit: int = 8) -> list[str]:
    return matched_terms(text, terms(criteria, group), limit)


def evidence_note(text: str, criteria: dict[str, Any], groups: tuple[str, ...] = ("population", "method", "data_source", "outcome")) -> str:
    parts = []
    for group in groups:
        hits = matched_group_terms(text, criteria, group, 5)
        if hits:
            parts.append(f"{group_label(criteria, group)}: {', '.join(hits)}")
    return " | ".join(parts)


def snippet(text: str, term_list: list[str], radius: int = 170) -> str:
    lower = text.lower()
    for term in term_list:
        idx = lower.find(term.lower())
        if idx >= 0:
            start = max(0, idx - radius)
            end = min(len(text), idx + len(term) + radius)
            return re.sub(r"\s+", " ", text[start:end]).strip()
    return ""


def parse_year(value: Any, text: str = "") -> int | None:
    raw = str(value or "")
    match = re.search(r"(20[0-3]\d|19\d{2})", raw)
    if match:
        return int(match.group(1))
    years = [int(item) for item in re.findall(r"\b(20[0-3]\d|19\d{2})\b", text[:6000])]
    return min(years) if years else None


def english_likely(text: str, language: Any = "") -> bool:
    lang = normalize(language)
    if lang and lang not in {"en", "eng", "english"} and "english" not in lang:
        return False
    if len(text) < 500:
        return True
    common = sum(text.lower().count(f" {word} ") for word in ["the", "and", "of", "in", "to", "with", "for"])
    return common >= 15


def _year_flag(criteria: dict[str, Any], year: int | None) -> str:
    date_range = criteria.get("date_range") or {}
    start = date_range.get("start_year")
    end = date_range.get("end_year")
    if not start and not end:
        return "not_configured"
    if year is None:
        return "unclear"
    if start and year < int(start):
        return "no"
    if end and year > int(end):
        return "no"
    return "yes"


def screen_eligibility(
    *,
    title: Any,
    body: Any,
    criteria: dict[str, Any],
    stage: str,
    document_type: Any = "",
    language: Any = "",
    year_value: Any = "",
) -> dict[str, str]:
    title_text = clean(title)
    body_text = clean(body)
    full = f"{title_text}\n{body_text}".strip()
    front = full[:9000]
    options = criteria.get("screening_options") or {}
    min_full_text_chars = int(options.get("min_full_text_chars", 3000))
    year = parse_year(year_value, body_text)

    flags = {
        "publication_year": _year_flag(criteria, year),
        "english": "yes" if english_likely(body_text or title_text, language) else "no",
        "population": "yes" if has_group(full, criteria, "population") else "no",
        "method": "yes" if has_group(full, criteria, "method") else "no",
        "data_source": "yes" if has_group(full, criteria, "data_source") else "no",
        "outcome": "yes" if has_group(full, criteria, "outcome") else "no",
        "human_participants": "yes" if has_group(full, criteria, "human") else "unclear",
        "primary_empirical": "yes" if has_group(full, criteria, "empirical") else "unclear",
        "secondary_research": "yes" if has_group(front + " " + clean(document_type), criteria, "secondary_research") else "no",
        "abstract_only": "yes" if has_group(front, criteria, "abstract_only") or (stage == "full_text" and len(body_text) < min_full_text_chars) else "no",
        "non_human_only": "yes" if has_group(full, criteria, "non_human") and not has_group(full, criteria, "human") else "no",
        "preprint_or_non_peer_review": "yes" if has_group(front + " " + clean(document_type), criteria, "preprint_or_non_peer_review") else "no",
        "intervention": "yes" if has_group(full, criteria, "intervention") else "no",
    }

    decision = "Include"
    reason = ""
    confidence = "medium"
    notes: list[str] = []

    if stage == "title":
        if not title_text:
            decision, reason, confidence = "Maybe", "", "low"
            notes.append("Missing title.")
        elif flags["secondary_research"] == "yes":
            decision, reason, confidence = "Exclude", group_reason(criteria, "secondary_research", "Secondary research"), "medium"
        elif flags["non_human_only"] == "yes":
            decision, reason, confidence = "Exclude", group_reason(criteria, "non_human", "Non-human subjects"), "medium"
        elif flags["population"] != "yes":
            decision, reason, confidence = "Exclude", group_reason(criteria, "population", "Target population not detected"), "high"
        elif flags["method"] != "yes":
            decision, reason, confidence = "Maybe", "", "low"
            notes.append(group_reason(criteria, "method", "Target method not detected in title."))
        elif flags["intervention"] == "yes" and flags["outcome"] != "yes":
            decision, reason, confidence = "Exclude", group_reason(criteria, "intervention", "Intervention-only focus without eligible outcome"), "medium"
        elif flags["data_source"] == "yes" or flags["outcome"] == "yes":
            decision, confidence = "Include", "medium"
        else:
            decision, confidence = "Maybe", "low"
            notes.append("Title contains population and method terms, but eligibility focus is unclear.")
    elif stage == "abstract":
        if not body_text:
            decision, reason, confidence = "Maybe", "", "low"
            notes.append("Missing abstract; manual lookup or full text review needed before final exclusion.")
        elif flags["english"] == "no":
            decision, reason, confidence = "Exclude", "Not English", "medium"
        elif flags["secondary_research"] == "yes":
            decision, reason, confidence = "Exclude", group_reason(criteria, "secondary_research", "Secondary research"), "medium"
        elif flags["non_human_only"] == "yes":
            decision, reason, confidence = "Exclude", group_reason(criteria, "non_human", "Non-human subjects"), "medium"
        elif flags["population"] != "yes":
            decision, reason, confidence = "Exclude", group_reason(criteria, "population", "Target population not detected"), "high"
        elif flags["method"] != "yes":
            decision, reason, confidence = "Exclude", group_reason(criteria, "method", "Target method not detected"), "medium"
        elif flags["data_source"] != "yes" and bool(options.get("require_data_source_for_abstract", True)):
            decision, reason, confidence = "Exclude", group_reason(criteria, "data_source", "Eligible data source not detected"), "medium"
        elif flags["outcome"] != "yes":
            if flags["intervention"] == "yes":
                decision, reason, confidence = "Exclude", group_reason(criteria, "intervention", "Intervention-only focus without eligible outcome"), "medium"
            else:
                decision, confidence = "Maybe", "low"
                notes.append(group_reason(criteria, "outcome", "Eligible outcome not clearly detected."))
    elif stage == "full_text":
        if flags["abstract_only"] == "yes":
            decision, reason, confidence = "Exclude", group_reason(criteria, "abstract_only", "Abstract-only or insufficient full text"), "medium"
        elif flags["publication_year"] == "no":
            date_range = criteria.get("date_range") or {}
            decision, reason, confidence = "Exclude", f"Outside eligible publication timeframe ({date_range.get('start_year', '')}-{date_range.get('end_year', '')})", "high"
        elif flags["english"] == "no":
            decision, reason, confidence = "Exclude", "Not English", "high"
        elif flags["secondary_research"] == "yes":
            decision, reason, confidence = "Exclude", group_reason(criteria, "secondary_research", "Secondary research"), "high"
        elif flags["preprint_or_non_peer_review"] == "yes" and bool(options.get("exclude_preprints", True)):
            decision, reason, confidence = "Exclude", group_reason(criteria, "preprint_or_non_peer_review", "Not peer reviewed"), "medium"
        elif flags["non_human_only"] == "yes":
            decision, reason, confidence = "Exclude", group_reason(criteria, "non_human", "Non-human subjects"), "high"
        elif flags["population"] != "yes":
            decision, reason, confidence = "Exclude", group_reason(criteria, "population", "Target population not detected"), "medium"
        elif flags["outcome"] != "yes":
            decision, reason, confidence = "Exclude", group_reason(criteria, "outcome", "Eligible outcome not detected"), "medium"
        elif flags["method"] != "yes":
            decision, reason, confidence = "Exclude", group_reason(criteria, "method", "Target method not detected"), "medium"
        elif flags["data_source"] != "yes":
            decision, reason, confidence = "Exclude", group_reason(criteria, "data_source", "Eligible data source not detected"), "medium"
        elif flags["primary_empirical"] != "yes":
            decision, reason, confidence = "Maybe", group_reason(criteria, "empirical", "Primary empirical evidence unclear"), "low"

    evidence = evidence_note(full, criteria)
    if decision == "Include":
        notes.append("Configured eligibility terms were detected.")
    if evidence:
        notes.append(f"Matched terms: {evidence}")
    if not notes:
        notes.append("No additional notes.")

    return {
        "decision": decision,
        "reason": reason,
        "confidence": confidence,
        "notes": " ".join(notes),
        "screened_year": str(year or ""),
        "criterion_publication_year": flags["publication_year"],
        "criterion_english": flags["english"],
        "criterion_population": flags["population"],
        "criterion_method": flags["method"],
        "criterion_data_source": flags["data_source"],
        "criterion_outcome": flags["outcome"],
        "criterion_human_participants": flags["human_participants"],
        "criterion_primary_empirical": flags["primary_empirical"],
        "criterion_secondary_research": flags["secondary_research"],
        "criterion_abstract_only": flags["abstract_only"],
        "criterion_non_human_only": flags["non_human_only"],
        "criterion_preprint_or_non_peer_review": flags["preprint_or_non_peer_review"],
        "evidence_population": snippet(full, terms(criteria, "population")),
        "evidence_method": snippet(full, terms(criteria, "method")),
        "evidence_data_source": snippet(full, terms(criteria, "data_source")),
        "evidence_outcome": snippet(full, terms(criteria, "outcome")),
        "evidence_empirical": snippet(full, terms(criteria, "empirical")),
    }
