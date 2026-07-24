#!/usr/bin/env python3
"""Find missing abstracts for title Include/Maybe PRISMA records.

The script starts from title_include_maybe_metadata.xlsx, preserves any abstracts
already present, then tries public scholarly metadata APIs for rows that are
still missing abstracts. It writes an audit-friendly Excel workbook containing
the enriched rows, lookup summary counts, source counts, and a review queue for
records that still need manual abstract lookup.
"""

from __future__ import annotations

import argparse
import difflib
import html
import json
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_INPUT = ROOT / "output" / "abstract_screening" / "title_include_maybe_metadata.xlsx"
DEFAULT_OUTPUT = ROOT / "output" / "abstract_finding" / "title_include_maybe_with_abstracts.xlsx"
DEFAULT_CACHE = ROOT / "output" / "abstract_finding" / "abstract_api_cache.json"


def temporary_output_path(output: Path) -> Path:
    if output.suffix:
        return output.with_name(f"{output.stem}.tmp{output.suffix}")
    return output.with_name(f"{output.name}.tmp")


HEADER_ALIASES = {
    "record_id": ["record_id", "record id", "id"],
    "title": ["title", "article title", "paper title"],
    "doi": ["doi", "digital object identifier"],
    "link": ["link", "url", "article link", "source url"],
    "journal": ["journal", "source_database", "source database", "source", "venue"],
    "year_published": ["year_published", "year published", "year", "publication year"],
    "authors": ["authors", "author"],
    "abstract": ["abstract"],
}


@dataclass
class AbstractCandidate:
    source: str
    abstract: str
    title: str = ""
    doi: str = ""
    year: str = ""
    journal: str = ""
    url: str = ""
    score: float = 0.0
    notes: str = ""


def clean_text(value: Any) -> str:
    text = html.unescape(str(value or ""))
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def normalize_header(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower().replace("_", " "))


def normalize_title(value: Any) -> str:
    text = html.unescape(str(value or "")).lower()
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"[\W_]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_doi(value: Any) -> str:
    doi = str(value or "").strip()
    doi = re.sub(r"^https?://(dx\.)?doi\.org/", "", doi, flags=re.IGNORECASE)
    doi = re.sub(r"^doi:\s*", "", doi, flags=re.IGNORECASE)
    return doi.strip().rstrip(".")


def doi_from_link(value: Any) -> str:
    text = str(value or "").strip()
    match = re.search(r"10\.\d{4,9}/[^\s\"'<>]+", text, flags=re.IGNORECASE)
    return normalize_doi(match.group(0)) if match else ""


def title_similarity(left: Any, right: Any) -> float:
    left_norm = normalize_title(left)
    right_norm = normalize_title(right)
    if not left_norm or not right_norm:
        return 0.0
    if left_norm == right_norm:
        return 1.0
    return difflib.SequenceMatcher(None, left_norm, right_norm).ratio()


def journal_bonus(input_journal: Any, candidate_journal: Any) -> float:
    left = normalize_title(input_journal)
    right = normalize_title(candidate_journal)
    if not left or not right:
        return 0.0
    if left == right:
        return 0.03
    if left in right or right in left:
        return 0.02
    return min(0.02, len(set(left.split()) & set(right.split())) * 0.004)


def confidence(score: float) -> str:
    if score >= 0.96:
        return "high"
    if score >= 0.88:
        return "medium"
    if score > 0:
        return "low"
    return "none"


def reconstruct_openalex_abstract(index: Any) -> str:
    if not isinstance(index, dict):
        return ""
    words: list[tuple[int, str]] = []
    for word, positions in index.items():
        if not isinstance(positions, list):
            continue
        for position in positions:
            if isinstance(position, int):
                words.append((position, word))
    return " ".join(word for _, word in sorted(words)).strip()


def inverted_abstract_to_text(index: Any) -> str:
    return clean_text(reconstruct_openalex_abstract(index))


def load_cache(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def save_cache(path: Path, cache: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = temporary_output_path(path)
    temp_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
    temp_path.replace(path)


def request_text(
    url: str,
    cache: dict[str, Any],
    *,
    user_agent: str,
    delay_seconds: float,
    timeout: int,
    retries: int,
    accept: str = "application/json",
) -> str | None:
    cache_key = f"TEXT::{url}"
    if cache_key in cache:
        return cache[cache_key]

    for attempt in range(retries + 1):
        if delay_seconds:
            time.sleep(delay_seconds)
        request = urllib.request.Request(url, headers={"User-Agent": user_agent, "Accept": accept})
        try:
            with urllib.request.urlopen(request, timeout=timeout) as response:
                text = response.read().decode("utf-8", errors="replace")
                cache[cache_key] = text
                return text
        except urllib.error.HTTPError as exc:
            if exc.code in {429, 500, 502, 503, 504} and attempt < retries:
                time.sleep(min(8, 2 ** attempt))
                continue
            cache[cache_key] = {"error": f"HTTP {exc.code}"}
            return None
        except (urllib.error.URLError, TimeoutError):
            if attempt < retries:
                time.sleep(min(8, 2 ** attempt))
                continue
            return None
    return None


def request_json(
    url: str,
    cache: dict[str, Any],
    *,
    user_agent: str,
    delay_seconds: float,
    timeout: int,
    retries: int,
) -> dict[str, Any] | None:
    cache_key = f"JSON::{url}"
    if cache_key in cache:
        cached = cache[cache_key]
        return cached if isinstance(cached, dict) else None
    text = request_text(
        url,
        cache,
        user_agent=user_agent,
        delay_seconds=delay_seconds,
        timeout=timeout,
        retries=retries,
    )
    if not text or not isinstance(text, str):
        return None
    try:
        payload = json.loads(text)
        cache[cache_key] = payload
        return payload
    except json.JSONDecodeError:
        return None


def crossref_candidate_from_item(item: dict[str, Any], source: str, input_title: str, input_journal: str) -> AbstractCandidate | None:
    abstract = clean_text(item.get("abstract", ""))
    if not abstract:
        return None
    title = clean_text((item.get("title") or [""])[0] if isinstance(item.get("title"), list) else item.get("title", ""))
    journal = clean_text((item.get("container-title") or [""])[0] if isinstance(item.get("container-title"), list) else "")
    score = title_similarity(input_title, title) + journal_bonus(input_journal, journal)
    return AbstractCandidate(
        source=source,
        abstract=abstract,
        title=title,
        doi=normalize_doi(item.get("DOI", "")),
        year=str(((item.get("published-print") or item.get("published-online") or item.get("published") or {}).get("date-parts") or [[""]])[0][0]),
        journal=journal,
        url=clean_text(item.get("URL", "")),
        score=min(score, 1.0),
        notes="Crossref abstract field.",
    )


def query_crossref_by_doi(doi: str, title: str, journal: str, cache: dict[str, Any], args: argparse.Namespace) -> AbstractCandidate | None:
    if not doi:
        return None
    url = f"https://api.crossref.org/works/{urllib.parse.quote(doi)}"
    payload = request_json(url, cache, user_agent=args.user_agent, delay_seconds=args.delay, timeout=args.timeout, retries=args.retries)
    item = (payload or {}).get("message", {})
    candidate = crossref_candidate_from_item(item, "crossref_doi", title, journal) if item else None
    if candidate:
        candidate.score = 1.0
    return candidate


def query_crossref_by_title(title: str, journal: str, cache: dict[str, Any], args: argparse.Namespace) -> list[AbstractCandidate]:
    if not title:
        return []
    params = urllib.parse.urlencode({"query.bibliographic": title, "rows": 5})
    url = f"https://api.crossref.org/works?{params}"
    payload = request_json(url, cache, user_agent=args.user_agent, delay_seconds=args.delay, timeout=args.timeout, retries=args.retries)
    items = ((payload or {}).get("message") or {}).get("items") or []
    candidates = []
    for item in items:
        candidate = crossref_candidate_from_item(item, "crossref_title", title, journal)
        if candidate:
            candidates.append(candidate)
    return candidates


def openalex_candidate_from_work(work: dict[str, Any], source: str, input_title: str, input_journal: str) -> AbstractCandidate | None:
    abstract = inverted_abstract_to_text(work.get("abstract_inverted_index"))
    if not abstract:
        return None
    title = clean_text(work.get("display_name", ""))
    primary = work.get("primary_location") or {}
    source_obj = primary.get("source") or {}
    journal = clean_text(source_obj.get("display_name", ""))
    ids = work.get("ids") or {}
    doi = normalize_doi(ids.get("doi", ""))
    score = title_similarity(input_title, title) + journal_bonus(input_journal, journal)
    return AbstractCandidate(
        source=source,
        abstract=abstract,
        title=title,
        doi=doi,
        year=str(work.get("publication_year", "") or ""),
        journal=journal,
        url=clean_text(ids.get("openalex", "") or primary.get("landing_page_url", "")),
        score=min(score, 1.0),
        notes="OpenAlex abstract_inverted_index.",
    )


def query_openalex_by_doi(doi: str, title: str, journal: str, cache: dict[str, Any], args: argparse.Namespace) -> AbstractCandidate | None:
    if not doi:
        return None
    params = urllib.parse.urlencode({"filter": f"doi:{doi}", "per-page": 1})
    url = f"https://api.openalex.org/works?{params}"
    payload = request_json(url, cache, user_agent=args.user_agent, delay_seconds=args.delay, timeout=args.timeout, retries=args.retries)
    results = (payload or {}).get("results") or []
    if not results:
        return None
    candidate = openalex_candidate_from_work(results[0], "openalex_doi", title, journal)
    if candidate:
        candidate.score = 1.0
    return candidate


def query_openalex_by_title(title: str, journal: str, cache: dict[str, Any], args: argparse.Namespace) -> list[AbstractCandidate]:
    if not title:
        return []
    params = urllib.parse.urlencode({"search": title, "per-page": 5})
    url = f"https://api.openalex.org/works?{params}"
    payload = request_json(url, cache, user_agent=args.user_agent, delay_seconds=args.delay, timeout=args.timeout, retries=args.retries)
    results = (payload or {}).get("results") or []
    candidates = []
    for work in results:
        candidate = openalex_candidate_from_work(work, "openalex_title", title, journal)
        if candidate:
            candidates.append(candidate)
    return candidates


def semantic_candidate_from_paper(paper: dict[str, Any], source: str, input_title: str, input_journal: str) -> AbstractCandidate | None:
    abstract = clean_text(paper.get("abstract", ""))
    if not abstract:
        return None
    title = clean_text(paper.get("title", ""))
    external = paper.get("externalIds") or {}
    doi = normalize_doi(external.get("DOI", ""))
    journal = clean_text(paper.get("venue", ""))
    score = title_similarity(input_title, title) + journal_bonus(input_journal, journal)
    return AbstractCandidate(
        source=source,
        abstract=abstract,
        title=title,
        doi=doi,
        year=str(paper.get("year", "") or ""),
        journal=journal,
        url=clean_text(paper.get("url", "")),
        score=min(score, 1.0),
        notes="Semantic Scholar abstract field.",
    )


def query_semantic_by_doi(doi: str, title: str, journal: str, cache: dict[str, Any], args: argparse.Namespace) -> AbstractCandidate | None:
    if not doi:
        return None
    fields = "title,abstract,year,venue,url,externalIds"
    url = f"https://api.semanticscholar.org/graph/v1/paper/DOI:{urllib.parse.quote(doi)}?fields={fields}"
    payload = request_json(url, cache, user_agent=args.user_agent, delay_seconds=args.delay, timeout=args.timeout, retries=args.retries)
    candidate = semantic_candidate_from_paper(payload or {}, "semantic_scholar_doi", title, journal) if payload else None
    if candidate:
        candidate.score = 1.0
    return candidate


def query_semantic_by_title(title: str, journal: str, cache: dict[str, Any], args: argparse.Namespace) -> list[AbstractCandidate]:
    if not title:
        return []
    fields = "title,abstract,year,venue,url,externalIds"
    params = urllib.parse.urlencode({"query": title, "limit": 5, "fields": fields})
    url = f"https://api.semanticscholar.org/graph/v1/paper/search?{params}"
    payload = request_json(url, cache, user_agent=args.user_agent, delay_seconds=args.delay, timeout=args.timeout, retries=args.retries)
    papers = (payload or {}).get("data") or []
    candidates = []
    for paper in papers:
        candidate = semantic_candidate_from_paper(paper, "semantic_scholar_title", title, journal)
        if candidate:
            candidates.append(candidate)
    return candidates


def europe_pmc_candidate_from_result(result: dict[str, Any], source: str, input_title: str, input_journal: str) -> AbstractCandidate | None:
    abstract = clean_text(result.get("abstractText", ""))
    if not abstract:
        return None
    title = clean_text(result.get("title", ""))
    journal = clean_text(result.get("journalTitle", ""))
    score = title_similarity(input_title, title) + journal_bonus(input_journal, journal)
    doi = normalize_doi(result.get("doi", ""))
    pmid = clean_text(result.get("pmid", ""))
    pmcid = clean_text(result.get("pmcid", ""))
    url = f"https://europepmc.org/article/MED/{pmid}" if pmid else f"https://europepmc.org/article/PMC/{pmcid}" if pmcid else ""
    return AbstractCandidate(
        source=source,
        abstract=abstract,
        title=title,
        doi=doi,
        year=clean_text(result.get("pubYear", "")),
        journal=journal,
        url=url,
        score=min(score, 1.0),
        notes="Europe PMC abstractText.",
    )


def query_europe_pmc(query: str, source: str, title: str, journal: str, cache: dict[str, Any], args: argparse.Namespace) -> list[AbstractCandidate]:
    if not query:
        return []
    params = urllib.parse.urlencode({"query": query, "format": "json", "pageSize": 5, "resultType": "core"})
    url = f"https://www.ebi.ac.uk/europepmc/webservices/rest/search?{params}"
    payload = request_json(url, cache, user_agent=args.user_agent, delay_seconds=args.delay, timeout=args.timeout, retries=args.retries)
    results = (((payload or {}).get("resultList") or {}).get("result")) or []
    candidates = []
    for result in results:
        candidate = europe_pmc_candidate_from_result(result, source, title, journal)
        if candidate:
            candidates.append(candidate)
    return candidates


def query_europe_pmc_by_doi(doi: str, title: str, journal: str, cache: dict[str, Any], args: argparse.Namespace) -> AbstractCandidate | None:
    if not doi:
        return None
    candidates = query_europe_pmc(f'DOI:"{doi}"', "europe_pmc_doi", title, journal, cache, args)
    if not candidates:
        return None
    best = max(candidates, key=lambda item: item.score)
    best.score = 1.0
    return best


def query_europe_pmc_by_title(title: str, journal: str, cache: dict[str, Any], args: argparse.Namespace) -> list[AbstractCandidate]:
    if not title:
        return []
    escaped = title.replace('"', " ")
    return query_europe_pmc(f'TITLE:"{escaped}"', "europe_pmc_title", title, journal, cache, args)


def pubmed_candidate_from_xml(xml_text: str, source: str, input_title: str, input_journal: str) -> AbstractCandidate | None:
    try:
        root = ET.fromstring(xml_text)
    except ET.ParseError:
        return None
    article = root.find(".//PubmedArticle")
    if article is None:
        return None
    title = clean_text("".join(article.findtext(".//ArticleTitle", default="") or ""))
    abstract_parts = []
    for node in article.findall(".//AbstractText"):
        label = node.attrib.get("Label", "")
        text = clean_text("".join(node.itertext()))
        if text:
            abstract_parts.append(f"{label}: {text}" if label else text)
    abstract = clean_text(" ".join(abstract_parts))
    if not abstract:
        return None
    journal = clean_text(article.findtext(".//Journal/Title", default="") or "")
    year = clean_text(article.findtext(".//PubDate/Year", default="") or "")
    doi = ""
    for node in article.findall(".//ArticleId"):
        if node.attrib.get("IdType") == "doi":
            doi = normalize_doi(node.text or "")
            break
    pmid = clean_text(article.findtext(".//PMID", default="") or "")
    score = title_similarity(input_title, title) + journal_bonus(input_journal, journal)
    return AbstractCandidate(
        source=source,
        abstract=abstract,
        title=title,
        doi=doi,
        year=year,
        journal=journal,
        url=f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/" if pmid else "",
        score=min(score, 1.0),
        notes="PubMed XML AbstractText.",
    )


def pubmed_ids_for_query(query: str, cache: dict[str, Any], args: argparse.Namespace) -> list[str]:
    params = urllib.parse.urlencode({"db": "pubmed", "term": query, "retmode": "json", "retmax": 5})
    url = f"https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi?{params}"
    payload = request_json(url, cache, user_agent=args.user_agent, delay_seconds=args.delay, timeout=args.timeout, retries=args.retries)
    return (((payload or {}).get("esearchresult") or {}).get("idlist")) or []


def query_pubmed_by_doi(doi: str, title: str, journal: str, cache: dict[str, Any], args: argparse.Namespace) -> AbstractCandidate | None:
    if not doi:
        return None
    ids = pubmed_ids_for_query(f"{doi}[AID]", cache, args)
    if not ids:
        return None
    params = urllib.parse.urlencode({"db": "pubmed", "id": ",".join(ids[:3]), "retmode": "xml"})
    url = f"https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi?{params}"
    xml_text = request_text(url, cache, user_agent=args.user_agent, delay_seconds=args.delay, timeout=args.timeout, retries=args.retries, accept="application/xml")
    candidate = pubmed_candidate_from_xml(xml_text or "", "pubmed_doi", title, journal)
    if candidate:
        candidate.score = 1.0
    return candidate


def query_pubmed_by_title(title: str, journal: str, cache: dict[str, Any], args: argparse.Namespace) -> list[AbstractCandidate]:
    if not title:
        return []
    ids = pubmed_ids_for_query(f'"{title}"[Title]', cache, args)
    if not ids:
        return []
    candidates = []
    for pmid in ids[:3]:
        params = urllib.parse.urlencode({"db": "pubmed", "id": pmid, "retmode": "xml"})
        url = f"https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi?{params}"
        xml_text = request_text(url, cache, user_agent=args.user_agent, delay_seconds=args.delay, timeout=args.timeout, retries=args.retries, accept="application/xml")
        candidate = pubmed_candidate_from_xml(xml_text or "", "pubmed_title", title, journal)
        if candidate:
            candidates.append(candidate)
    return candidates


def best_candidate_for_row(row: pd.Series, cache: dict[str, Any], args: argparse.Namespace) -> AbstractCandidate | None:
    title = str(row.get("title", "") or "").strip()
    journal = str(row.get("journal", "") or row.get("source_database", "") or "").strip()
    doi = normalize_doi(row.get("doi", "")) or doi_from_link(row.get("link", ""))
    providers = [provider.strip().lower() for provider in args.providers.split(",") if provider.strip()]

    if doi:
        if "europe_pmc" in providers:
            candidate = query_europe_pmc_by_doi(doi, title, journal, cache, args)
            if candidate and candidate.abstract:
                return candidate
        if "pubmed" in providers:
            candidate = query_pubmed_by_doi(doi, title, journal, cache, args)
            if candidate and candidate.abstract:
                return candidate
        if "openalex" in providers:
            candidate = query_openalex_by_doi(doi, title, journal, cache, args)
            if candidate and candidate.abstract:
                return candidate
        if "semantic_scholar" in providers:
            candidate = query_semantic_by_doi(doi, title, journal, cache, args)
            if candidate and candidate.abstract:
                return candidate
        if "crossref" in providers:
            candidate = query_crossref_by_doi(doi, title, journal, cache, args)
            if candidate and candidate.abstract:
                return candidate

    fallback_candidates: list[AbstractCandidate] = []
    if title:
        if "europe_pmc" in providers:
            candidates = query_europe_pmc_by_title(title, journal, cache, args)
            good = [item for item in candidates if item.score >= args.min_title_score]
            if good:
                return max(good, key=lambda item: (item.score, len(item.abstract)))
            fallback_candidates.extend(candidates)
        if "pubmed" in providers:
            candidates = query_pubmed_by_title(title, journal, cache, args)
            good = [item for item in candidates if item.score >= args.min_title_score]
            if good:
                return max(good, key=lambda item: (item.score, len(item.abstract)))
            fallback_candidates.extend(candidates)
        if "openalex" in providers:
            candidates = query_openalex_by_title(title, journal, cache, args)
            good = [item for item in candidates if item.score >= args.min_title_score]
            if good:
                return max(good, key=lambda item: (item.score, len(item.abstract)))
            fallback_candidates.extend(candidates)
        if "semantic_scholar" in providers:
            candidates = query_semantic_by_title(title, journal, cache, args)
            good = [item for item in candidates if item.score >= args.min_title_score]
            if good:
                return max(good, key=lambda item: (item.score, len(item.abstract)))
            fallback_candidates.extend(candidates)
        if "crossref" in providers:
            candidates = query_crossref_by_title(title, journal, cache, args)
            good = [item for item in candidates if item.score >= args.min_title_score]
            if good:
                return max(good, key=lambda item: (item.score, len(item.abstract)))
            fallback_candidates.extend(candidates)

    if not fallback_candidates:
        return None
    return max(fallback_candidates, key=lambda item: item.score)


def canonicalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    normalized_to_original = {normalize_header(column): column for column in df.columns}
    for canonical, aliases in HEADER_ALIASES.items():
        if canonical in df.columns:
            continue
        for alias in aliases:
            original = normalized_to_original.get(normalize_header(alias))
            if original:
                df[canonical] = df[original]
                break
        if canonical not in df.columns:
            df[canonical] = ""
    return df.fillna("")


def read_input(path: Path, sheet: str | None) -> pd.DataFrame:
    if path.suffix.lower() in {".xlsx", ".xls"}:
        if sheet:
            return pd.read_excel(path, sheet_name=sheet, dtype=str).fillna("")
        workbook = pd.ExcelFile(path)
        sheet_name = "Title_Include_Maybe_Metadata" if "Title_Include_Maybe_Metadata" in workbook.sheet_names else workbook.sheet_names[0]
        return pd.read_excel(path, sheet_name=sheet_name, dtype=str).fillna("")
    return pd.read_csv(path, dtype=str).fillna("")


def add_lookup_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy().fillna("")
    defaults = {
        "abstract_lookup_status": "",
        "abstract_lookup_source": "",
        "abstract_lookup_match_score": "",
        "abstract_lookup_match_confidence": "",
        "abstract_lookup_matched_title": "",
        "abstract_lookup_matched_doi": "",
        "abstract_lookup_matched_year": "",
        "abstract_lookup_matched_journal": "",
        "abstract_lookup_matched_url": "",
        "abstract_lookup_notes": "",
    }
    for column, default in defaults.items():
        if column not in df.columns:
            df[column] = default
    return df


def enrich_abstracts(df: pd.DataFrame, cache: dict[str, Any], args: argparse.Namespace) -> pd.DataFrame:
    df = add_lookup_columns(canonicalize_columns(df))
    total = len(df)
    processed = 0
    found = 0
    already = 0

    for idx, row in df.iterrows():
        existing = str(row.get("abstract", "") or "").strip()
        record_id = str(row.get("record_id", "") or "").strip() or str(idx + 1)
        if existing:
            prior_status = str(row.get("abstract_lookup_status", "") or "").strip()
            if not prior_status:
                df.at[idx, "abstract_lookup_status"] = "already_present"
                df.at[idx, "abstract_lookup_source"] = row.get("metadata_source", "") or "input_file"
                df.at[idx, "abstract_lookup_match_confidence"] = "existing"
                df.at[idx, "abstract_lookup_notes"] = "Abstract was already present in the input workbook."
            already += 1
        else:
            candidate = best_candidate_for_row(row, cache, args)
            if candidate and candidate.abstract and (candidate.source.endswith("_doi") or candidate.score >= args.min_title_score):
                df.at[idx, "abstract"] = candidate.abstract
                df.at[idx, "abstract_lookup_status"] = "found"
                df.at[idx, "abstract_lookup_source"] = candidate.source
                df.at[idx, "abstract_lookup_match_score"] = f"{candidate.score:.3f}"
                df.at[idx, "abstract_lookup_match_confidence"] = "high" if candidate.source.endswith("_doi") else confidence(candidate.score)
                df.at[idx, "abstract_lookup_matched_title"] = candidate.title
                df.at[idx, "abstract_lookup_matched_doi"] = candidate.doi
                df.at[idx, "abstract_lookup_matched_year"] = candidate.year
                df.at[idx, "abstract_lookup_matched_journal"] = candidate.journal
                df.at[idx, "abstract_lookup_matched_url"] = candidate.url
                df.at[idx, "abstract_lookup_notes"] = candidate.notes
                found += 1
            elif candidate and candidate.abstract:
                df.at[idx, "abstract_lookup_status"] = "not_found"
                df.at[idx, "abstract_lookup_source"] = candidate.source
                df.at[idx, "abstract_lookup_match_score"] = f"{candidate.score:.3f}"
                df.at[idx, "abstract_lookup_match_confidence"] = "low"
                df.at[idx, "abstract_lookup_matched_title"] = candidate.title
                df.at[idx, "abstract_lookup_matched_doi"] = candidate.doi
                df.at[idx, "abstract_lookup_matched_year"] = candidate.year
                df.at[idx, "abstract_lookup_matched_journal"] = candidate.journal
                df.at[idx, "abstract_lookup_matched_url"] = candidate.url
                df.at[idx, "abstract_lookup_notes"] = "Candidate found but title match was below threshold; not inserted."
            else:
                df.at[idx, "abstract_lookup_status"] = "not_found"
                df.at[idx, "abstract_lookup_notes"] = "No abstract found in queried public metadata APIs."

        processed += 1
        if processed % args.save_every == 0:
            save_cache(args.cache, cache)
        if processed % args.progress_every == 0 or processed == total:
            missing = int(df["abstract"].astype(str).str.strip().eq("").sum())
            print(f"Processed {processed}/{total} | already {already} | found {found} | still missing {missing}", flush=True)

    return df


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="31572C")
    font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        if ws.max_row and ws.max_column:
            ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for column_cells in ws.columns:
            letter = get_column_letter(column_cells[0].column)
            max_len = max(len(str(cell.value or "")[:100]) for cell in column_cells)
            ws.column_dimensions[letter].width = max(12, min(max_len + 2, 55))

    for sheet_name in ["Records_With_Abstracts", "Still_Missing_Abstracts"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = {cell.value: cell.column for cell in ws[1]}
        for name, width in {
            "title": 58,
            "authors": 42,
            "abstract": 90,
            "abstract_lookup_notes": 55,
            "link": 45,
            "doi": 28,
        }.items():
            if name in headers:
                ws.column_dimensions[get_column_letter(headers[name])].width = width
    wb.save(path)


def write_output(path: Path, df: pd.DataFrame) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    abstract_present = df["abstract"].astype(str).str.strip().ne("")
    missing = ~abstract_present
    lookup_counts = df["abstract_lookup_status"].replace("", "blank").value_counts(dropna=False).rename_axis("abstract_lookup_status").reset_index(name="count")
    source_counts = df.loc[df["abstract_lookup_source"].astype(str).str.strip().ne(""), "abstract_lookup_source"].value_counts(dropna=False).rename_axis("abstract_lookup_source").reset_index(name="count")
    summary = pd.DataFrame(
        [
            {"metric": "Input records", "value": len(df)},
            {"metric": "Records with abstracts after lookup", "value": int(abstract_present.sum())},
            {"metric": "Records still missing abstracts", "value": int(missing.sum())},
            {"metric": "Abstracts already present in input", "value": int((df["abstract_lookup_status"] == "already_present").sum())},
            {"metric": "New abstracts found by script", "value": int((df["abstract_lookup_status"] == "found").sum())},
            {"metric": "Not found by script", "value": int((df["abstract_lookup_status"] == "not_found").sum())},
        ]
    )
    temp_path = temporary_output_path(path)
    with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        lookup_counts.to_excel(writer, sheet_name="Lookup_Status_Counts", index=False)
        source_counts.to_excel(writer, sheet_name="Lookup_Source_Counts", index=False)
        df.to_excel(writer, sheet_name="Records_With_Abstracts", index=False)
        df.loc[missing].to_excel(writer, sheet_name="Still_Missing_Abstracts", index=False)
    style_workbook(temp_path)
    temp_path.replace(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Find missing abstracts for title Include/Maybe records.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="Input xlsx/csv file.")
    parser.add_argument("--sheet", default=None, help="Input sheet name. Defaults to Title_Include_Maybe_Metadata when present.")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Output xlsx file.")
    parser.add_argument("--cache", type=Path, default=DEFAULT_CACHE, help="JSON API cache path.")
    parser.add_argument(
        "--providers",
        default="europe_pmc,pubmed,openalex,semantic_scholar,crossref",
        help="Comma-separated providers: europe_pmc,pubmed,openalex,semantic_scholar,crossref.",
    )
    parser.add_argument("--min-title-score", type=float, default=0.88, help="Minimum title-match score for title-search abstract insertion.")
    parser.add_argument("--delay", type=float, default=0.12, help="Delay between uncached API requests, in seconds.")
    parser.add_argument("--timeout", type=int, default=25, help="API timeout in seconds.")
    parser.add_argument("--retries", type=int, default=2, help="Retries for transient API failures.")
    parser.add_argument("--save-every", type=int, default=50, help="Save API cache every N records.")
    parser.add_argument("--progress-every", type=int, default=25, help="Print progress every N records.")
    parser.add_argument("--limit", type=int, default=0, help="Optional row limit for test runs; 0 means all rows.")
    parser.add_argument(
        "--user-agent",
        default="prisma-abstract-finder/1.0",
        help="User-Agent sent to public scholarly APIs.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if not args.input.exists():
        print(f"Input file not found: {args.input}", file=sys.stderr)
        return 1

    cache = load_cache(args.cache)
    df = read_input(args.input, args.sheet)
    if args.limit and args.limit > 0:
        df = df.head(args.limit).copy()

    print(f"Loaded {len(df)} records from {args.input}", flush=True)
    enriched = enrich_abstracts(df, cache, args)
    save_cache(args.cache, cache)
    write_output(args.output, enriched)

    abstract_count = int(enriched["abstract"].astype(str).str.strip().ne("").sum())
    found_count = int((enriched["abstract_lookup_status"] == "found").sum())
    missing_count = len(enriched) - abstract_count
    print(f"Wrote output: {args.output}", flush=True)
    print(f"Abstracts available after lookup: {abstract_count}/{len(enriched)}", flush=True)
    print(f"New abstracts found: {found_count}", flush=True)
    print(f"Still missing abstracts: {missing_count}", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
