#!/usr/bin/env python3
"""Merge supplemental metadata into a PRISMA Include/Maybe abstract workbook.

This script matches Supplemental rows to the Records_With_Abstracts sheet by DOI,
article link, exact title, then near-title match. It fills missing abstracts and
blank metadata fields while preserving the original PRISMA screening columns.
"""

from __future__ import annotations

import argparse
import difflib
import html
import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
DEFAULT_BASE = ROOT / "output" / "abstract_finding" / "title_include_maybe_with_abstracts.xlsx"
DEFAULT_SUPPLEMENTAL = ROOT / "data" / "manual" / "supplemental_metadata.xlsx"
DEFAULT_OUTPUT = ROOT / "output" / "abstract_finding" / "title_include_maybe_with_supplemental_metadata.xlsx"


def temporary_output_path(output: Path) -> Path:
    if output.suffix:
        return output.with_name(f"{output.stem}.tmp{output.suffix}")
    return output.with_name(f"{output.name}.tmp")


def clean(value: Any) -> str:
    text = html.unescape(str(value or ""))
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def normalize_title(value: Any) -> str:
    text = html.unescape(str(value or "")).lower()
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"[\W_]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_doi(value: Any) -> str:
    doi = str(value or "").strip()
    doi = re.sub(r"^https?://(dx\.)?doi\.org/", "", doi, flags=re.IGNORECASE)
    doi = re.sub(r"^doi:\s*", "", doi, flags=re.IGNORECASE)
    return doi.strip().rstrip(".").lower()


def normalize_link(value: Any) -> str:
    return str(value or "").strip().rstrip("/").lower()


def normalize_header(value: Any) -> str:
    return re.sub(r"\s+", "_", str(value or "").strip().lower())


def read_records(path: Path, sheet: str | None) -> pd.DataFrame:
    if path.suffix.lower() in {".xlsx", ".xls"}:
        if sheet:
            return pd.read_excel(path, sheet_name=sheet, dtype=str).fillna("")
        xl = pd.ExcelFile(path)
        sheet_name = "Records_With_Abstracts" if "Records_With_Abstracts" in xl.sheet_names else xl.sheet_names[0]
        return pd.read_excel(path, sheet_name=sheet_name, dtype=str).fillna("")
    return pd.read_csv(path, dtype=str).fillna("")


def read_supplemental(path: Path, sheet: str | None) -> pd.DataFrame:
    if path.suffix.lower() in {".xlsx", ".xls"}:
        xl = pd.ExcelFile(path)
        sheet_name = sheet or xl.sheet_names[0]
        df = pd.read_excel(path, sheet_name=sheet_name, dtype=str).fillna("")
    else:
        df = pd.read_csv(path, dtype=str).fillna("")
    df.columns = [normalize_header(col) for col in df.columns]
    aliases = {
        "title": ["title", "article_title", "paper_title"],
        "link": ["link", "url", "article_link"],
        "keyword": ["keyword", "keywords", "search_query"],
        "database": ["database", "source_database", "source", "journal"],
        "year": ["year", "year_published", "publication_year"],
        "abstract": ["abstract"],
        "doi": ["doi"],
        "authors": ["authors", "author"],
        "document_type": ["document_type", "documenttype", "type"],
        "language": ["language", "lang"],
    }
    for canonical, names in aliases.items():
        if canonical in df.columns:
            continue
        for name in names:
            if name in df.columns:
                df[canonical] = df[name]
                break
        if canonical not in df.columns:
            df[canonical] = ""
    return df.fillna("")


def build_indexes(book: pd.DataFrame) -> tuple[dict[str, list[int]], dict[str, list[int]], dict[str, list[int]]]:
    by_doi: dict[str, list[int]] = {}
    by_link: dict[str, list[int]] = {}
    by_title: dict[str, list[int]] = {}
    book = book.copy()
    book["_n_doi"] = book["doi"].map(normalize_doi)
    book["_n_link"] = book["link"].map(normalize_link)
    book["_n_title"] = book["title"].map(normalize_title)
    for idx, row in book.iterrows():
        if clean(row.get("abstract", "")):
            if row["_n_doi"]:
                by_doi.setdefault(row["_n_doi"], []).append(idx)
            if row["_n_link"]:
                by_link.setdefault(row["_n_link"], []).append(idx)
            if row["_n_title"]:
                by_title.setdefault(row["_n_title"], []).append(idx)
    return by_doi, by_link, by_title


def choose_best(candidates: list[tuple[str, float, int]], book: pd.DataFrame) -> tuple[str, float, int] | None:
    if not candidates:
        return None
    # Prefer longer abstracts when match quality is tied.
    return max(candidates, key=lambda item: (item[1], len(clean(book.loc[item[2], "abstract"]))))


def match_row(row: pd.Series, book: pd.DataFrame, by_doi: dict[str, list[int]], by_link: dict[str, list[int]], by_title: dict[str, list[int]], min_score: float) -> tuple[str, float, int] | None:
    doi = normalize_doi(row.get("doi", ""))
    link = normalize_link(row.get("link", ""))
    title = normalize_title(row.get("title", ""))

    if doi and doi in by_doi:
        return choose_best([("doi", 1.0, idx) for idx in by_doi[doi]], book)
    if link and link in by_link:
        return choose_best([("link", 1.0, idx) for idx in by_link[link]], book)
    if title and title in by_title:
        return choose_best([("exact_title", 1.0, idx) for idx in by_title[title]], book)
    if title:
        near: list[tuple[str, float, int]] = []
        for book_title, indexes in by_title.items():
            score = difflib.SequenceMatcher(None, title, book_title).ratio()
            if score >= min_score:
                near.extend(("near_title", score, idx) for idx in indexes)
        return choose_best(near, book)
    return None


def ensure_columns(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    df = df.copy()
    for col in columns:
        if col not in df.columns:
            df[col] = ""
    return df.fillna("")


def fill_metadata(base: pd.DataFrame, book: pd.DataFrame, min_score: float) -> tuple[pd.DataFrame, pd.DataFrame]:
    base = ensure_columns(
        base,
        [
            "record_id", "title", "link", "journal", "source_database", "year_published", "doi", "authors",
            "abstract", "document_type", "language", "keywords", "abstract_lookup_status", "abstract_lookup_source",
            "abstract_lookup_match_score", "abstract_lookup_match_confidence", "abstract_lookup_matched_title",
            "abstract_lookup_matched_doi", "abstract_lookup_matched_year", "abstract_lookup_matched_journal",
            "abstract_lookup_matched_url", "abstract_lookup_notes",
        ],
    )
    by_doi, by_link, by_title = build_indexes(book)
    audit_rows: list[dict[str, Any]] = []

    for idx, row in base.iterrows():
        missing_abstract_before = not clean(row.get("abstract", ""))
        match = match_row(row, book, by_doi, by_link, by_title, min_score)
        if not match:
            continue
        match_type, score, book_idx = match
        b = book.loc[book_idx]
        book_abstract = clean(b.get("abstract", ""))
        if not book_abstract:
            continue

        fields_filled: list[str] = []
        if missing_abstract_before:
            base.at[idx, "abstract"] = book_abstract
            base.at[idx, "abstract_lookup_status"] = "found_from_supplemental"
            base.at[idx, "abstract_lookup_source"] = "Supplemental"
            base.at[idx, "abstract_lookup_match_score"] = f"{score:.3f}"
            base.at[idx, "abstract_lookup_match_confidence"] = "high" if score >= 0.98 else "medium"
            base.at[idx, "abstract_lookup_matched_title"] = clean(b.get("title", ""))
            base.at[idx, "abstract_lookup_matched_doi"] = clean(b.get("doi", ""))
            base.at[idx, "abstract_lookup_matched_year"] = clean(b.get("year", ""))
            base.at[idx, "abstract_lookup_matched_journal"] = clean(b.get("database", ""))
            base.at[idx, "abstract_lookup_matched_url"] = clean(b.get("link", ""))
            base.at[idx, "abstract_lookup_notes"] = f"Filled from Supplemental by {match_type} match. Original title preserved."
            fields_filled.append("abstract")

        fill_map = {
            "link": "link",
            "keywords": "keyword",
            "source_database": "database",
            "journal": "database",
            "year_published": "year",
            "doi": "doi",
            "authors": "authors",
            "document_type": "document_type",
            "language": "language",
        }
        for target, source in fill_map.items():
            if not clean(base.at[idx, target]) and clean(b.get(source, "")):
                base.at[idx, target] = clean(b.get(source, ""))
                fields_filled.append(target)

        if fields_filled:
            audit_rows.append(
                {
                    "record_id": base.at[idx, "record_id"],
                    "base_title": base.at[idx, "title"],
                    "supplemental_title": clean(b.get("title", "")),
                    "match_type": match_type,
                    "match_score": f"{score:.3f}",
                    "base_doi": row.get("doi", ""),
                    "supplemental_doi": clean(b.get("doi", "")),
                    "base_link": row.get("link", ""),
                    "supplemental_link": clean(b.get("link", "")),
                    "filled_fields": "; ".join(fields_filled),
                    "abstract_filled": "yes" if "abstract" in fields_filled else "no",
                }
            )

    return base, pd.DataFrame(audit_rows)


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="264653")
    font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        if ws.max_row and ws.max_column:
            ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for col_cells in ws.columns:
            letter = get_column_letter(col_cells[0].column)
            max_len = max(len(str(cell.value or "")[:90]) for cell in col_cells)
            ws.column_dimensions[letter].width = max(12, min(max_len + 2, 55))
        headers = {cell.value: cell.column for cell in ws[1]}
        for name, width in {"title": 58, "base_title": 58, "supplemental_title": 58, "abstract": 90, "authors": 42, "link": 45, "abstract_lookup_notes": 55}.items():
            if name in headers:
                ws.column_dimensions[get_column_letter(headers[name])].width = width
    wb.save(path)


def write_workbook(output: Path, df: pd.DataFrame, audit: pd.DataFrame, input_name: str, supplemental_name: str) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    abstract_mask = df["abstract"].astype(str).str.strip().ne("")
    missing_mask = ~abstract_mask
    status_counts = df["abstract_lookup_status"].replace("", "blank").value_counts(dropna=False).rename_axis("abstract_lookup_status").reset_index(name="count")
    match_counts = audit["match_type"].value_counts(dropna=False).rename_axis("supplemental_match_type").reset_index(name="count") if not audit.empty else pd.DataFrame(columns=["supplemental_match_type", "count"])
    summary = pd.DataFrame(
        [
            {"metric": "Base workbook", "value": input_name},
            {"metric": "Supplemental metadata workbook", "value": supplemental_name},
            {"metric": "Total records", "value": len(df)},
            {"metric": "Records with abstracts after Supplemental merge", "value": int(abstract_mask.sum())},
            {"metric": "Records still missing abstracts", "value": int(missing_mask.sum())},
            {"metric": "Rows matched to Supplemental with fields filled", "value": len(audit)},
            {"metric": "Missing abstracts filled from Supplemental", "value": int((audit.get("abstract_filled", pd.Series(dtype=str)) == "yes").sum()) if not audit.empty else 0},
        ]
    )
    temp_output = temporary_output_path(output)
    with pd.ExcelWriter(temp_output, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        match_counts.to_excel(writer, sheet_name="Supplemental_Match_Counts", index=False)
        status_counts.to_excel(writer, sheet_name="Lookup_Status_Counts", index=False)
        df.to_excel(writer, sheet_name="Records_With_Abstracts", index=False)
        df.loc[abstract_mask].to_excel(writer, sheet_name="Papers_With_Abstracts", index=False)
        df.loc[missing_mask].to_excel(writer, sheet_name="Still_Missing_Abstracts", index=False)
        audit.to_excel(writer, sheet_name="Supplemental_Match_Audit", index=False)
    style_workbook(temp_output)
    temp_output.replace(output)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Merge Supplemental metadata into an abstract-finding workbook.")
    parser.add_argument("--input", type=Path, default=DEFAULT_BASE)
    parser.add_argument("--input-sheet", default="Records_With_Abstracts")
    parser.add_argument("--supplemental", type=Path, default=DEFAULT_SUPPLEMENTAL)
    parser.add_argument("--supplemental-sheet", default=None)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--min-near-title-score", type=float, default=0.94)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if not args.input.exists():
        raise SystemExit(f"Input workbook not found: {args.input}")
    if not args.supplemental.exists():
        raise SystemExit(f"Supplemental workbook not found: {args.supplemental}")
    base = read_records(args.input, args.input_sheet)
    book = read_supplemental(args.supplemental, args.supplemental_sheet)
    before = int(base["abstract"].astype(str).str.strip().ne("").sum()) if "abstract" in base.columns else 0
    merged, audit = fill_metadata(base, book, args.min_near_title_score)
    after = int(merged["abstract"].astype(str).str.strip().ne("").sum())
    write_workbook(args.output, merged, audit, args.input.name, args.supplemental.name)
    print(f"Input records: {len(merged)}")
    print(f"Abstracts before Supplemental merge: {before}")
    print(f"Abstracts after Supplemental merge: {after}")
    print(f"New abstracts filled from Supplemental: {after - before}")
    print(f"Still missing abstracts: {len(merged) - after}")
    print(f"Rows with Supplemental fields filled: {len(audit)}")
    print(f"Wrote: {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
