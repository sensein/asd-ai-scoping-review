"""Build the ICR reliability dataset and Krippendorff alpha outputs.

The pipeline imports the same shared codebook and parsing helpers used by the
Results scripts. Source-script labels are preserved in the output codebook and
run log for traceability.
"""

from __future__ import annotations

import hashlib
import math
import os
import platform
import random
import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter

from codebook import (
    ALGORITHM_FAMILY_PATTERNS,
    HYBRID_MODEL_PATTERN,
    LEARNING_TYPE_PATTERNS,
    NOT_GIVEN_TASK_PATTERN,
    TASK_TYPE_PATTERNS,
)
from helper_functions_ import (
    evaluation_metric_categories as shared_evaluation_metric_categories,
    extract_accuracy_percent as shared_extract_accuracy_percent,
    is_invalid as shared_is_invalid,
    normalize_text as shared_normalize_text,
    parse_numeric_age_value,
    yes_no_nominal as shared_yes_no_nominal,
)
from reliability import is_missing, krippendorff_alpha


RANDOM_SEED = 20260707
BOOTSTRAP_ITERATIONS = 300

SCRIPT_DIR = Path(__file__).resolve().parent

# Repository root:
# asd-ai-scoping-review/
REPO_ROOT = SCRIPT_DIR.parent
DATA_ROOT = Path(os.environ.get("ASD_REVIEW_DATA_ROOT", REPO_ROOT / "data")).expanduser()
if not DATA_ROOT.is_absolute():
    DATA_ROOT = REPO_ROOT / DATA_ROOT
OUTPUT_ROOT = Path(os.environ.get("ASD_REVIEW_OUTPUT_ROOT", REPO_ROOT / "output")).expanduser()
if not OUTPUT_ROOT.is_absolute():
    OUTPUT_ROOT = REPO_ROOT / OUTPUT_ROOT

DEFAULT_WORKBOOK = DATA_ROOT / "ICR.xlsx"

TYPE_CLASSIFICATION = DATA_ROOT / "ICR_variable_type_classification.xlsx"

# All ICR outputs will be saved here:
OUTPUT_DIR = OUTPUT_ROOT / "icr_results"

SOURCE_SCRIPT_DIR = SCRIPT_DIR
# The Results scripts are in the same folder as this script:
# asd-ai-scoping-review/scripts/

SOURCE_SCRIPT_MANIFEST = [
    ("setup_data_.py", SOURCE_SCRIPT_DIR / "setup_data_.py"),
    (
        "helper_functions_.py",
        SOURCE_SCRIPT_DIR / "helper_functions_.py",
    ),
    ("rq1_.py", SOURCE_SCRIPT_DIR / "rq1_.py"),
    ("rq2_.py", SOURCE_SCRIPT_DIR / "rq2_.py"),
    ("rq3_.py", SOURCE_SCRIPT_DIR / "rq3_.py"),
    ("rq4_.py", SOURCE_SCRIPT_DIR / "rq4_.py"),
    ("rq5_.py", SOURCE_SCRIPT_DIR / "rq5_.py"),
    ("codebook.py", SOURCE_SCRIPT_DIR / "codebook.py"),
    ("reliability.py", SOURCE_SCRIPT_DIR / "reliability.py"),
]

SOURCE_SCRIPTS = [
    path for _, path in SOURCE_SCRIPT_MANIFEST
]


def repo_relative_path(path: Path) -> str:
    try:
        return str(path.resolve().relative_to(REPO_ROOT.resolve()))
    except Exception:
        return path.name


def extract_coder_override_from_note(note: str) -> str:
    match = re.search(r"novelty.*sensitivity.*done by\s+([A-Za-z][A-Za-z0-9_.-]*)", note, flags=re.IGNORECASE)
    return match.group(1).strip() if match else ""


def anonymize_coder_id(raw_coder: Any, coder_label_map: dict[str, str]) -> str:
    raw = display_raw(raw_coder).strip()
    if not raw:
        return ""
    key = normalize_text(raw)
    if key not in coder_label_map:
        coder_label_map[key] = f"Coder{len(coder_label_map) + 1}"
    return coder_label_map[key]


def anonymize_note_text(note: str, coder_label_map: dict[str, str]) -> str:
    override = extract_coder_override_from_note(note)
    if not override:
        return note
    return re.sub(re.escape(override), anonymize_coder_id(override, coder_label_map), note, count=1)


INVALID_TEXT_VALUES = {
    "",
    "-",
    "--",
    "---",
    "nan",
    "none",
    "no information",
    "na",
    "n/a",
    "n.a",
    "n.a.",
    "n.d",
    "n.d.",
    "nd",
    "n/d",
    "nr",
    "n.r",
    "n.r.",
    "n/r",
    "not given",
    "not givven",
    "not reported",
    "not specified",
    "not stated",
    "not mentioned",
    "not provided",
    "not available",
    "not applicable",
    "not explicitly stated",
    "unknown",
    "unkown",
    "unclear",
    "not clear",
    "none specified",
}

NEGATIVE_TEXT_VALUES = {"no", "n", "false", "0", "absent", "not included", "not used"}

AGE_CATEGORIES = {
    "Infants": (0.0, 1.0),
    "Toddlers": (1.0, 3.0),
    "Pre-schoolers": (3.0, 6.0),
    "Grade-schoolers": (6.0, 12.0),
    "Teens": (12.0, 18.0),
    "Adults": (18.0, float("inf")),
}


PATTERN_CLASSICAL_ML = ALGORITHM_FAMILY_PATTERNS["classical_machine_learning_models"]
PATTERN_ENSEMBLE = ALGORITHM_FAMILY_PATTERNS["ensemble_models"]
PATTERN_NEURAL = ALGORITHM_FAMILY_PATTERNS["neural_network_models"]
PATTERN_STATISTICAL_SPECIALISED = ALGORITHM_FAMILY_PATTERNS["statistical_and_other_specialised_models"]
PATTERN_HYBRID_MODEL = HYBRID_MODEL_PATTERN


@dataclass(frozen=True)
class VariablePair:
    domain: str
    subgroup: str
    variable_label: str
    variable_name: str
    reviewer_col_idx: int
    final_col_idx: int
    reviewer_excel_col: str
    final_excel_col: str
    reviewer_block: str
    final_block: str
    position_in_block: int
    processor: str
    variable_type: str
    measurement_level: str
    source_script: str
    results_function: str
    is_exception: bool = False
    is_new_icr_rule: bool = False


def normalize_text(value: Any) -> str:
    return shared_normalize_text(value)


def display_raw(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value)


def is_invalid(value: Any) -> bool:
    return shared_is_invalid(value)


def is_negative_marker(value: Any) -> bool:
    text = normalize_text(value)
    return text in INVALID_TEXT_VALUES or text in NEGATIVE_TEXT_VALUES


def slugify(value: str) -> str:
    text = normalize_text(value)
    text = text.replace("#", "number")
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return re.sub(r"_+", "_", text).strip("_") or "blank"


def source_hash(path: Path) -> str:
    if not path.exists():
        return "missing"
    return hashlib.sha256(path.read_bytes()).hexdigest()[:16]


def regex_categories(text: str, patterns: dict[str, str]) -> dict[str, int]:
    return {category: int(bool(re.search(pattern, text, flags=re.IGNORECASE))) for category, pattern in patterns.items()}


def add_unmatched_if_needed(
    unmatched: list[dict[str, Any]],
    paper_id: str,
    variable_name: str,
    rating_role: str,
    coder_id: str,
    raw_value: Any,
    normalized_value: str,
    category_family: str,
    matched_categories: dict[str, int],
    substantive_missing_ok: bool = False,
) -> None:
    if is_invalid(raw_value) or (substantive_missing_ok and normalize_text(raw_value) in {"no", "none"}):
        return
    review_only_markers = (
        "other_uncategorized",
        "manual_review",
        "need_manual_revision",
        "no_category_matched",
        "unclear",
    )
    meaningful_matches = [
        key
        for key, value in matched_categories.items()
        if value == 1 and not any(marker in key for marker in review_only_markers)
    ]
    if meaningful_matches:
        return
    if normalize_text(raw_value):
        unmatched.append(
            {
                "paper_id": paper_id,
                "variable_name": variable_name,
                "rating_role": rating_role,
                "coder_id": coder_id,
                "raw_value": display_raw(raw_value),
                "normalized_value": normalized_value,
                "suggested_category": "",
                "manual_review_decision": "",
                "final_category": "",
                "category_family": category_family,
            }
        )


def parse_numeric_value(value: Any, percent: bool = False) -> float:
    if value is None:
        return np.nan
    if isinstance(value, datetime):
        return np.nan
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        val = float(value)
        if percent and 0 < val <= 1:
            return val * 100
        return val
    text = normalize_text(value)
    if text in INVALID_TEXT_VALUES or text in NEGATIVE_TEXT_VALUES:
        return np.nan
    text = text.replace(",", "")
    nums = re.findall(r"-?\d+(?:\.\d+)?", text)
    if not nums:
        return np.nan
    val = float(nums[0])
    if percent and 0 < val <= 1:
        val *= 100
    return val


def parse_count_value(value: Any) -> float:
    if value is None or isinstance(value, datetime):
        return np.nan
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = normalize_text(value)
    if text in INVALID_TEXT_VALUES or text in NEGATIVE_TEXT_VALUES:
        return np.nan
    nums = [float(x) for x in re.findall(r"\d+(?:\.\d+)?", text)]
    if not nums:
        return np.nan
    if re.search(r"\+| and ", text) and len(nums) > 1:
        return float(sum(nums))
    return nums[0]


def quality_issue_indicator(value: Any) -> tuple[int, str]:
    text = normalize_text(value)
    if text in {"", "-", "--", "no", "n", "false", "0", "none", "n/a", "na", "n.d", "nd", "n/d"}:
        return 0, "no_quality_issue_recorded"
    if re.fullmatch(r"no[\s.,;:]*", text or ""):
        return 0, "no_quality_issue_recorded"
    return 1, "quality_issue_recorded"


def yes_no_nominal(value: Any, blank_as_no: bool = False) -> str:
    return shared_yes_no_nominal(value, blank_as_no=blank_as_no)


def binary_from_nominal(category: str) -> dict[str, int]:
    return {category: 1}


def date_as_age_range(value: Any) -> str:
    if isinstance(value, datetime):
        return f"{value.month}-{value.day}"
    return display_raw(value)


def extract_age_numbers(text: Any) -> list[float]:
    return [float(n) for n in re.findall(r"\d+\.?\d*", str(text))]


def contains_months(text: Any) -> bool:
    return "month" in normalize_text(text)


def convert_if_months(age_min: float, age_max: float, text: Any) -> tuple[float, float]:
    if contains_months(text):
        return age_min / 12.0, age_max / 12.0
    return age_min, age_max


def is_reasonable_age_range(age_min: float, age_max: float) -> bool:
    if pd.isna(age_min) or pd.isna(age_max):
        return False
    if age_min < 0 or age_max < 0:
        return False
    if age_max > 120:
        return False
    return True


def parse_multiple_age_ranges(value: Any) -> list[tuple[float, float]]:
    raw = date_as_age_range(value)
    text = normalize_text(raw)
    dash_matches = re.findall(r"(\d+\.?\d*)\s*[-/]\s*(\d+\.?\d*)", text)
    to_matches = re.findall(r"(\d+\.?\d*)\s+to\s+(\d+\.?\d*)", text)
    ranges: list[tuple[float, float]] = []
    for start, end in dash_matches + to_matches:
        age_min = float(start)
        age_max = float(end)
        age_min, age_max = convert_if_months(age_min, age_max, text)
        age_min, age_max = min(age_min, age_max), max(age_min, age_max)
        if is_reasonable_age_range(age_min, age_max):
            ranges.append((age_min, age_max))
    return ranges


def parse_under_over_age(value: Any) -> list[tuple[float, float]]:
    text = normalize_text(date_as_age_range(value))
    nums = extract_age_numbers(text)
    if len(nums) != 1:
        return []
    age = nums[0] / 12.0 if contains_months(text) else nums[0]
    if "under" in text or "below" in text:
        return [(0.0, age)] if is_reasonable_age_range(0.0, age) else []
    if "over" in text or "above" in text:
        return [(age, float("inf"))] if age <= 120 else []
    return []


def parse_mean_sd_age_from_text(value: Any) -> list[tuple[float, float]]:
    text = normalize_text(date_as_age_range(value))
    if "±" not in text and "+/-" not in text:
        return []
    nums = extract_age_numbers(text)
    if len(nums) < 2:
        return []
    mean, sd = nums[0], nums[1]
    age_min = max(mean - sd, 0.0)
    age_max = mean + sd
    age_min, age_max = convert_if_months(age_min, age_max, text)
    return [(age_min, age_max)] if is_reasonable_age_range(age_min, age_max) else []


def parse_regular_age(value: Any) -> list[tuple[float, float]]:
    text = normalize_text(date_as_age_range(value))
    if text in INVALID_TEXT_VALUES:
        return []
    nums = extract_age_numbers(text)
    if not nums:
        return []
    if len(nums) == 1:
        age_min = age_max = nums[0]
    else:
        age_min, age_max = min(nums), max(nums)
    age_min, age_max = convert_if_months(age_min, age_max, text)
    return [(age_min, age_max)] if is_reasonable_age_range(age_min, age_max) else []


def parse_age_range_cell(value: Any) -> tuple[list[tuple[float, float]], str]:
    text = normalize_text(date_as_age_range(value))
    if text in INVALID_TEXT_VALUES:
        return [], "invalid"
    if any(word in text for word in ["under", "below", "over", "above"]):
        ranges = parse_under_over_age(value)
        if ranges:
            return ranges, "under_over"
    if text.count("-") >= 1 or "/" in text or re.search(r"\bto\b", text):
        ranges = parse_multiple_age_ranges(value)
        if ranges:
            return ranges, "range_field"
    ranges = parse_mean_sd_age_from_text(value)
    if ranges:
        return ranges, "mean_sd_text_in_range_field"
    ranges = parse_regular_age(value)
    if ranges:
        return ranges, "single_or_regular_age"
    return [], "unparsed"


def categories_for_age_ranges(ranges: list[tuple[float, float]]) -> list[str]:
    matched: set[str] = set()
    for age_min, age_max in ranges:
        for category, (low, high) in AGE_CATEGORIES.items():
            if high == float("inf"):
                if age_max >= low:
                    matched.add(category)
            elif age_max >= low and age_min < high:
                matched.add(category)
    return [cat for cat in AGE_CATEGORIES if cat in matched]


def age_range_categories(value: Any) -> tuple[dict[str, int], str, str]:
    ranges, parser = parse_age_range_cell(value)
    cats = {cat: 0 for cat in AGE_CATEGORIES}
    cats["age_range_not_given"] = 0
    cats["multiple_age_groups"] = 0
    matched = categories_for_age_ranges(ranges) if ranges else []
    if not matched:
        cats["age_range_not_given"] = 1
    else:
        for cat in matched:
            cats[cat] = 1
        cats["multiple_age_groups"] = int(len(matched) >= 2)
    norm = "; ".join(matched) if matched else "age_range_not_given"
    return cats, norm, parser


def diagnostic_method_categories(value: Any) -> dict[str, int]:
    text = normalize_text(value)
    patterns = {
        "dsm": r"\bdsm\b|diagnostic and statistical manual",
        "icd": r"\bicd\b",
        "ados": r"\bados\b|autism diagnostic observation schedule",
        "adi_r": r"\badi[- ]?r\b|autism diagnostic interview",
        "cars": r"\bcars\b|childhood autism rating scale",
        "clinical_diagnosis": r"clinical diagnos|clinically diagnosed|expert diagnos",
        "k_sads": r"k[- ]?sads|schedule for affective disorders",
        "iq_or_developmental_assessment": r"iq\b|wechsler|bayley|gesell|intelligence|adaptive behavior|cognitive",
        "other_assessment": r"assessment|scale|questionnaire|screening|test|interview",
    }
    cats = regex_categories(text, patterns)
    cats["not_reported"] = int(is_invalid(value))
    if cats["not_reported"]:
        for key in list(patterns):
            cats[key] = 0
    return cats


def terminology_categories(value: Any, subgroup: str) -> dict[str, int]:
    text = normalize_text(value)
    if subgroup == "asd":
        patterns = {
            "official_diagnostic_terminology": r"\basd\b|autism spectrum disorder|autism spectrum condition",
            "identity_first_language": r"\bautistic\b|autistics",
            "person_first_language": r"with asd|with autism|individuals? with|participants? with|children with|child with",
            "generic_autism_language": r"\bautism\b|autistic disorder",
        }
    elif subgroup == "neurotypical":
        patterns = {
            "typically_developing": r"typically develop|td group|td children|td individuals",
            "neurotypical": r"neurotypical|\bnt\b",
            "normal_or_healthy_controls": r"normal subject|healthy control|control group|controls",
            "non_autistic": r"non[- ]?asd|non[- ]?autistic",
        }
    else:
        patterns = {
            "adhd": r"\badhd\b|attention deficit",
            "cerebral_palsy": r"cerebral palsy|\bcp\b",
            "fragile_x": r"fragile x",
            "schizophrenia": r"schizophrenia|\bsz\b",
            "developmental_or_intellectual_disability": r"developmental|intellectual disability|mental retardation",
            "other_specific_diagnosis": r"diagnos|disorder|clinical|syndrome|condition",
        }
    cats = regex_categories(text, patterns)
    cats["not_reported"] = int(is_invalid(value))
    if not cats["not_reported"] and not any(cats[k] for k in patterns):
        cats["other_uncategorized_terminology"] = 1
    else:
        cats["other_uncategorized_terminology"] = 0
    return cats


def study_goal_categories(value: Any) -> dict[str, int]:
    text = normalize_text(value)
    patterns = {
        "prediction_of_outcome": r"\bprediction\b|\bpredict\w*\b|\bforecast\w*\b|\boutcome prediction\b",
        "screening_detection": r"\bscreen\w*\b|\brecognition\b|\bdetection\b|\bdetect\w*\b|\bidentification\b|\bidentify\w*\b",
        "severity_detection": r"\bseverity\b|\bsevere\b|\bseverely\b|\bsymptom severity\b",
        "classification": r"\bclassif\w*\b|\bclassifier\b|\bdistinguish\w*\b|\bdifferentiat\w*\b",
        "diagnosis": r"\bdiagnos\w*\b|\bdiagnostic\b",
        "identifying_symptoms_biomarkers": r"\bindicator\w*\b|\bsymptom\w*\b|\binvestigat\w*\b|\bbiomarker\w*\b|\bmarker\w*\b",
        "other_goal": r"\battention\b|\bstratification\b|\bintervention\b|\btreatment\b|\btherapy\b|\bfeasibility\b",
    }
    cats = regex_categories(text, patterns)
    cats["unclear"] = int(not is_invalid(value) and not any(cats.values()))
    cats["not_reported"] = int(is_invalid(value))
    return cats


def study_setting_nominal(value: Any) -> str:
    text = normalize_text(value)
    if text in INVALID_TEXT_VALUES or text == "no":
        return "not_reported"
    controlled = bool(
        re.search(
            r"\bclinic\w*\b|\bclinical\b|\bcontrolled\b|\blab\w*\b|\blaboratory\b|\bhospital\w*\b|"
            r"\bschool\w*\b|\bcenter\w*\b|\bcentre\w*\b|\buniversity\b|\bresearch facility\b|\bvr\b|virtual reality",
            text,
        )
    )
    uncontrolled = bool(
        re.search(r"\bremote\b|\buncontrolled\b|\bonline\b|\bhome\w*\b|\bnaturalistic\b|in[- ]?the[- ]?wild|real[- ]?world", text)
    )
    explicit_both = bool(re.search(r"\bboth\b|clinic.*home|home.*clinic|lab.*home|home.*lab|controlled.*uncontrolled", text))
    if explicit_both or (controlled and uncontrolled):
        return "both_controlled_and_uncontrolled"
    if controlled:
        return "controlled_setting"
    if uncontrolled:
        return "uncontrolled_naturalistic_remote"
    return "unclear"


def task_type_categories(value: Any) -> dict[str, int]:
    text = normalize_text(value)
    cats = regex_categories(text, TASK_TYPE_PATTERNS)
    matched = sum(cats.values())
    cats["not_given"] = int(bool(re.search(NOT_GIVEN_TASK_PATTERN, text)))
    cats["multiple_task_types"] = int(matched >= 2)
    cats["unclear"] = int(not cats["not_given"] and matched == 0)
    return cats


def modality_nominal(value: Any) -> str:
    return yes_no_nominal(value, blank_as_no=False)


def other_behavioral_categories(value: Any) -> dict[str, int]:
    text = normalize_text(value)
    patterns = {
        "facial_expression_emotion_recognition": r"\bface\b|\bfaces\b|\bfacial\b|\bfacial expression(?:s)?\b|\bemotion(?:s)?\b|\bemotion recognition\b",
        "nonverbal_other_speech_language": r"\btranscript\b|\btext\b|\btweet\b|questionnaire|echolalia|audio-visual|audiovisual",
        "social_interaction": r"interaction|interact|gesture|social|vocali[sz]ation",
        "joint_attention": r"joint attention|jointattention",
        "video_analysis_data": r"video|video frame|audio-visual|audiovisual",
        "decision_making": r"decision making|decision-making",
        "sensor_data": r"sensor|inertial sensor|wearable",
        "other_movement_data": r"inertial|kinematic|grasp|\bpose\b|angle|fine[- ]motor|motor abnormalities|gait|posture",
        "other_gaze_data": r"eye[- ]gaze|scan[- ]?path|saccade|fixation|eye[- ]tracking|gaze pattern",
        "eeg": r"\beeg\b|electroencephalograph",
    }
    cats = regex_categories(text, patterns)
    cats["not_reported"] = int(is_invalid(value) or text == "no")
    if not cats["not_reported"] and not any(cats[k] for k in patterns):
        cats["other_uncategorized_behavioral_data"] = 1
    else:
        cats["other_uncategorized_behavioral_data"] = 0
    return cats


def feature_fusion_categories(value: Any) -> dict[str, int]:
    text = normalize_text(value)
    no_fusion = bool(re.fullmatch(r"\s*|[-\s]+|not specified|not reported|unclear|unknown|na|n/a|nd|n/d|no|none|nan", text))
    patterns = {
        "concatenating_features": r"concat(?:enate|enated|enation|)|feature concat",
        "early_fusion": r"early fusion|early-fusion|input[- ]level fusion|feature[- ]level fusion",
        "late_fusion": r"late fusion|late-fusion|decision[- ]level fusion|score[- ]level fusion",
        "hybrid_fusion": r"hybrid fusion|\bhybrid\b|ensemble fusion",
        "multimodal_fusion": r"multimodal fusion|multi-modal fusion|cross[- ]modal|modality fusion",
        "feature_graph": r"feature graph|feature-graph|graph fusion",
        "feature_selection_before_combination": r"recursive feature elimination|\brfe\b|\brecursive\b|feature selection",
        "weighted_or_rule_based_combination": r"weighted|weighting|rule[- ]based|voting|majority vote",
    }
    cats = {category: 0 for category in patterns}
    if not no_fusion:
        cats.update(regex_categories(text, patterns))
    cats["used_fusion_or_feature_combination"] = int(not no_fusion and any(cats[k] for k in patterns))
    cats["no_fusion_or_not_reported"] = int(no_fusion)
    cats["non_placeholder_needs_manual_review"] = int(not no_fusion and not any(cats[k] for k in patterns))
    return cats


def learning_type_categories(value: Any) -> dict[str, int]:
    text = normalize_text(value)
    cats = regex_categories(text, LEARNING_TYPE_PATTERNS)
    cats["not_reported"] = int(is_invalid(value))
    cats["unclear"] = int(not cats["not_reported"] and not any(cats[k] for k in LEARNING_TYPE_PATTERNS))
    return cats


def features_broad_categories(value: Any) -> dict[str, int]:
    text = normalize_text(value)
    patterns = {
        "gaze_eye_tracking_features": r"\baoi\b|\broi\b|area of interest|region of interest|fixation|saccade|blink|eye[- ]?movement|eye[- ]?tracking|scanpath|scan[- ]?path|gaze|visual attention|visit count|revisit|heatmap|saliency",
        "facial_expression_face_features": r"facial landmark|face landmark|openface|facial expression|\bface\b|facial|action unit|\bau\d+\b|smile|mouth|eyes|eyebrow|lip|emotion recognition|facial dynamics",
        "motor_pose_kinematic_features": r"speed|acceleration|velocity|duration|movement|motion|amplitude|deceleration|distance|displacement|openpose|head pose|head movement|rotation|joint movement|skeleton|keypoint|\bpose\b|grip force|sway|jerk|kinematic|gait|stride|walking|gesture|tablet|touch|wheel rotation|rmse",
        "social_interaction_behavioral_features": r"reaction|latency|response latency|eye contact|social engagement|human behavior coding|observation coding|imitation|social influence|response bias|correctness of response|turn[- ]?taking|joint attention|interaction",
        "language_speech_acoustic_features": r"\bword\b|word count|tf[- ]?idf|word2vec|wav2vec|bert|transformer|nlp|natural language|sentence embedding|\btext embedding\b|tweet|questionnaire|q[- ]?chat|audio|raw audio|spectrogram|prosody|pitch|voice|vocalization|vocalisation|mfcc|speech rhythm|acoustic",
        "vector_or_embedding_features": r"presence vector|weighted presence vector|feature vector|behavioral vector|behavioural vector|embedding|embeddings|latent representation|latent vector|vector\b|vectors\b",
        "image_video_visual_features": r"raw video|\bvideo\b|\bvideos\b|\bimage\b|\bimages\b|frame|frames|rgb|optical flow|visual features|video analysis|image analysis|heatmap|scanpath image|facial image",
        "demographic_developmental_background_features": r"\bage\b|\bsex\b|\bgender\b|developmental history|age of walking|age of first words|pregnancy|delivery|premature|family history|parental|vaccination|sensory|adaptive behavior|iq\b",
    }
    cats = regex_categories(text, patterns)
    cats["not_reported"] = int(is_invalid(value))
    cats["other_uncategorized_features"] = int(not cats["not_reported"] and not any(cats[k] for k in patterns))
    return cats


def algorithm_family_categories(value: Any) -> dict[str, int]:
    text = normalize_text(value)
    cats = {
        "classical_machine_learning_models": int(bool(re.search(PATTERN_CLASSICAL_ML, text))),
        "ensemble_models": int(bool(re.search(PATTERN_ENSEMBLE, text))),
        "neural_network_models": int(bool(re.search(PATTERN_NEURAL, text))),
        "statistical_and_other_specialised_models": int(bool(re.search(PATTERN_STATISTICAL_SPECIALISED, text))),
        "hybrid_or_multi_model_architectures": int(bool(re.search(PATTERN_HYBRID_MODEL, text))),
    }
    cats["multi_model_papers"] = int(sum(cats[k] for k in [
        "classical_machine_learning_models",
        "ensemble_models",
        "neural_network_models",
        "statistical_and_other_specialised_models",
    ]) > 1)
    cats["not_given"] = int(is_invalid(value))
    cats["other_uncategorized_algorithm"] = int(not cats["not_given"] and not any(v for k, v in cats.items() if k not in {"not_given", "other_uncategorized_algorithm"}))
    return cats


def algorithm_group_nominal(value: Any) -> str:
    cats = algorithm_family_categories(value)
    if cats["not_given"]:
        return "not_given"
    if cats["hybrid_or_multi_model_architectures"]:
        return "hybrid model"
    families = []
    if cats["classical_machine_learning_models"]:
        families.append("classical machine learning")
    if cats["ensemble_models"]:
        families.append("ensemble model")
    if cats["neural_network_models"]:
        families.append("neural network")
    if cats["statistical_and_other_specialised_models"]:
        families.append("statistical / specialised model")
    if len(families) > 1:
        return "multiple model families"
    if len(families) == 1:
        return families[0]
    return "unclear"


def exact_model_name(value: Any) -> str:
    text = normalize_text(value)
    if text in INVALID_TEXT_VALUES:
        return "not_given"
    replacements = {
        r"\bsvm\b|support vector machine": "support_vector_machine",
        r"\brf\b|random forest": "random_forest",
        r"\bknn\b|k[- ]?nearest": "knn",
        r"naive|naïve": "naive_bayes",
        r"\bcnn\b|convolutional": "cnn",
        r"\blstm\b": "lstm",
        r"cnn[-+ ]?lstm": "cnn_lstm",
        r"xgboost": "xgboost",
        r"lightgbm|light gbm": "lightgbm",
        r"adaboost|ada boost": "adaboost",
        r"decision tree|treebagger": "decision_tree",
        r"\bmlp\b|multi[- ]?layer perceptron": "mlp",
        r"\bhmm\b|hidden markov": "hidden_markov_model",
        r"pomdp": "pomdp",
        r"mobilenet": "mobilenet",
        r"resnet": "resnet",
        r"vgg": "vgg",
        r"fine[- ]?tuning": "fine_tuning",
        r"ensemble": "ensemble_model",
    }
    for pattern, name in replacements.items():
        if re.search(pattern, text):
            return name
    return slugify(text)[:80] or "unclear"


def evaluation_metric_categories(value: Any) -> dict[str, int]:
    return shared_evaluation_metric_categories(value)


def extract_accuracy_percent(value: Any, evaluation_metrics_value: Any = None) -> float:
    return shared_extract_accuracy_percent(value, evaluation_metrics_value)


def feature_importance_method_categories(value: Any) -> dict[str, int]:
    text = normalize_text(value)
    patterns = {
        "statistical_feature_evaluation": r"t[- ]?test|ttest|anova|kruskal|wallis|kolmogorov|smirnov|mann|whitney|mwu|pearson|spearman|correlation|permutation|discriminative|p[- ]?value|statistical test",
        "filter_based_feature_selection": r"relief|relieff|information gain|mutual information|\bmic\b|maximal information coefficient|mrmr|cfs|fisher|fdr|chi[- ]?square|chi2|correlation[- ]?based",
        "wrapper_based_feature_selection": r"forward|backward|recursive|\brfe\b|svm[- ]?rfe|stepwise|swda|genetic|wrapper|boruta|sequential feature",
        "embedded_feature_importance": r"svm weights?|rf weights?|random forest|feature weights?|weights?|weighted|gini importance|tree importance|xgboost importance|lightgbm importance|lasso",
        "explainable_ai_model_interpretation": r"shap|shapley|lime|grad[- ]?cam|cam\b|integrated gradients|saliency|attention|attention map|feature importance|importance|ablation|leave[- ]?one[- ]?out|permutation importance|pdp|partial dependence|explainable|xai|interpretability",
        "dimensionality_reduction_representation_learning": r"pca|principal component|vae|autoencoder|latent|representation|dimensionality|embedding visualization|t[- ]?sne|umap",
        "ensemble_hybrid_feature_selection": r"ensemble|ensembled|voting|combined|combining|combination|hybrid|fusion based feature selection",
        "feature_engineering_exploratory_analysis": r"spatial|temporal|spatiotemporal|spatio-temporal|k[- ]?means|clustering|lmem|linear mixed|exploratory|feature engineering",
    }
    cats = regex_categories(text, patterns)
    cats["not_reported"] = int(is_invalid(value))
    cats["other_uncategorized_feature_importance_method"] = int(not cats["not_reported"] and not any(cats[k] for k in patterns))
    return cats


def feature_importance_result_categories(value: Any) -> dict[str, int]:
    text = normalize_text(value)
    patterns = {
        "motor_and_kinematic_features": r"speed|velocity|acceleration|movement|motion|kinematic|gait|stride|walking|grip|force|gesture|grasp|head rotation|yaw|roll|amplitude|tablet|touch",
        "gaze_and_visual_attention_features": r"gaze|fixation|saccade|eye movement|eye tracking|scanpath|visual focus|attention|\baoi\b|mouth|eye contact|heatmap|saliency",
        "speech_and_acoustic_features": r"speech|acoustic|voice|vocal|prosody|pitch|\bf0\b|mfcc|rhythm|sentiment|emotional|\btext feature\b|\bword\b|embedding",
        "facial_and_social_features": r"\bau\d+\b|facial|smiling|social|interaction|presence of face|emotion",
        "multimodal_combination_features": r"fusion|fused|combined|combining|concatenat|integrat|multimodal|all features|feature combination|dual[- ]?stream",
        "other_behavioral_features": r"entropy|visual focus|steerable|rgb|color|colour|intensity|orientation|biological movement|questionnaire|score",
    }
    cats = regex_categories(text, patterns)
    cats["not_reported"] = int(is_invalid(value))
    cats["other_uncategorized_feature_importance_result"] = int(not cats["not_reported"] and not any(cats[k] for k in patterns))
    return cats


def bias_mitigation_categories(value: Any) -> dict[str, int]:
    text = normalize_text(value)
    patterns = {
        "smote": r"smote|synthetic minority|over[- ]?sampling|oversampling",
        "adasyn_or_synthetic_sampling": r"adasyn|synthetic data|synthetic samples|synthetic examples|data synthesis",
        "class_weights_or_cost_sensitive_learning": r"class weights?|weighted loss|cost[- ]?sensitive|sample weights?|focal loss|balanced loss",
        "stratified_sampling": r"stratified|stratification|stratify|stratifying|stratified k[- ]?fold",
        "under_or_over_sampling": r"under[- ]?sampling|undersampling|over[- ]?sampling|oversampling|random under|random over",
        "data_augmentation": r"data augmentation|augment|augmentation|rotation|flip|noise injection|cropping|synthetic augmentation|gan augmentation",
        "balanced_split_or_matching": r"balanced dataset|balanced groups|matched groups|age[- ]?matched|gender[- ]?matched|sex[- ]?matched|similar ratio",
    }
    cats = regex_categories(text, patterns)
    cats["not_reported"] = int(is_invalid(value) or text == "no")
    cats["other_uncategorized_bias_mitigation"] = int(not cats["not_reported"] and not any(cats[k] for k in patterns))
    return cats


def cross_dataset_validation_categories(value: Any) -> dict[str, int]:
    text = normalize_text(value)
    strict = bool(
        re.search(
            r"^yes$|cross[-\s]?corpus|cross[-\s]?dataset|cross[-\s]?site|leave[-\s]?one[-\s]?dataset[-\s]?out|"
            r"leave[-\s]?one[-\s]?site[-\s]?out|loxo|loso|train(?:ed)?\s+on\s+.*test(?:ed)?\s+on|"
            r"train(?:ed)?\s+on\s+.*evaluat(?:ed|ion)\s+on|external validation",
            text,
        )
    )
    external = bool(
        re.search(
            r"second dataset|secondary dataset|independent dataset|external dataset|separate dataset|another dataset|"
            r"two datasets|multiple datasets|multi[- ]?dataset|separate cohort|independent cohort|different cohort|"
            r"held[- ]?out dataset|unseen dataset|unseen site|out[- ]?of[- ]?sample|generalization dataset|generalisation dataset",
            text,
        )
    )
    no = bool(re.search(r"^no$|^no,|^no\.|^no ", text))
    missing = is_invalid(value)
    return {
        "strict_cross_corpus_validation": int(strict and not missing),
        "external_dataset_multi_dataset_validation": int(external and not missing),
        "any_cross_dataset_generalizability_evaluation": int((strict or external) and not missing),
        "reported_no_cross_dataset_validation": int(no and not missing),
        "not_reported": int(missing),
        "other_valid_unclear": int(not missing and not (strict or external or no)),
    }


def cross_validation_categories(value: Any) -> tuple[dict[str, int], float]:
    text = normalize_text(value)
    pattern = r"^yes$|cross[- ]?validation|\bcv\b|k[- ]?fold|\d+[- ]?fold|folds?\b|leave[- ]?one[- ]?out|loocv|stratified k[- ]?fold|nested cross[- ]?validation|subject[- ]?independent cross[- ]?validation"
    reported = int(bool(re.search(pattern, text)) and not is_invalid(value))
    folds = np.nan
    match = re.search(r"(\d+)\s*[- ]?fold", text)
    if match:
        folds = float(match.group(1))
    elif "loocv" in text or "leave-one-out" in text or "leave one out" in text:
        folds = np.nan
    return {"cross_validation_reported": reported, "not_reported": int(is_invalid(value)), "reported_no_cross_validation": int(text == "no")}, folds


def loxo_categories(value: Any) -> dict[str, int]:
    text = normalize_text(value)
    reported = bool(re.search(r"^yes$|\bloxo\b|\bloocv\b|leave[- ]?one[- ]?out|leave[- ]?one[- ]?subject[- ]?out|leave[- ]?one[- ]?site[- ]?out|leave[- ]?one[- ]?dataset[- ]?out|loso|lopo", text))
    return {"loxo_reported": int(reported and not is_invalid(value)), "not_reported": int(is_invalid(value)), "reported_no_loxo": int(text == "no")}


def real_time_categories(value: Any) -> dict[str, int]:
    text = normalize_text(value)
    reported = bool(re.search(r"^yes$|real[- ]?time|online inference|live|real time analysis|real[- ]?time detection|app[- ]?based|mobile app|deployed|deployment", text))
    return {"real_time_analysis_reported": int(reported and not is_invalid(value)), "not_reported": int(is_invalid(value)), "reported_no_real_time": int(text == "no")}


def dataset_type_nominal(value: Any) -> str:
    text = normalize_text(value)
    if text in INVALID_TEXT_VALUES:
        return "not_reported_or_placeholder"
    if re.search(r"^\s*yes\b|\bnew\b|\bcollected\b|\bcreated\b", text):
        return "new_dataset_or_primary_data_collection"
    if re.search(r"^\s*no\b|\bexisting\b|\bprevious\b|\bpreviously\b|\bearlier\b|\bpublic\b|\bopen dataset\b|\bsecondary\b", text):
        return "existing_dataset_or_secondary_data"
    return "manual_review_unclear"


DATASET_ALIAS_PATTERNS = {
    "dataset_abide": r"\babide\b|autism brain imaging data exchange",
    "dataset_sfari": r"\bsfari\b",
    "dataset_saliency4asd": r"saliency4asd|saliency for asd",
    "dataset_childes": r"\bchildes\b",
    "dataset_guesswhat": r"guesswhat",
    "dataset_affectnet": r"affectnet",
    "dataset_de_enigma": r"de[- ]?enigma",
    "dataset_kaggle": r"kaggle",
    "dataset_geneva_autism_cohort": r"geneva autism cohort",
    "dataset_ext_dataset": r"ext[- ]?dataset",
    "dataset_autistic_children_facial_image": r"autistic children facial image",
    "dataset_et_support_asd": r"et dataset to support|support the research on asd",
    "dataset_visualization_eye_tracking_asd": r"visuali[sz]ation.*eye tracking.*autism|eye tracking scanpaths",
    "dataset_asd_toddler": r"asd toddler|screening data for toddlers|toddlers.*saudi",
    "dataset_child_pathological_speech": r"child pathalogical speech|child pathological speech",
    "dataset_newly_collected": r"newly collected|primary data|own dataset|collected",
}


def dataset_name_categories(value: Any) -> dict[str, int]:
    text = normalize_text(value)
    cats = regex_categories(text, DATASET_ALIAS_PATTERNS)
    cats["dataset_not_given"] = int(text in INVALID_TEXT_VALUES or text in {"not provided", "not specified"})
    if cats["dataset_not_given"]:
        for key in DATASET_ALIAS_PATTERNS:
            cats[key] = 0
    cats["dataset_other_uncategorized"] = int(not cats["dataset_not_given"] and not any(cats[k] for k in DATASET_ALIAS_PATTERNS))
    return cats


def data_description_categories(value: Any) -> dict[str, int]:
    text = normalize_text(value)
    patterns = {
        "newly_collected_data": r"collected|recorded|acquired|interacting|played|participants|children",
        "public_existing_dataset": r"dataset|kaggle|childes|affectnet|saliency4asd|public|existing",
        "private_or_restricted_dataset": r"private|restricted|cohort|hospital|clinical records",
        "clinical_data": r"clinical|diagnostic|ados|patient|cohort",
        "video_data": r"\bvideo\b|\bvideos\b|clips|footage|recordings",
        "image_data": r"\bimage\b|\bimages\b|facial image|thermal image|mri scan|scan path images",
        "eye_tracking_data": r"eye[- ]?tracking|gaze|scanpath|fixation|aoi|x-y eye positions|eye movements",
        "audio_or_speech_data": r"audio|speech|voice|transcript|language|acoustic|samples",
        "movement_or_pose_data": r"motor|movement|motion|\bpose\b|kinematic|force plate|center of pressure|accelerat|wheel rotation|gait",
        "questionnaire_data": r"questionnaire|survey|scale|q-chat|features",
        "physiological_data": r"\beeg\b|physiological|thermal",
        "neuroimaging_data": r"\bmri\b|fmri|neuroimaging|scans",
        "robot_interaction_data": r"robot|nao",
        "game_iot_sensor_data": r"game|iot|toy car|sensorized|sensor",
        "multimodal_data": r"multimodal|video and audio|eeg.*video|voice.*eye|facial.*voice|audio/video|audio and video",
    }
    cats = regex_categories(text, patterns)
    cats["data_description_not_given"] = int(is_invalid(value))
    cats["other_uncategorized_data_description"] = int(not cats["data_description_not_given"] and not any(cats[k] for k in patterns))
    return cats


RECOMMENDATION_PATTERNS = {
    "larger_sample_size": r"larger sample|bigger sample|increase sample|more subjects|more participants|more patients|sample size|population size|larger number",
    "larger_or_new_dataset": r"larger dataset|bigger dataset|more data|new dataset|collect(?:ing)? new data|expand dataset|dataset size|additional data",
    "participant_diversity": r"diverse|diversity|race|racial|ethnic|culture|cultural|age range|younger|adults|less verbal|balanced dataset|wider population|different disorders",
    "sex_gender_representation": r"\bsex\b|gender|female|females|girls|male|males|sex differences",
    "group_matching_or_comparison": r"match(?:ing)?|comparison group|control group|typically developing|iq|baseline group|compare groups",
    "severity_subtyping_differential_diagnosis": r"severity|subtyp|differential diagnosis|cognitive deficits|other diagnos|pd|clinical vs nonclinical",
    "longitudinal_followup": r"longitudinal|follow[- ]?up|over time|developmental|track change|predictive value over time",
    "external_or_clinical_validation": r"external validation|clinical validation|validate.*hospital|hospital|real[- ]?world validation|validate.*clinical|collaborate with hospitals|cooperative hospitals",
    "multimodal_or_additional_modalities": r"multimodal|additional modalit|more modalit|combine.*(voice|eye|movement|physiological|video|audio)|add modalities|eye tracking.*movement",
    "feature_engineering_or_additional_features": r"feature|features|feature engineering|additional motion|image features|kinematic|demographic information|object present",
    "advanced_models_or_model_expansion": r"advanced (?:deep learning|model)|different ml models|more models|ensemble models|attention mechanism|sophisticated|improve vgg|test.*models|model expansion",
    "data_augmentation_or_synthetic_data": r"augmentation|synthetic|gan|\bpose variation\b|occlusion",
    "model_evaluation_and_comparison": r"model evaluation|evaluation|compare|comparison|baseline|cross[- ]?validation|performance metrics|greater eval",
    "interpretability_explainability": r"interpret|explain|shap|lime|transparency|trust|feature importance|layer-wise",
    "task_or_stimulus_adaptation": r"task|stimuli|stimulus|adapt|suitable|object|design choices|video as stimuli",
    "automation_efficiency_accessibility": r"automated|automation|efficiency|time-efficient|accessible|autonomy|reduced human|fully automated",
    "clinical_implementation_or_utility": r"clinical usability|clinical utility|implementation|deploy|screening tools|diagnostic procedures|hospitals|assistive technolog",
    "reporting_and_data_transparency": r"reporting|provide more information|mentioning|transparency|data transparency|participants in each group|diagnosis information",
    "diagnostic_assessment_comparison": r"formal asd|assessment|diagnostic assessment|m-chat|asrs|confirmed diagnosis|diagnosis comparison",
}


def recommendation_categories(value: Any) -> dict[str, int]:
    text = normalize_text(value)
    cats = regex_categories(text, RECOMMENDATION_PATTERNS)
    missing = int(is_invalid(value) or text in {"no", "none"})
    cats["no_recommendation_reported"] = missing
    cats["other_uncategorized_recommendation"] = int(not missing and not any(cats[k] for k in RECOMMENDATION_PATTERNS))
    return cats


def open_source_access_nominal(value: Any) -> str:
    text = normalize_text(value)
    if text in INVALID_TEXT_VALUES:
        return "empty_not_reported"
    if re.search(r"request|upon request", text):
        return "available_on_request"
    if re.search(r"limited|partial|only|features only|processed data|not raw|subset|de-identified|anonymi[sz]ed|restricted", text):
        return "limited_access"
    if re.fullmatch(r"\s*yes[.,;:]*\s*", text):
        return "yes_open"
    if re.fullmatch(r"\s*no[.,;:]*\s*", text):
        return "no_not_open"
    return "manual_annotation"


def time_frame_categories(value: Any) -> tuple[dict[str, int], float]:
    text = normalize_text(value)
    not_long = bool(re.search(r"^\s*no\b|\bsingle\b|\bcross[- ]?sectional\b", text))
    cross = bool(not not_long and re.search(r"\bcross[- ]?sectional\b", text))
    long = bool(not not_long and re.search(r"^\s*yes\b|longitudinal|follow[- ]?up|multiple|month|week|year|day", text))
    cats = {
        "longitudinal_or_repeated_time_points": int(long),
        "not_longitudinal_or_single_time_point": int(not_long),
        "cross_sectional": int(cross),
        "unreported_or_unclear": int(not (long or not_long or cross)),
    }
    duration = extract_duration_days(value)
    return cats, duration


def extract_duration_days(value: Any) -> float:
    text = normalize_text(value)
    if not re.search(r"month|week|year|day", text):
        return np.nan
    unit = None
    mult = None
    if "week" in text:
        unit, mult = "week", 7.0
    elif "month" in text:
        unit, mult = "month", 30.5
    elif "year" in text:
        unit, mult = "year", 365.25
    elif "day" in text:
        unit, mult = "day", 1.0
    if unit is None or mult is None:
        return np.nan
    m = re.search(r"(\d+(?:\.\d+)?)\s*(?:-|to)\s*(\d+(?:\.\d+)?)", text)
    if m:
        return ((float(m.group(1)) + float(m.group(2))) / 2.0) * mult
    m = re.search(r"(\d+(?:\.\d+)?)", text)
    if m:
        return float(m.group(1)) * mult
    return np.nan


def data_collection_tool_categories(value: Any) -> dict[str, int]:
    text = normalize_text(value)
    patterns = {
        "tobii": r"\btobii\b",
        "smi": r"\bsmi\b|sensomotoric",
        "nao_robot": r"\bnao\b|robot",
        "gazefinder": r"gazefinder",
        "eyelink": r"eyelink|eye\s*link",
        "gazepoint": r"gazepoint",
        "force_plates": r"force plate|plates",
        "kinect": r"kinect",
        "openpose": r"openpose|open pose",
        "openface": r"openface|open face",
        "accelerometer": r"accelerometer|imu|inertial measurement|sensor",
        "eeg_electrodes_or_sensor_net": r"electrode|hydrocel|sensor net|geodesic|eeg",
        "mri_scanner": r"\bmri\b|\bfmri\b|scanner",
        "camera_or_webcam": r"camera|webcam|video camera|thermal",
        "microphone_or_audio_recorder": r"\bmic\b|microphone|audio recorder|recorder",
    }
    cats = regex_categories(text, patterns)
    cats["not_reported"] = int(is_invalid(value))
    cats["other_uncategorized_tool"] = int(not cats["not_reported"] and not any(cats[k] for k in patterns))
    return cats


def limitation_categories(value: Any) -> dict[str, int]:
    text = normalize_text(value)
    patterns = {
        "diagnosis_limitations": r"ados|adi[- ]?r|cars|clinical diagnos|standardized diagnostic|validated diagnostic|not confirmed|diagnostic tool|diagnos.*self.report|diagnos.*parent",
        "small_sample_size": r"small sample|sample size|limited sample|limited number|limited data|limited dataset|few subject|few participant|larger dataset|larger sample|more data|small cohort|pilot analysis",
        "lack_of_demographic_diversity": r"cultur\w*|demographic|divers\w*|race|racial|ethnic|socioeconomic|language|geographic|homogene\w*|selection bias|population bias",
        "lack_of_sex_gender_balance": r"\bsex\b|gender|female|females|girls|male|males|single gender|sex imbalance|gender imbalance|few females",
        "limited_age_generalizability": r"age group|age range|age variability|younger than|very young children|only adults|only adolescents|only children|specific.*age",
        "lack_of_iq_adaptive_behavior_measures": r"\biq\b|intelligence quotient|cognitive ability|adaptive behaviou?r|functional level|developmental level|not measured|participant information",
        "single_site_sample": r"single site|single-site|single cent(?:er|re)|one clinic|one school|one hospital|one university|single institution",
        "lack_of_external_validation": r"external validation|independent dataset|external dataset|cross[- ]?corpus|cross[- ]?site|replication|overfitting|standard dataset",
        "limited_generalizability": r"generaliz|generalis|not representative|may not apply|limited coverage|specific population|selection bias",
        "lack_of_longitudinal_data": r"cross[- ]?sectional|longitudinal|follow[- ]?up|over time|developmental trajectories",
        "missing_comparison_group": r"no control group|comparison group|control group|typically developing|neurotypical|non[- ]?asd|not matched|mismatched",
        "measurement_tool_limitations": r"tool|device|instrument|eye tracker|camera|microphone|audio quality|speech-to-text|robot|\bvr\b|headset|not evaluated|not described",
        "task_paradigm_limitations": r"task|paradigm|stimuli|stimulus|short duration|short time window|single images|structured setting|real-world|not suitable",
        "child_task_feasibility_limitations": r"attention|young children|shy.*children|bored|fatigue|discomfort|headset|immersive environment|sensor|eye tracker.*affect",
        "model_interpretability_limitations": r"interpret|explain|black[- ]?box|feature importance|unclear.*model|conclusive explanation",
        "data_quality_missing_data_limitations": r"unbalanced|imbalanced|class imbalance|missing data|data quality|poor accuracy|false negatives|data availability|lack of quality",
        "experimental_setup_limitations": r"setting|experimental setup|experiment|procedure|protocol|setup|not strictly controlled|manual|human oversight|therapist|room|environment",
        "analysis_setup_limitations": r"computational cost|autoencoder|supervised|pre[- ]?processing|diarization|segmentation|feature extraction|performance metrics|ml models|not clear.*analysis|not enough info",
    }
    cats = regex_categories(text, patterns)
    placeholder = bool(re.fullmatch(r"\s*|[-\s]+|(?:no|yes|n/d|nd|n/a|na|nan|none|no limitation|no limitations|not reported|not applicable|n\.a\.)(?:\s+(?:no|yes|n/d|nd|n/a|na|nan|none|no limitation|no limitations|not reported|not applicable|n\.a\.|-))*\s*", text))
    cats["empty_or_not_applicable"] = int(placeholder)
    cats["need_manual_revision"] = int(not placeholder and not any(cats[k] for k in patterns))
    return cats


def novelty_yes_no_categories(value: Any) -> dict[str, int]:
    nominal = yes_no_nominal(value)
    return {"yes": int(nominal == "yes"), "no": int(nominal == "no"), "empty_or_not_reported": int(nominal == "not_reported"), "other_nonempty_manual_review": int(nominal == "unclear")}


def sensitive_data_protection_categories(value: Any) -> dict[str, int]:
    text = normalize_text(value)
    patterns = {
        "deidentification_or_participant_id": r"de[- ]?identif\w*|participant id|subject id|anonymous id|coded id|personal information removed",
        "face_blurring_or_mosaicing": r"mosaic\w*|blur\w*|face.*removed|removed.*face|mask\w*.*face|face.*mask\w*",
        "consent_or_parent_choice": r"consent|parent\w*.*choose|choose when|permission",
        "restricted_access_or_secure_storage": r"restricted|secure|encrypted|password|access control|stored securely|protected server",
    }
    cats = regex_categories(text, patterns)
    placeholder = is_invalid(value) or text == "no"
    cats["not_reported_or_placeholder"] = int(placeholder)
    cats["manual_review_non_placeholder"] = int(not placeholder and not any(cats[k] for k in patterns))
    return cats


def future_goal_categories(value: Any) -> dict[str, int]:
    text = normalize_text(value)
    patterns = {
        "dataset_size_and_diversity": r"larger data|more data|increase dataset|population diversity|female|gender differences|balanced groups|group variability|age|younger|other diagnos|neurological|different disorders",
        "feature_expansion_and_fusion": r"more feature|feature fusion|combine features|data fusion|behavioral modalit|correlation between features|demographic",
        "validation_and_generalizability": r"\bvalidation\b|external data|multiple data|longitudinal|follow[- ]?up|different tasks|better experimental setup|generaliz|external validation",
        "model_performance_and_optimization": r"\baccuracy\b|classification performance|improve performance|optimi[sz]e|efficiency|different ml models|improve diagnosis",
        "automation_tools_and_implementation": r"fully automated|unified tool|sustainability|implementation|deploy",
        "further_analysis_unspecified": r"\bfurther\b|\bfuture work\b|futher",
        "not_given": r"not given|not specified|not reported|n/a|na|n\.a|nd|n\.d|n/d",
    }
    cats = regex_categories(text, patterns)
    cats["no_category_matched"] = int(not any(cats.values()) and not is_invalid(value))
    return cats


def classify_variable_pair(domain: str, subgroup: str, variable_label: str) -> tuple[str, str, str, str, str, bool, bool]:
    label = normalize_text(variable_label)
    if domain == "quality":
        return "quality_issue", "binary", "nominal", "ICR exception rule", "ICR pipeline PDF sections 3.1-3.2", True, False
    if label == "range age":
        return "age_range", "multi_label", "nominal", "parse_age_range_cell/categories_for_age_ranges", "helper_functions_.py/rq1_.py plus ICR PDF section 4.1", True, False
    if label in {"mean age", "std age"}:
        return "numeric_age", "numeric", "interval", "parse_numeric_age_value", "helper_functions_.py/rq1_.py plus ICR PDF section 4.2", True, False
    if label.startswith("#") or label in {"# male participants"}:
        return "numeric_count", "numeric", "interval", "parse_subgroup_count_sum/parse_numeric", "rq1_.py", False, False
    if label == "label":
        return "terminology", "multi_label", "nominal", "categorize_terminology", "rq1_.py", False, False
    if "assessment method" in label or "assesment method" in label:
        return "diagnostic_method", "multi_label", "nominal", "compute_diagnosis_methods", "rq1_.py", False, False
    if "other assessment" in label:
        return "assessment_presence", "binary", "nominal", "count_nonmissing_rows", "rq1_.py", False, False
    if label == "comorbidities":
        return "presence_binary", "binary", "nominal", "count_comorbidities", "rq1_.py", False, False
    if label.startswith("match in"):
        return "yes_no_nominal", "nominal", "nominal", "count_yes_rows", "rq1_.py", False, False
    if domain == "behaviors" and label in {"gaze", "speech", "motor"}:
        return "modality_yes_no", "nominal", "nominal", "yes_no_modality_summary", "helper_functions_.py/rq3_.py", False, False
    if domain == "behaviors" and label in {"other behavioural", "other type of data"}:
        return "other_behavioral", "multi_label", "nominal", "compute_other_behavioral_keywords", "helper_functions_.py/rq3_.py", False, False
    if domain == "behaviors" and "fusion" in label:
        return "feature_fusion", "multi_label", "nominal", "feature_fusion_summary", "rq3_.py", False, False
    if domain == "ai" and label == "algorithms used":
        return "algorithms", "multi_label", "nominal", "algorithms_broad/detect_algorithm_families", "rq4_.py", False, False
    if domain == "ai" and label == "features (or e2e)":
        return "features", "multi_label", "nominal", "features_broad", "rq4_.py", False, False
    if domain == "ai" and "learning type" in label:
        return "learning_type", "multi_label", "nominal", "machine_learning_paradigm", "rq4_.py", False, False
    if domain == "ai" and label == "evaluation metrics":
        return "evaluation_metrics", "multi_label", "nominal", "evaluation_metrics", "rq4_.py", False, False
    if domain == "ai" and label == "best performing model":
        return "best_model", "mixed", "nominal", "detect_algorithm_families/exact_model_name", "rq4_.py plus ICR PDF section 5.4", False, True
    if domain == "ai" and label == "best performance":
        return "best_performance", "numeric", "interval", "extract_accuracy_from_row", "rq4_.py", False, False
    if domain == "ai" and label == "features importance technique":
        return "feature_importance_method", "multi_label", "nominal", "compute_interpretation_methods", "rq4_.py", False, False
    if domain == "ai" and label == "features importance result":
        return "feature_importance_result", "multi_label", "nominal", "compute_feature_importance_result", "rq4_.py", False, False
    if domain == "ai" and "balancing" in label:
        return "bias_mitigation", "multi_label", "nominal", "compute_bias_mitigation", "rq4_.py", False, False
    if domain == "ai" and label == "cross corpus validation":
        return "cross_dataset_validation", "multi_label", "nominal", "compute_cross_dataset_validation", "rq4_.py", False, False
    if domain == "ai" and label == "x-fold cross-validation":
        return "cross_validation", "mixed", "nominal", "compute_cross_validation", "rq4_.py", False, False
    if domain == "ai" and label == "loxo":
        return "loxo", "multi_label", "nominal", "compute_LOXO", "rq4_.py", False, False
    if domain == "ai" and "real time" in label:
        return "real_time", "multi_label", "nominal", "compute_real_time_analysis", "rq4_.py", False, False
    if domain == "study" and label == "study setting":
        return "study_setting", "nominal", "nominal", "compute_study_setting", "helper_functions_.py/rq2_.py", False, False
    if domain == "study" and label == "study goal":
        return "study_goal", "multi_label", "nominal", "compute_study_goals", "rq2_.py", False, False
    if domain == "study" and label == "data collection?":
        return "dataset_type", "nominal", "nominal", "compute_dataset_type", "rq2_.py", False, False
    if domain == "study" and label == "dataset name":
        return "dataset_name", "multi_label", "nominal", "dataset alias dictionary", "ICR pipeline PDF section 5.1", False, True
    if domain == "study" and label == "is the data open source?":
        return "open_source_access", "nominal", "nominal", "open_source_access_summary", "rq2_.py", False, False
    if domain == "study" and label == "data description":
        return "data_description", "multi_label", "nominal", "ICR data description codebook", "ICR pipeline PDF section 5.2", False, True
    if domain == "study" and label == "longitudinally (how long)":
        return "time_frame", "mixed", "nominal", "compute_time_frame/extract_duration_to_days", "rq2_.py", False, False
    if domain == "study" and label == "data collection tool":
        return "data_collection_tool", "multi_label", "nominal", "data_collection_tool_summary", "rq2_.py", False, False
    if domain == "study" and label == "task for the participants":
        return "task_type", "multi_label", "nominal", "compute_task_type", "helper_functions_.py/rq2_.py", False, False
    if domain == "study" and label == "open source code":
        return "open_source_access", "nominal", "nominal", "open_source_access_summary", "rq2_.py", False, False
    if domain == "study" and label == "study limitations":
        return "limitations", "multi_label", "nominal", "rq_limitation_categories", "rq2_.py", False, False
    if domain == "study" and "reccomendations" in label:
        return "recommendations", "multi_label", "nominal", "recommendation regex framework", "ICR pipeline PDF section 5.3", False, True
    if domain == "study" and label == "main findings":
        return "main_findings", "multi_label", "nominal", "goal_finding_hybrid_summary", "rq2_.py", False, False
    if domain == "novelty_and_sensitivity" and label == "future research pipelines":
        return "future_goals", "multi_label", "nominal", "compute_future_goals_categories", "rq2_.py", False, False
    if domain == "novelty_and_sensitivity" and label == "measures taken to protect sensitive data":
        return "sensitive_data_protection", "multi_label", "nominal", "sensitive_data_protection_summary", "rq2_.py", False, False
    if domain == "novelty_and_sensitivity":
        return "novelty_yes_no", "multi_label", "nominal", "compute_yes_no_text_count", "rq2_.py", False, False
    return "normalized_text_nominal", "nominal", "nominal", "normalize_text", "ICR fallback", False, False


def get_block_domain(block_name: str) -> str:
    base = re.sub(r"_?coder[12]$", "", block_name, flags=re.IGNORECASE).strip()
    return slugify(base)


def participant_subgroup(position: int) -> str:
    if 0 <= position <= 8:
        return "asd"
    if 9 <= position <= 17:
        return "neurotypical"
    if 18 <= position <= 27:
        return "other_diagnosis"
    return ""


def canonical_variable_label(raw_label: str, domain: str, subgroup: str, position: int) -> str:
    label = (raw_label or "").strip()
    normalized = normalize_text(label)
    replacements = {
        "# nd participants": "# neurotypical participants",
        "assesment method": "Assessment method",
        "other assessment": "Other assessments",
        "fusion tecnique": "fusion technique",
        "balancing/unbiasing tecnique": "balancing/unbiasing technique",
        "reccomendations": "recommendations",
    }
    if normalized in replacements:
        label = replacements[normalized]
    if domain == "participants":
        prefix = subgroup or f"position_{position}"
        return f"{prefix} {label}".strip()
    return label


def build_workbook_model(workbook_path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[VariablePair], dict[str, int], list[dict[str, Any]]]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = wb["Sheet1"]
    merged_map: dict[tuple[int, int], Any] = {}
    for rg in ws.merged_cells.ranges:
        val = ws.cell(rg.min_row, rg.min_col).value
        for r in range(rg.min_row, rg.max_row + 1):
            for c in range(rg.min_col, rg.max_col + 1):
                merged_map[(r, c)] = val

    def cell_value(row: int, col: int) -> Any:
        v = ws.cell(row, col).value
        return v if v is not None else merged_map.get((row, col))

    columns: list[dict[str, Any]] = []
    for c in range(1, ws.max_column + 1):
        columns.append(
            {
                "col_idx": c,
                "excel_col": get_column_letter(c),
                "row1": cell_value(1, c),
                "row2": cell_value(2, c),
            }
        )

    rows: list[dict[str, Any]] = []
    for r in range(3, ws.max_row + 1):
        row = {"excel_row": r}
        for col in columns:
            row[col["excel_col"]] = ws.cell(r, col["col_idx"]).value
        rows.append(row)

    idx_by_block: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for col in columns:
        if col["row1"] not in {None, "Paper_id", "Title", "Coder", "Notes"}:
            idx_by_block[str(col["row1"])].append(col)

    block_pairs = [
        ("Quality_Coder1", "Quality_Coder2"),
        ("Participants_Coder1", "Participants_Coder2"),
        ("Behaviors_Coder1", "Behaviors_Coder2"),
        ("AI_Coder1", "AI_Coder2"),
        ("study_coder1", "study_coder2"),
        ("Novelty_and_sensitivity_Coder1", "Novelty_and_sensitivity_Coder2"),
    ]

    variable_pairs: list[VariablePair] = []
    mapping_issues: list[dict[str, Any]] = []
    for reviewer_block, final_block in block_pairs:
        reviewer_cols = idx_by_block[reviewer_block]
        final_cols = idx_by_block[final_block]
        if len(reviewer_cols) != len(final_cols):
            mapping_issues.append({"issue_type": "block_length_mismatch", "reviewer_block": reviewer_block, "final_block": final_block})
        for pos, (rcol, fcol) in enumerate(zip(reviewer_cols, final_cols)):
            domain = get_block_domain(reviewer_block)
            subgroup = participant_subgroup(pos) if domain == "participants" else ""
            raw_label = str(rcol.get("row2") or fcol.get("row2") or "")
            canonical_label = canonical_variable_label(raw_label, domain, subgroup, pos)
            processor, vtype, level, result_function, source_script, is_exception, is_new = classify_variable_pair(domain, subgroup, raw_label)
            variable_name = "__".join([x for x in [domain, subgroup, slugify(raw_label)] if x])
            variable_pairs.append(
                VariablePair(
                    domain=domain,
                    subgroup=subgroup,
                    variable_label=canonical_label,
                    variable_name=variable_name,
                    reviewer_col_idx=rcol["col_idx"],
                    final_col_idx=fcol["col_idx"],
                    reviewer_excel_col=rcol["excel_col"],
                    final_excel_col=fcol["excel_col"],
                    reviewer_block=reviewer_block,
                    final_block=final_block,
                    position_in_block=pos,
                    processor=processor,
                    variable_type=vtype,
                    measurement_level=level,
                    source_script=source_script,
                    results_function=result_function,
                    is_exception=is_exception,
                    is_new_icr_rule=is_new,
                )
            )

    special_cols = {}
    for col in columns:
        row2 = normalize_text(col["row2"])
        row1 = normalize_text(col["row1"])
        if row1 == "paper_id":
            special_cols["paper_id"] = col["col_idx"]
        elif row1 == "title":
            special_cols["title"] = col["col_idx"]
        elif row2 == "final annotation coder":
            special_cols["final_coder"] = col["col_idx"]
        elif row2.startswith("review coder"):
            special_cols["review_coder"] = col["col_idx"]
        elif row1 == "notes":
            special_cols["notes"] = col["col_idx"]

    return rows, columns, variable_pairs, special_cols, mapping_issues


def process_value(
    pair: VariablePair,
    raw: Any,
    paper_id: str,
    rating_role: str,
    coder_id: str,
    unmatched: list[dict[str, Any]],
    evaluation_metrics_value: Any = None,
) -> dict[str, Any]:
    processor = pair.processor
    text_norm = normalize_text(date_as_age_range(raw) if processor == "age_range" else raw)
    result = {
        "normalized_value": text_norm,
        "nominal_value": None,
        "categories": {},
        "numeric_value": np.nan,
        "parser_note": "",
    }

    if processor == "quality_issue":
        binary, nominal = quality_issue_indicator(raw)
        result.update({"normalized_value": nominal, "nominal_value": str(binary), "categories": {"quality_issue_recorded": binary}})
    elif processor in {"numeric_age", "numeric_count"}:
        num = parse_numeric_age_value(raw) if processor == "numeric_age" else parse_count_value(raw)
        result.update({"normalized_value": "" if pd.isna(num) else str(num), "numeric_value": num})
    elif processor == "age_range":
        cats, norm, parser = age_range_categories(raw)
        result.update({"normalized_value": norm, "categories": cats, "parser_note": parser})
    elif processor == "terminology":
        cats = terminology_categories(raw, pair.subgroup)
        result.update({"normalized_value": "; ".join([k for k, v in cats.items() if v]) or "not_reported", "categories": cats})
        add_unmatched_if_needed(unmatched, paper_id, pair.variable_name, rating_role, coder_id, raw, result["normalized_value"], "terminology", cats)
    elif processor == "diagnostic_method":
        cats = diagnostic_method_categories(raw)
        result.update({"normalized_value": "; ".join([k for k, v in cats.items() if v]) or "not_reported", "categories": cats})
    elif processor in {"assessment_presence", "presence_binary"}:
        present = int(not is_invalid(raw) and normalize_text(raw) not in {"no", "none", "not applicable"})
        result.update({"normalized_value": "reported" if present else "not_reported", "nominal_value": str(present), "categories": {"reported": present}})
    elif processor == "yes_no_nominal":
        nominal = yes_no_nominal(raw)
        result.update({"normalized_value": nominal, "nominal_value": nominal})
    elif processor == "modality_yes_no":
        nominal = modality_nominal(raw)
        result.update({"normalized_value": nominal, "nominal_value": nominal})
    elif processor == "other_behavioral":
        cats = other_behavioral_categories(raw)
        result.update({"normalized_value": "; ".join([k for k, v in cats.items() if v]) or "not_reported", "categories": cats})
        add_unmatched_if_needed(unmatched, paper_id, pair.variable_name, rating_role, coder_id, raw, result["normalized_value"], "other_behavioral", cats)
    elif processor == "feature_fusion":
        cats = feature_fusion_categories(raw)
        result.update({"normalized_value": "; ".join([k for k, v in cats.items() if v]) or "no_fusion_or_not_reported", "categories": cats})
    elif processor == "algorithms":
        cats = algorithm_family_categories(raw)
        result.update({"normalized_value": "; ".join([k for k, v in cats.items() if v]) or "not_given", "categories": cats, "nominal_value": algorithm_group_nominal(raw)})
        add_unmatched_if_needed(unmatched, paper_id, pair.variable_name, rating_role, coder_id, raw, result["normalized_value"], "algorithm", cats)
    elif processor == "features":
        cats = features_broad_categories(raw)
        result.update({"normalized_value": "; ".join([k for k, v in cats.items() if v]) or "not_reported", "categories": cats})
        add_unmatched_if_needed(unmatched, paper_id, pair.variable_name, rating_role, coder_id, raw, result["normalized_value"], "features", cats)
    elif processor == "learning_type":
        cats = learning_type_categories(raw)
        result.update({"normalized_value": "; ".join([k for k, v in cats.items() if v]) or "not_reported", "categories": cats})
    elif processor == "evaluation_metrics":
        cats = evaluation_metric_categories(raw)
        result.update({"normalized_value": "; ".join([k for k, v in cats.items() if v]) or "not_reported", "categories": cats})
        add_unmatched_if_needed(unmatched, paper_id, pair.variable_name, rating_role, coder_id, raw, result["normalized_value"], "evaluation_metrics", cats)
    elif processor == "best_model":
        exact = exact_model_name(raw)
        cats = algorithm_family_categories(raw)
        exact_cat = {f"best_model_exact__{exact}": int(exact != "not_given")}
        combined = {**cats, **exact_cat}
        result.update({"normalized_value": exact, "nominal_value": algorithm_group_nominal(raw), "categories": combined})
        add_unmatched_if_needed(unmatched, paper_id, pair.variable_name, rating_role, coder_id, raw, exact, "best_model", cats)
    elif processor == "best_performance":
        num = extract_accuracy_percent(raw, evaluation_metrics_value)
        result.update({"normalized_value": "" if pd.isna(num) else str(num), "numeric_value": num})
    elif processor == "feature_importance_method":
        cats = feature_importance_method_categories(raw)
        result.update({"normalized_value": "; ".join([k for k, v in cats.items() if v]) or "not_reported", "categories": cats})
    elif processor == "feature_importance_result":
        cats = feature_importance_result_categories(raw)
        result.update({"normalized_value": "; ".join([k for k, v in cats.items() if v]) or "not_reported", "categories": cats})
        add_unmatched_if_needed(unmatched, paper_id, pair.variable_name, rating_role, coder_id, raw, result["normalized_value"], "feature_importance_result", cats)
    elif processor == "bias_mitigation":
        cats = bias_mitigation_categories(raw)
        result.update({"normalized_value": "; ".join([k for k, v in cats.items() if v]) or "not_reported", "categories": cats})
    elif processor == "cross_dataset_validation":
        cats = cross_dataset_validation_categories(raw)
        result.update({"normalized_value": "; ".join([k for k, v in cats.items() if v]) or "not_reported", "categories": cats})
    elif processor == "cross_validation":
        cats, folds = cross_validation_categories(raw)
        result.update({"normalized_value": "; ".join([k for k, v in cats.items() if v]) or "not_reported", "categories": cats, "numeric_value": folds})
    elif processor == "loxo":
        cats = loxo_categories(raw)
        result.update({"normalized_value": "; ".join([k for k, v in cats.items() if v]) or "not_reported", "categories": cats})
    elif processor == "real_time":
        cats = real_time_categories(raw)
        result.update({"normalized_value": "; ".join([k for k, v in cats.items() if v]) or "not_reported", "categories": cats})
    elif processor == "study_setting":
        nominal = study_setting_nominal(raw)
        result.update({"normalized_value": nominal, "nominal_value": nominal})
    elif processor == "study_goal":
        cats = study_goal_categories(raw)
        result.update({"normalized_value": "; ".join([k for k, v in cats.items() if v]) or "not_reported", "categories": cats})
    elif processor == "dataset_type":
        nominal = dataset_type_nominal(raw)
        result.update({"normalized_value": nominal, "nominal_value": nominal})
    elif processor == "dataset_name":
        cats = dataset_name_categories(raw)
        result.update({"normalized_value": "; ".join([k for k, v in cats.items() if v]) or "dataset_not_given", "categories": cats})
        add_unmatched_if_needed(unmatched, paper_id, pair.variable_name, rating_role, coder_id, raw, result["normalized_value"], "dataset_name", cats)
    elif processor == "open_source_access":
        nominal = open_source_access_nominal(raw)
        result.update({"normalized_value": nominal, "nominal_value": nominal})
    elif processor == "data_description":
        cats = data_description_categories(raw)
        result.update({"normalized_value": "; ".join([k for k, v in cats.items() if v]) or "data_description_not_given", "categories": cats})
        add_unmatched_if_needed(unmatched, paper_id, pair.variable_name, rating_role, coder_id, raw, result["normalized_value"], "data_description", cats)
    elif processor == "time_frame":
        cats, days = time_frame_categories(raw)
        result.update({"normalized_value": "; ".join([k for k, v in cats.items() if v]) or "unreported_or_unclear", "categories": cats, "numeric_value": days})
    elif processor == "data_collection_tool":
        cats = data_collection_tool_categories(raw)
        result.update({"normalized_value": "; ".join([k for k, v in cats.items() if v]) or "not_reported", "categories": cats})
        add_unmatched_if_needed(unmatched, paper_id, pair.variable_name, rating_role, coder_id, raw, result["normalized_value"], "data_collection_tool", cats)
    elif processor == "task_type":
        cats = task_type_categories(raw)
        result.update({"normalized_value": "; ".join([k for k, v in cats.items() if v]) or "not_given", "categories": cats})
    elif processor == "limitations":
        cats = limitation_categories(raw)
        result.update({"normalized_value": "; ".join([k for k, v in cats.items() if v]) or "empty_or_not_applicable", "categories": cats})
        add_unmatched_if_needed(unmatched, paper_id, pair.variable_name, rating_role, coder_id, raw, result["normalized_value"], "limitations", cats)
    elif processor == "recommendations":
        cats = recommendation_categories(raw)
        result.update({"normalized_value": "; ".join([k for k, v in cats.items() if v]) or "no_recommendation_reported", "categories": cats})
        if cats.get("other_uncategorized_recommendation"):
            add_unmatched_if_needed(unmatched, paper_id, pair.variable_name, rating_role, coder_id, raw, result["normalized_value"], "recommendations", {})
    elif processor == "main_findings":
        cats = study_goal_categories(raw)
        result.update({"normalized_value": "; ".join([k for k, v in cats.items() if v]) or "not_reported", "categories": cats})
    elif processor == "future_goals":
        cats = future_goal_categories(raw)
        result.update({"normalized_value": "; ".join([k for k, v in cats.items() if v]) or "not_given", "categories": cats})
    elif processor == "novelty_yes_no":
        cats = novelty_yes_no_categories(raw)
        result.update({"normalized_value": "; ".join([k for k, v in cats.items() if v]) or "empty_or_not_reported", "categories": cats})
    elif processor == "sensitive_data_protection":
        cats = sensitive_data_protection_categories(raw)
        result.update({"normalized_value": "; ".join([k for k, v in cats.items() if v]) or "not_reported_or_placeholder", "categories": cats})
        add_unmatched_if_needed(unmatched, paper_id, pair.variable_name, rating_role, coder_id, raw, result["normalized_value"], "sensitive_data_protection", cats)
    else:
        nominal = text_norm if text_norm not in INVALID_TEXT_VALUES else "not_reported"
        result.update({"normalized_value": nominal, "nominal_value": nominal})

    return result


def krippendorff_alpha_from_unit_values(unit_values: list[list[Any]], level: str) -> tuple[float, str]:
    return krippendorff_alpha(unit_values, level=level)


def missing_value(value: Any) -> bool:
    return is_missing(value)


def alpha_records(records: list[dict[str, Any]], level: str, value_key: str) -> dict[str, Any]:
    unit_to_values: dict[str, list[Any]] = defaultdict(list)
    coder_unit_seen: set[tuple[str, str]] = set()
    duplicate_coder_unit = 0
    for rec in records:
        value = rec.get(value_key)
        if missing_value(value):
            continue
        unit_id = str(rec["unit_id"])
        coder_id = str(rec["coder_id"])
        key = (unit_id, coder_id)
        if key in coder_unit_seen:
            duplicate_coder_unit += 1
            continue
        coder_unit_seen.add(key)
        unit_to_values[unit_id].append(value)

    unit_values = list(unit_to_values.values())
    valid_pair_groups = [vals for vals in unit_values if len(vals) >= 2]
    n_paired = len(valid_pair_groups)
    all_values = [v for vals in unit_values for v in vals if not missing_value(v)]
    alpha, reason = krippendorff_alpha_from_unit_values(unit_values, level)

    agreements = []
    abs_diffs = []
    for vals in valid_pair_groups:
        if len(vals) < 2:
            continue
        a, b = vals[0], vals[1]
        if level == "interval":
            agreements.append(float(a) == float(b))
            abs_diffs.append(abs(float(a) - float(b)))
        else:
            agreements.append(str(a) == str(b))

    observed_agreement = float(np.mean(agreements)) if agreements else np.nan
    prevalence = np.nan
    try:
        numeric_values = [float(v) for v in all_values]
        if set(numeric_values).issubset({0.0, 1.0}):
            prevalence = float(np.mean(numeric_values)) if numeric_values else np.nan
    except Exception:
        prevalence = np.nan

    ci_low = np.nan
    ci_high = np.nan
    if not pd.isna(alpha) and len(valid_pair_groups) >= 3:
        rng = random.Random(RANDOM_SEED)
        boot = []
        units = list(unit_to_values.keys())
        for _ in range(BOOTSTRAP_ITERATIONS):
            sampled_units = [rng.choice(units) for _ in units]
            sampled_groups = [unit_to_values[u] for u in sampled_units]
            a, _ = krippendorff_alpha_from_unit_values(sampled_groups, level)
            if not pd.isna(a):
                boot.append(a)
        if len(boot) >= 20:
            ci_low, ci_high = np.percentile(boot, [2.5, 97.5])

    out = {
        "n_units": len(unit_to_values),
        "n_paired": n_paired,
        "n_values": len(all_values),
        "observed_agreement": observed_agreement,
        "krippendorff_alpha": alpha,
        "ci_low": ci_low,
        "ci_high": ci_high,
        "undefined_reason": reason,
        "prevalence": prevalence,
        "duplicate_coder_unit_records": duplicate_coder_unit,
        "exact_agreement": observed_agreement,
        "mean_absolute_difference": float(np.mean(abs_diffs)) if abs_diffs else np.nan,
        "median_absolute_difference": float(np.median(abs_diffs)) if abs_diffs else np.nan,
        "min_absolute_difference": float(np.min(abs_diffs)) if abs_diffs else np.nan,
        "max_absolute_difference": float(np.max(abs_diffs)) if abs_diffs else np.nan,
    }
    return out


def build_outputs(workbook_path: Path, output_dir: Path) -> dict[str, Any]:
    missing_result_scripts = [f"{label} ({path})" for label, path in SOURCE_SCRIPT_MANIFEST if not path.exists()]
    if missing_result_scripts:
        raise FileNotFoundError(f"Missing finalized Results scripts: {missing_result_scripts}")

    rows, columns, pairs, special_cols, mapping_issues = build_workbook_model(workbook_path)
    col_by_idx = {col["col_idx"]: col for col in columns}
    col_letter_by_idx = {col["col_idx"]: col["excel_col"] for col in columns}

    run_warnings: list[str] = []
    coder_issues = list(mapping_issues)
    unmatched: list[dict[str, Any]] = []
    normalized_records: list[dict[str, Any]] = []
    binary_records: list[dict[str, Any]] = []
    numeric_records: list[dict[str, Any]] = []
    long_records: list[dict[str, Any]] = []
    coder_label_map: dict[str, str] = {}

    required_special = {"paper_id", "review_coder", "final_coder", "notes"}
    missing_special = required_special - set(special_cols)
    if missing_special:
        raise ValueError(f"Missing required special columns: {sorted(missing_special)}")

    novelty_final_override_coder_raw = ""
    evaluation_metric_columns: dict[tuple[str, str], int] = {}
    for candidate in pairs:
        if candidate.processor == "evaluation_metrics":
            evaluation_metric_columns[("reviewer", candidate.reviewer_block)] = candidate.reviewer_col_idx
            evaluation_metric_columns[("final_annotation", candidate.final_block)] = candidate.final_col_idx

    for row in rows:
        note = display_raw(row.get(col_letter_by_idx[special_cols["notes"]]))
        novelty_final_override_coder_raw = extract_coder_override_from_note(note)
        if novelty_final_override_coder_raw:
            run_warnings.append(
                "Notes column states Novelty/Sensitivity final annotation coding was done by a specific coder; "
                "Coder2 novelty/sensitivity records are assigned to the corresponding neutral coder ID."
            )
            break

    for row in rows:
        paper_id = display_raw(row.get(col_letter_by_idx[special_cols["paper_id"]])).strip()
        if not paper_id:
            continue
        title = display_raw(row.get(col_letter_by_idx.get(special_cols.get("title", -1), ""))).strip()
        review_coder_raw = display_raw(row.get(col_letter_by_idx[special_cols["review_coder"]])).strip()
        final_coder_raw = display_raw(row.get(col_letter_by_idx[special_cols["final_coder"]])).strip()
        review_coder = anonymize_coder_id(review_coder_raw, coder_label_map)
        final_coder = anonymize_coder_id(final_coder_raw, coder_label_map)
        note = display_raw(row.get(col_letter_by_idx[special_cols["notes"]])).strip()
        note_for_output = anonymize_note_text(note, coder_label_map)

        if not review_coder or not final_coder:
            coder_issues.append({"paper_id": paper_id, "issue_type": "missing_coder_identity", "review_coder": review_coder, "final_coder": final_coder, "notes": note_for_output})
        if review_coder and final_coder and review_coder == final_coder:
            coder_issues.append({"paper_id": paper_id, "issue_type": "same_review_and_final_coder", "review_coder": review_coder, "final_coder": final_coder, "notes": note_for_output})
        if note:
            coder_issues.append({"paper_id": paper_id, "issue_type": "notes_recorded", "review_coder": review_coder, "final_coder": final_coder, "notes": note_for_output})

        for pair in pairs:
            ratings = [
                ("reviewer", review_coder, pair.reviewer_col_idx),
                ("final_annotation", final_coder, pair.final_col_idx),
            ]
            for rating_role, coder_id, col_idx in ratings:
                actual_coder = coder_id
                assignment_note = ""
                if rating_role == "final_annotation" and pair.domain == "novelty_and_sensitivity" and novelty_final_override_coder_raw:
                    override_coder = anonymize_coder_id(novelty_final_override_coder_raw, coder_label_map)
                    if coder_id != override_coder:
                        assignment_note = f"Final annotation coder overridden from {coder_id} to {override_coder} for Novelty/Sensitivity per Notes exception."
                    actual_coder = override_coder
                raw = row.get(col_letter_by_idx[col_idx])
                block = pair.reviewer_block if rating_role == "reviewer" else pair.final_block
                metric_col_idx = evaluation_metric_columns.get((rating_role, block))
                metric_raw = row.get(col_letter_by_idx[metric_col_idx]) if metric_col_idx is not None else None
                processed = process_value(
                    pair,
                    raw,
                    paper_id,
                    rating_role,
                    actual_coder,
                    unmatched,
                    evaluation_metrics_value=metric_raw,
                )
                unit_id = f"{paper_id}::{pair.variable_name}"
                base_category = ""
                if pair.measurement_level == "nominal" and missing_value(processed["numeric_value"]):
                    base_category = processed["nominal_value"] if processed["nominal_value"] is not None else processed["normalized_value"]
                base = {
                    "paper_id": paper_id,
                    "title": title,
                    "domain": pair.domain,
                    "subgroup": pair.subgroup,
                    "variable_name": pair.variable_name,
                    "variable_label": pair.variable_label,
                    "rating_role": rating_role,
                    "coder_id": actual_coder,
                    "raw_value": display_raw(raw),
                    "normalized_value": processed["normalized_value"],
                    "category": base_category,
                    "binary_value": np.nan,
                    "numeric_value": processed["numeric_value"],
                    "nominal_value": processed["nominal_value"],
                    "measurement_level": pair.measurement_level,
                    "processor": pair.processor,
                    "source_script": pair.source_script,
                    "parser_note": processed["parser_note"],
                    "assignment_note": assignment_note,
                    "unit_id": unit_id,
                }
                normalized_records.append(base.copy())
                if not missing_value(processed["numeric_value"]):
                    numeric_records.append(base.copy())
                for category, binary in processed["categories"].items():
                    b = np.nan if binary is None else int(binary)
                    assert b in (0, 1) or pd.isna(b)
                    cat_record = {
                        **base,
                        "category": category,
                        "binary_value": b,
                        "numeric_value": np.nan,
                        "nominal_value": None,
                        "unit_id": f"{unit_id}::{category}",
                    }
                    binary_records.append(cat_record)

    mapping_df = pd.DataFrame([pair.__dict__ for pair in pairs])
    codebook_rows = []
    for pair in pairs:
        codebook_rows.append(
            {
                "variable_name": pair.variable_name,
                "domain": pair.domain,
                "subgroup": pair.subgroup,
                "variable_label": pair.variable_label,
                "variable_type": pair.variable_type,
                "measurement_level": pair.measurement_level,
                "processor": pair.processor,
                "category_definitions": describe_processor_categories(pair.processor),
                "invalid_value_rules": "Common invalid markers: blank, dash, N/A, N/D, not given, not reported, unknown, unclear; exceptions documented in processor.",
                "source_script": pair.source_script,
                "results_function": pair.results_function,
                "is_exception": pair.is_exception,
                "is_new_icr_rule": pair.is_new_icr_rule,
            }
        )
    codebook_df = pd.DataFrame(codebook_rows)

    normalized_df = pd.DataFrame(normalized_records)
    binary_df = pd.DataFrame(binary_records)
    numeric_df = pd.DataFrame(numeric_records)
    long_df = normalized_df.copy()
    long_df["row_type"] = "normalized"
    unmatched_df = pd.DataFrame(unmatched)
    coder_issues_df = pd.DataFrame(coder_issues)

    alpha_rows = []
    for (variable_name, category), group in binary_df.groupby(["variable_name", "category"], dropna=False):
        records = group.to_dict("records")
        stats = alpha_records(records, "nominal", "binary_value")
        pos_reviewer = int(group.loc[group["rating_role"] == "reviewer", "binary_value"].fillna(0).sum())
        pos_final = int(group.loc[group["rating_role"] == "final_annotation", "binary_value"].fillna(0).sum())
        alpha_rows.append(
            {
                "variable_name": variable_name,
                "category": category,
                "data_type": "binary_category",
                "measurement_level": "nominal",
                "positive_reviewer_ratings": pos_reviewer,
                "positive_final_ratings": pos_final,
                **stats,
            }
        )

    nominal_candidates = normalized_df[normalized_df["nominal_value"].notna()].copy()
    for variable_name, group in nominal_candidates.groupby("variable_name", dropna=False):
        if variable_name in set(binary_df["variable_name"]):
            pair_type = mapping_df.loc[mapping_df["variable_name"] == variable_name, "variable_type"].iloc[0]
            if pair_type != "nominal":
                continue
        records = group.to_dict("records")
        stats = alpha_records(records, "nominal", "nominal_value")
        alpha_rows.append(
            {
                "variable_name": variable_name,
                "category": "",
                "data_type": "nominal_variable",
                "measurement_level": "nominal",
                "positive_reviewer_ratings": np.nan,
                "positive_final_ratings": np.nan,
                **stats,
            }
        )

    for variable_name, group in numeric_df.groupby("variable_name", dropna=False):
        records = group.to_dict("records")
        stats = alpha_records(records, "interval", "numeric_value")
        alpha_rows.append(
            {
                "variable_name": variable_name,
                "category": "",
                "data_type": "numeric_variable",
                "measurement_level": "interval",
                "positive_reviewer_ratings": np.nan,
                "positive_final_ratings": np.nan,
                **stats,
            }
        )

    alpha_df = pd.DataFrame(alpha_rows).sort_values(["variable_name", "category", "data_type"]).reset_index(drop=True)

    # Descriptive summary across estimable alpha coefficients
    valid_alpha = alpha_df["krippendorff_alpha"].dropna()

    n_estimable_alpha = int(valid_alpha.shape[0])
    median_alpha = float(valid_alpha.median())
    q1_alpha = float(valid_alpha.quantile(0.25))
    q3_alpha = float(valid_alpha.quantile(0.75))
    mean_alpha = float(valid_alpha.mean())

    print("\nUndefined alpha reason counts:")
    print(
        alpha_df.loc[
            alpha_df["krippendorff_alpha"].isna(),
            "undefined_reason"
            ].value_counts(dropna=False)
            )

    raw_rows = []
    for row in rows:
        out = {"excel_row": row["excel_row"]}
        for col in columns:
            header = f"{col['excel_col']}__{col['row1'] or ''}__{col['row2'] or ''}".strip("_")
            out[header] = display_raw(row.get(col["excel_col"]))
        raw_rows.append(out)
    raw_df = pd.DataFrame(raw_rows)

    summary_rows = [
        {
             "metric": "execution_status",
            "value": "success",
            },
        {
            "metric": "studies_processed",
            "value": normalized_df["paper_id"].nunique(),
            },
        {
            "metric": "paired_variables",
            "value": len(pairs),
            },
        
        { 
            "metric": "generated_binary_category_rows",
            "value": len(binary_df),
            },
        { 
            "metric": "generated_binary_categories",
            "value": binary_df[
                ["variable_name", "category"]
                ].drop_duplicates().shape[0],
                 },
        {
             "metric": "unmatched_values_requiring_review",
            "value": len(unmatched_df),
            },
        {
            "metric": "coder_assignment_issues",
            "value": len(coder_issues_df),
            },
        {
            "metric": "variables_or_categories_with_calculable_alpha",
            "value": n_estimable_alpha,
            },
        {
            "metric": "variables_or_categories_with_undefined_alpha",
            "value": int(alpha_df["krippendorff_alpha"].isna().sum()),
            },

        {
            "metric": "median_alpha_across_estimable_coefficients",
            "value": median_alpha,
            },
        {
            "metric": "alpha_first_quartile",
            "value": q1_alpha,
            },
        {
            "metric": "alpha_third_quartile",
            "value": q3_alpha,
            },
        {
            "metric": "mean_alpha_across_estimable_coefficients",
            "value": mean_alpha,
            },

        {"metric": "random_seed", "value": RANDOM_SEED},
        {"metric": "bootstrap_iterations", "value": BOOTSTRAP_ITERATIONS},
    ]

    for warning in sorted(set(run_warnings)):
        summary_rows.append({"metric": "warning", "value": warning})
    summary_df = pd.DataFrame(summary_rows)

    validate_outputs(rows, pairs, normalized_df, binary_df, numeric_df, alpha_df, special_cols, col_letter_by_idx)

    output_dir.mkdir(parents=True, exist_ok=True)
    mapping_path = output_dir / "icr_variable_mapping.csv"
    codebook_path = output_dir / "icr_codebook.csv"
    xlsx_path = output_dir / "ICR_processed.xlsx"
    alpha_path = output_dir / "ICR_alpha_results.csv"
    unmatched_path = output_dir / "ICR_unmatched_values.csv"
    log_path = output_dir / "ICR_run_log.txt"
    method_path = output_dir / "ICR_method_summary.md"
    proposed_changes_path = output_dir / "ICR_proposed_category_changes.csv"

    mapping_df.to_csv(mapping_path, index=False)
    codebook_df.to_csv(codebook_path, index=False)
    alpha_df.to_csv(alpha_path, index=False)
    unmatched_df.to_csv(unmatched_path, index=False)
    pd.DataFrame(columns=["variable_name", "raw_value", "proposed_category", "reason", "status"]).to_csv(proposed_changes_path, index=False)

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        raw_df.to_excel(writer, sheet_name="raw_preserved", index=False)
        long_df.to_excel(writer, sheet_name="long_format", index=False)
        normalized_df.to_excel(writer, sheet_name="normalized", index=False)
        binary_df.to_excel(writer, sheet_name="binary_categories", index=False)
        numeric_df.to_excel(writer, sheet_name="numeric_variables", index=False)
        unmatched_df.to_excel(writer, sheet_name="unmatched_values", index=False)
        coder_issues_df.to_excel(writer, sheet_name="coder_assignment_issues", index=False)
        alpha_df.to_excel(writer, sheet_name="alpha_results", index=False)
        summary_df.to_excel(writer, sheet_name="summary", index=False)

    style_workbook(xlsx_path)
    write_run_log(log_path, workbook_path, pairs, run_warnings, mapping_issues, alpha_df)
    write_method_summary(method_path, normalized_df, mapping_df, alpha_df, coder_issues_df, unmatched_df, bool(novelty_final_override_coder_raw))

    return {
        "mapping": mapping_df,
        "codebook": codebook_df,
        "raw": raw_df,
        "long": long_df,
        "normalized": normalized_df,
        "binary": binary_df,
        "numeric": numeric_df,
        "unmatched": unmatched_df,
        "coder_issues": coder_issues_df,
        "alpha": alpha_df,
        "summary": summary_df,
        "paths": {
            "mapping": mapping_path,
            "codebook": codebook_path,
            "workbook": xlsx_path,
            "alpha": alpha_path,
            "unmatched": unmatched_path,
            "log": log_path,
            "method": method_path,
            "proposed_changes": proposed_changes_path,
        },
    }


def describe_processor_categories(processor: str) -> str:
    examples = {
        "quality_issue": "0=no quality issue; 1=quality issue. Blank/dash/plain No are 0; Yes or substantive concern text is 1.",
        "age_range": "Infants, Toddlers, Pre-schoolers, Grade-schoolers, Teens, Adults, age_range_not_given, multiple_age_groups.",
        "numeric_age": "Continuous numeric; invalid values are missing, never zero-filled.",
        "numeric_count": "Continuous count; invalid values are missing.",
        "recommendations": "20 frozen non-mutually-exclusive recommendation categories plus no_recommendation_reported and other_uncategorized_recommendation.",
        "dataset_name": "Alias-normalized dataset indicators plus dataset_not_given and dataset_other_uncategorized.",
        "data_description": "Non-mutually-exclusive data-source/modality codebook plus data_description_not_given and other_uncategorized_data_description.",
        "best_model": "Exact normalized best model plus broad algorithm-family indicators reused from algorithms used.",
    }
    return examples.get(processor, "See processor/source function columns for category definitions; invalid/missing handling follows source Results logic.")


def validate_outputs(
    rows: list[dict[str, Any]],
    pairs: list[VariablePair],
    normalized_df: pd.DataFrame,
    binary_df: pd.DataFrame,
    numeric_df: pd.DataFrame,
    alpha_df: pd.DataFrame,
    special_cols: dict[str, int],
    col_letter_by_idx: dict[int, str],
) -> None:
    n_studies = len([r for r in rows if display_raw(r.get(col_letter_by_idx[special_cols["paper_id"]])).strip()])
    assert normalized_df["paper_id"].nunique() == n_studies, "Number of studies changed during normalization."
    assert len(pairs) == 67, f"Expected 67 paired variables, found {len(pairs)}."
    assert set(binary_df["binary_value"].dropna().unique()).issubset({0, 1}), "Binary category values must be only 0/1/missing."
    if not numeric_df.empty:
        missing_raw_numeric = numeric_df[numeric_df["raw_value"].str.lower().isin(INVALID_TEXT_VALUES)]
        assert missing_raw_numeric.empty, "Invalid numeric raw values entered numeric output."
    assert recommendation_categories("cross-validation").get("external_or_clinical_validation", 0) == 0, "Plain cross-validation categorized as external validation."
    assert not alpha_df.empty, "Alpha table is empty."
    expected_roles = {"reviewer", "final_annotation"}
    assert set(normalized_df["rating_role"].unique()) == expected_roles, "Missing reviewer/final records."
    role_counts = normalized_df.groupby(["paper_id", "variable_name"])["rating_role"].nunique()
    assert int((role_counts == 2).sum()) == len(role_counts), "Some study-variable pairs do not have two rating roles."


def style_workbook(path: Path) -> None:
    wb = openpyxl.load_workbook(path)
    header_fill = "1F4E79"
    header_font = openpyxl.styles.Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor=header_fill)
            cell.font = header_font
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
        for col_cells in ws.columns:
            col_letter = get_column_letter(col_cells[0].column)
            max_len = 0
            for cell in col_cells[:200]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(value), 80))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 48)
        ws.sheet_view.showGridLines = False
    wb.save(path)


def write_run_log(path: Path, workbook_path: Path, pairs: list[VariablePair], warnings: list[str], mapping_issues: list[dict[str, Any]], alpha_df: pd.DataFrame) -> None:
    lines = []
    lines.append("ICR pipeline run log")
    lines.append("Execution timestamp: omitted to keep rerun outputs deterministic")
    lines.append(f"Execution status: success")
    lines.append(f"Python executable: {Path(sys.executable).name}")
    lines.append(f"Python version: {sys.version.replace(os.linesep, ' ')}")
    lines.append(f"Platform: {platform.platform()}")
    for module_name, module in [("pandas", pd), ("numpy", np), ("openpyxl", openpyxl)]:
        lines.append(f"{module_name}: {getattr(module, '__version__', 'unknown')}")
    lines.append(f"Krippendorff alpha implementation: internal pairwise nominal/interval implementation")
    lines.append(f"Random seed: {RANDOM_SEED}")
    lines.append(f"Bootstrap iterations: {BOOTSTRAP_ITERATIONS}")
    lines.append("")
    lines.append("Input files:")
    for p in [workbook_path, TYPE_CLASSIFICATION]:
        status = "present" if p.exists() else "missing"
        lines.append(
            f"- {repo_relative_path(p)} | {status} | "
            f"sha256={source_hash(p)}"
            )
    lines.append("")
    lines.append("Finalized Results scripts used/refactored:")
    for label, p in SOURCE_SCRIPT_MANIFEST:
        status = "present" if p.exists() else "missing"
        actual_note = "" if p.name == label else f" | actual_filename={p.name}"
        lines.append(f"- {label} | {repo_relative_path(p)} | {status} | sha256={source_hash(p)}{actual_note}")
    lines.append("")
    lines.append(f"Paired variables: {len(pairs)}")
    lines.append(f"Calculable alpha rows: {int(alpha_df['krippendorff_alpha'].notna().sum())}")
    lines.append(f"Undefined alpha rows: {int(alpha_df['krippendorff_alpha'].isna().sum())}")
    lines.append("")
    lines.append("Warnings and manual mappings:")
    if warnings:
        for w in sorted(set(warnings)):
            lines.append(f"- {w}")
    else:
        lines.append("- None")
    if mapping_issues:
        for issue in mapping_issues:
            lines.append(f"- Mapping issue: {issue}")
    lines.append("- Dataset aliases and recommendation regexes are frozen in run_icr_pipeline.py.")
    lines.append("- Standalone Results scripts were not imported because they execute analyses at import time; needed logic was refactored with source-script labels preserved.")
    lines.append("- The rq5_.py script is hashed and logged for the manuscript Results set; no ICR worksheet variable maps uniquely to an RQ5-only output.")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_method_summary(
    path: Path,
    normalized_df: pd.DataFrame,
    mapping_df: pd.DataFrame,
    alpha_df: pd.DataFrame,
    coder_issues_df: pd.DataFrame,
    unmatched_df: pd.DataFrame,
    novelty_override: bool,
) -> None:
    n_studies = normalized_df["paper_id"].nunique()
    n_pairs = mapping_df.shape[0]
    levels = mapping_df.groupby("measurement_level")["variable_name"].nunique().to_dict()
    lines = [
        "# ICR Method Summary",
        "",
        f"The pipeline processed {n_studies} studies and {n_pairs} reviewer-to-final variable pairs from `ICR.xlsx`.",
        "",
        "Coder1 columns were treated as the independent reviewer values and assigned to `Review Coder`. "
        "Coder2 columns were treated as final annotation values and assigned to `Final annotation Coder`. "
        + ("A Notes exception assigned Novelty/Sensitivity Coder2 records to a neutral coder ID." if novelty_override else "No coder-identity override was applied."),
        "",
        "Krippendorff's alpha was used because coder pairs vary across studies; the reliability matrix can preserve actual coder identities with missing cells for coder-study combinations that were not observed.",
        "",
        "Free text was lowercased, whitespace-normalized, punctuation-normalized, and matched to frozen regex/category dictionaries refactored from the Results scripts or specified by the ICR PDF.",
        "",
        f"Measurement levels used: nominal={levels.get('nominal', 0)}, interval={levels.get('interval', 0)}. Ordinal alpha was not used because no included variable had a genuine ordered codebook.",
        "",
        f"Unmatched substantive values exported for review: {len(unmatched_df)}.",
        f"Coder assignment/notes rows exported: {len(coder_issues_df)}.",
        f"Calculable alpha rows: {int(alpha_df['krippendorff_alpha'].notna().sum())}. Undefined alpha rows: {int(alpha_df['krippendorff_alpha'].isna().sum())}.",
        "",
        "Undefined alpha rows are reported as NA with agreement/prevalence information and should not be interpreted as poor reliability.",
    ]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    workbook_path = Path(sys.argv[1]).expanduser() if len(sys.argv) > 1 else DEFAULT_WORKBOOK
    if not workbook_path.exists():
        raise FileNotFoundError(workbook_path)
    result = build_outputs(workbook_path, OUTPUT_DIR)
    summary = result["summary"]
    for _, row in summary.iterrows():
        print(f"{row['metric']}: {row['value']}")
    print("Created files:")
    for key, value in result["paths"].items():
        print(f"- {key}: {value}")


if __name__ == "__main__":
    main()
