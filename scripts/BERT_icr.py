#!/usr/bin/env python3
"""BERT-based semantic-category ICR analysis for the ICR workbook.

This script treats the input workbook as read-only. It detects the same
Coder1/Coder2 paired blocks used in the prior ICR analysis, builds shared
data-driven semantic categories from pooled Coder1+Coder2 text for each
eligible free-text variable, and calculates nominal Krippendorff's alpha.

Install commands, if dependencies are missing:
    python3 -m pip install sentence-transformers

Default model:
    sentence-transformers/all-MiniLM-L6-v2
"""

from __future__ import annotations

import argparse
import csv
import importlib.util
import json
import math
import os
import platform
import random
import re
import sys
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter
from sklearn.cluster import AgglomerativeClustering
from sklearn.metrics.pairwise import cosine_similarity

_ORIGINAL_FIND_SPEC = importlib.util.find_spec
_OPTIONAL_BACKENDS_TO_HIDE = {
    "torchvision",
    "tensorflow",
    "tensorflow_probability",
    "keras",
    "tf_keras",
}


def _hide_broken_optional_transformers_backends(name: str, *args: Any, **kwargs: Any):
    """Prevent optional vision/TF backends from breaking text-only embeddings.

    This Python environment has a broken optional torchvision/lzma path and a
    TensorFlow build compiled against an older NumPy ABI. Sentence embeddings
    only require the PyTorch text stack, so hiding those optional backends keeps
    transformers on the intended code path without uninstalling packages.
    """
    if name in _OPTIONAL_BACKENDS_TO_HIDE or any(name.startswith(prefix + ".") for prefix in _OPTIONAL_BACKENDS_TO_HIDE):
        return None
    return _ORIGINAL_FIND_SPEC(name, *args, **kwargs)


importlib.util.find_spec = _hide_broken_optional_transformers_backends
try:
    from sentence_transformers import SentenceTransformer
except Exception as exc:  # pragma: no cover - shown as a runtime dependency error
    SentenceTransformer = None
    SENTENCE_TRANSFORMER_IMPORT_ERROR = repr(exc)
else:
    SENTENCE_TRANSFORMER_IMPORT_ERROR = ""


RANDOM_SEED = 20260709
DEFAULT_MODEL_NAME = "sentence-transformers/all-MiniLM-L6-v2"
PRIMARY_DISTANCE_THRESHOLD = 0.35
SENSITIVITY_THRESHOLDS = [0.25, 0.45]
MIN_UNIQUE_FOR_CLUSTERING = 3
RARE_CATEGORY_MAX_ASSIGNMENTS = 1
SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parent
DEFAULT_INPUT_WORKBOOK = REPO_ROOT / "data" / "ICR.xlsx"
DEFAULT_OUTPUT_DIR = REPO_ROOT / "output" / "bert_icr_results"


INVALID_TEXT_VALUES = {
    "",
    "-",
    "--",
    "---",
    "nan",
    "none",
    "no information",
    "na",
    "n/a",
    "n.a",
    "n.a.",
    "n.d",
    "n.d.",
    "nd",
    "n/d",
    "nr",
    "n.r",
    "n.r.",
    "n/r",
    "not given",
    "not reported",
    "not specified",
    "not stated",
    "not mentioned",
    "not available",
    "not provided",
    "not applicable",
    "unknown",
    "unkown",
    "unclear",
    "no information",
    "not clear",
    "not explicitly stated",
    "none specified",
    "not givven",
}

BLOCK_PAIRS = [
    ("Quality_Coder1", "Quality_Coder2"),
    ("Participants_Coder1", "Participants_Coder2"),
    ("Behaviors_Coder1", "Behaviors_Coder2"),
    ("AI_Coder1", "AI_Coder2"),
    ("study_coder1", "study_coder2"),
    ("Novelty_and_sensitivity_Coder1", "Novelty_and_sensitivity_Coder2"),
]


def extract_coder_override_from_note(note: str) -> str:
    match = re.search(r"novelty.*sensitivity.*done by\s+([A-Za-z][A-Za-z0-9_.-]*)", note, flags=re.IGNORECASE)
    return match.group(1).strip() if match else ""


def anonymize_coder_id(raw_coder: Any, coder_label_map: dict[str, str]) -> str:
    raw = display_raw(raw_coder).strip()
    if not raw:
        return ""
    key = normalize_response(raw) or raw.lower()
    if key not in coder_label_map:
        coder_label_map[key] = f"Coder{len(coder_label_map) + 1}"
    return coder_label_map[key]


@dataclass(frozen=True)
class VariablePair:
    domain: str
    subgroup: str
    variable_label: str
    variable_name: str
    reviewer_col_idx: int
    final_col_idx: int
    reviewer_excel_col: str
    final_excel_col: str
    reviewer_block: str
    final_block: str
    position_in_block: int
    prior_variable_type: str
    free_text_eligible: bool
    eligibility_reason: str


def slugify(value: str) -> str:
    text = normalize_label(value)
    text = text.replace("#", "number")
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return re.sub(r"_+", "_", text).strip("_") or "blank"


def normalize_label(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u00a0", " ")
    text = re.sub(r"\s+", " ", text).strip().lower()
    return text


def display_raw(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value)


def normalize_response(value: Any) -> str | None:
    """Minimal text normalization used for duplicates and exact-match checks."""
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    text = str(value).replace("\u00a0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    if text.lower() in INVALID_TEXT_VALUES:
        return None
    return text.lower()


def text_for_embedding(value: Any) -> str | None:
    """Whitespace-normalized text for BERT; words/content are otherwise preserved."""
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    text = str(value).replace("\u00a0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    if text.lower() in INVALID_TEXT_VALUES:
        return None
    return text


def get_block_domain(block_name: str) -> str:
    base = re.sub(r"_?coder[12]$", "", block_name, flags=re.IGNORECASE).strip()
    return slugify(base)


def participant_subgroup(position: int) -> str:
    if 0 <= position <= 8:
        return "asd"
    if 9 <= position <= 17:
        return "neurotypical"
    if 18 <= position <= 27:
        return "other_diagnosis"
    return ""


def canonical_variable_label(raw_label: str, domain: str, subgroup: str, position: int) -> str:
    label = (raw_label or "").strip()
    normalized = normalize_label(label)
    replacements = {
        "# nd participants": "# neurotypical participants",
        "assesment method": "Assessment method",
        "other assessment": "Other assessments",
        "fusion tecnique": "fusion technique",
        "balancing/unbiasing tecnique": "balancing/unbiasing technique",
        "reccomendations": "recommendations",
    }
    if normalized in replacements:
        label = replacements[normalized]
    if domain == "participants":
        return f"{subgroup or f'position_{position}'} {label}".strip()
    return label


def prior_variable_type(domain: str, raw_label: str) -> str:
    label = normalize_label(raw_label)
    if domain == "quality":
        return "binary_quality"
    if label.startswith("#") or label in {"mean age", "std age", "best performance"}:
        return "numeric"
    if label == "range age":
        return "structured_age_range"
    if label.startswith("match in"):
        return "binary_match"
    if domain == "behaviors" and label in {"gaze", "speech", "motor"}:
        return "binary_modality"
    if label in {"is the data open source?", "open source code"}:
        return "access_nominal"
    return "free_text_candidate"


def looks_like_numeric_or_range(values: list[str]) -> bool:
    if not values:
        return False
    numeric_like = 0
    for value in values:
        v = value.strip().lower()
        if re.fullmatch(r"[0-9]+(?:\.[0-9]+)?", v):
            numeric_like += 1
        elif re.fullmatch(r"[0-9]+(?:\.[0-9]+)?\s*[-/]\s*[0-9]+(?:\.[0-9]+)?(?:\s*(?:months?|years?|yrs?|y/o)?)?", v):
            numeric_like += 1
        elif re.fullmatch(r"\d{4}-\d{1,2}-\d{1,2}", v):
            numeric_like += 1
    return numeric_like / max(len(values), 1) >= 0.8


def is_simple_binary(values: list[str]) -> bool:
    if not values:
        return False
    yes_no_like = {
        "yes",
        "no",
        "y",
        "n",
        "true",
        "false",
        "0",
        "1",
        "yes.",
        "no.",
    }
    unique = {v.strip().lower() for v in values}
    return len(unique) <= 4 and unique.issubset(yes_no_like)


def read_workbook_model(workbook_path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[VariablePair], dict[str, int], list[dict[str, Any]]]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=False)
    ws = wb["Sheet1"]

    merged_map: dict[tuple[int, int], Any] = {}
    for rg in ws.merged_cells.ranges:
        val = ws.cell(rg.min_row, rg.min_col).value
        for r in range(rg.min_row, rg.max_row + 1):
            for c in range(rg.min_col, rg.max_col + 1):
                merged_map[(r, c)] = val

    def cell_value(row: int, col: int) -> Any:
        v = ws.cell(row, col).value
        return merged_map.get((row, col), v) if v is None else v

    columns = []
    for c in range(1, ws.max_column + 1):
        columns.append(
            {
                "col_idx": c,
                "excel_col": get_column_letter(c),
                "row1": cell_value(1, c),
                "row2": cell_value(2, c),
            }
        )

    rows: list[dict[str, Any]] = []
    for r in range(3, ws.max_row + 1):
        row = {"excel_row": r}
        for col in columns:
            row[col["excel_col"]] = ws.cell(r, col["col_idx"]).value
        rows.append(row)

    idx_by_block: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for col in columns:
        if col["row1"] not in {None, "Paper_id", "Title", "Coder", "Notes"}:
            idx_by_block[str(col["row1"])].append(col)

    special_cols: dict[str, int] = {}
    for col in columns:
        row1 = normalize_label(col["row1"])
        row2 = normalize_label(col["row2"])
        if row1 == "paper_id":
            special_cols["paper_id"] = col["col_idx"]
        elif row1 == "title":
            special_cols["title"] = col["col_idx"]
        elif row2 == "review coder":
            special_cols["review_coder"] = col["col_idx"]
        elif row2 == "final annotation coder":
            special_cols["final_coder"] = col["col_idx"]
        elif row1 == "notes":
            special_cols["notes"] = col["col_idx"]

    mapping_issues: list[dict[str, Any]] = []
    variable_pairs: list[VariablePair] = []
    col_letter_by_idx = {col["col_idx"]: col["excel_col"] for col in columns}

    data_rows = [
        row
        for row in rows
        if normalize_response(row.get(col_letter_by_idx.get(special_cols.get("paper_id", -1), ""))) is not None
    ]

    for reviewer_block, final_block in BLOCK_PAIRS:
        reviewer_cols = idx_by_block[reviewer_block]
        final_cols = idx_by_block[final_block]
        if len(reviewer_cols) != len(final_cols):
            mapping_issues.append(
                {
                    "issue_type": "block_length_mismatch",
                    "reviewer_block": reviewer_block,
                    "final_block": final_block,
                    "reviewer_n": len(reviewer_cols),
                    "final_n": len(final_cols),
                }
            )
        for pos, (rcol, fcol) in enumerate(zip(reviewer_cols, final_cols)):
            domain = get_block_domain(reviewer_block)
            subgroup = participant_subgroup(pos) if domain == "participants" else ""
            raw_label = str(rcol.get("row2") or fcol.get("row2") or "")
            canonical_label = canonical_variable_label(raw_label, domain, subgroup, pos)
            vtype = prior_variable_type(domain, raw_label)
            variable_name = "__".join([x for x in [domain, subgroup, slugify(raw_label)] if x])

            paired_values: list[str] = []
            for row in data_rows:
                rv = normalize_response(row.get(rcol["excel_col"]))
                fv = normalize_response(row.get(fcol["excel_col"]))
                if rv is not None:
                    paired_values.append(rv)
                if fv is not None:
                    paired_values.append(fv)

            eligible = True
            reason = "eligible_free_text"
            if vtype != "free_text_candidate":
                eligible = False
                reason = f"excluded_{vtype}"
            elif looks_like_numeric_or_range(paired_values):
                eligible = False
                reason = "excluded_numeric_or_structured_range_values"
            elif is_simple_binary(paired_values):
                eligible = False
                reason = "excluded_simple_binary_values"

            variable_pairs.append(
                VariablePair(
                    domain=domain,
                    subgroup=subgroup,
                    variable_label=canonical_label,
                    variable_name=variable_name,
                    reviewer_col_idx=rcol["col_idx"],
                    final_col_idx=fcol["col_idx"],
                    reviewer_excel_col=rcol["excel_col"],
                    final_excel_col=fcol["excel_col"],
                    reviewer_block=reviewer_block,
                    final_block=final_block,
                    position_in_block=pos,
                    prior_variable_type=vtype,
                    free_text_eligible=eligible,
                    eligibility_reason=reason,
                )
            )

    return data_rows, columns, variable_pairs, special_cols, mapping_issues


def krippendorff_nominal_alpha(unit_to_values: dict[str, list[Any]]) -> tuple[float | None, str, float | None]:
    observed_pairs = 0
    observed_disagreements = 0
    all_values: list[Any] = []
    exact_agreements = 0
    exact_total = 0

    for values in unit_to_values.values():
        clean = [v for v in values if v is not None and not (isinstance(v, float) and math.isnan(v))]
        all_values.extend(clean)
        if len(clean) >= 2:
            exact_total += 1
            exact_agreements += int(clean[0] == clean[1])
        for i in range(len(clean)):
            for j in range(i + 1, len(clean)):
                observed_pairs += 1
                observed_disagreements += int(clean[i] != clean[j])

    if observed_pairs == 0:
        return None, "too_few_paired_observations", None
    if len(all_values) < 2:
        return None, "too_few_values", exact_agreements / exact_total if exact_total else None

    total_pairs = 0
    expected_disagreements = 0
    for i in range(len(all_values)):
        for j in range(i + 1, len(all_values)):
            total_pairs += 1
            expected_disagreements += int(all_values[i] != all_values[j])

    exact_agreement = exact_agreements / exact_total if exact_total else None
    if total_pairs == 0 or expected_disagreements == 0:
        return None, "no_variation", exact_agreement

    do = observed_disagreements / observed_pairs
    de = expected_disagreements / total_pairs
    if de == 0:
        return None, "no_variation", exact_agreement
    return 1.0 - (do / de), "valid", exact_agreement


def medoid_response(texts: list[str], embeddings: np.ndarray) -> str:
    if len(texts) == 1:
        return texts[0]
    sims = cosine_similarity(embeddings)
    idx = int(np.argmax(sims.mean(axis=1)))
    return texts[idx]


def cluster_unique_responses(
    normalized_to_embedding_text: dict[str, str],
    embeddings: np.ndarray,
    distance_threshold: float,
    method_label: str,
) -> tuple[dict[str, str], list[dict[str, Any]], str]:
    normalized_values = list(normalized_to_embedding_text.keys())
    display_values = [normalized_to_embedding_text[n] for n in normalized_values]
    n_unique = len(normalized_values)

    if n_unique == 0:
        return {}, [], "no_valid_responses"

    if n_unique < MIN_UNIQUE_FOR_CLUSTERING:
        raw_labels = np.arange(n_unique)
        rule = f"exact_normalized_fallback_n_unique_lt_{MIN_UNIQUE_FOR_CLUSTERING}"
    else:
        clustering = AgglomerativeClustering(
            n_clusters=None,
            metric="cosine",
            linkage="average",
            distance_threshold=distance_threshold,
        )
        raw_labels = clustering.fit_predict(embeddings)
        rule = f"{method_label}_distance_threshold_{distance_threshold}"

    cluster_to_indices: dict[int, list[int]] = defaultdict(list)
    for idx, raw_label in enumerate(raw_labels):
        cluster_to_indices[int(raw_label)].append(idx)

    ordered_clusters = sorted(
        cluster_to_indices.items(),
        key=lambda item: (
            medoid_response([display_values[i] for i in item[1]], embeddings[item[1]]).lower(),
            min(item[1]),
        ),
    )

    norm_to_category: dict[str, str] = {}
    summaries: list[dict[str, Any]] = []
    for cat_num, (raw_label, indices) in enumerate(ordered_clusters, start=1):
        category_id = f"Category_{cat_num}"
        cluster_texts = [display_values[i] for i in indices]
        cluster_norms = [normalized_values[i] for i in indices]
        cluster_embeddings = embeddings[indices]
        representative = medoid_response(cluster_texts, cluster_embeddings)
        examples = " | ".join(sorted(cluster_texts, key=lambda x: (len(x), x.lower()))[:5])
        for norm in cluster_norms:
            norm_to_category[norm] = category_id
        summaries.append(
            {
                "category_id": category_id,
                "representative_response": representative,
                "category_description": representative,
                "n_unique_responses": len(indices),
                "example_responses": examples,
                "cluster_rule": rule,
                "raw_cluster_label": raw_label,
            }
        )
    return norm_to_category, summaries, rule


def alpha_for_labels(paired_rows: list[dict[str, Any]], label_key_1: str, label_key_2: str, unit_prefix: str) -> tuple[float | None, str, float | None]:
    unit_to_values = {}
    for row in paired_rows:
        unit = f"{unit_prefix}::{row['paper_id']}"
        unit_to_values[unit] = [row[label_key_1], row[label_key_2]]
    return krippendorff_nominal_alpha(unit_to_values)


def build_assignments_for_variable(
    pair: VariablePair,
    rows: list[dict[str, Any]],
    col_letter_by_idx: dict[int, str],
    special_cols: dict[str, int],
    norm_to_primary_category: dict[str, str],
    category_summary_lookup: dict[str, dict[str, Any]],
    novelty_final_override_coder_raw: str,
    coder_label_map: dict[str, str],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], int, int, int]:
    paper_col = col_letter_by_idx[special_cols["paper_id"]]
    title_col = col_letter_by_idx.get(special_cols.get("title", -1), "")
    review_col = col_letter_by_idx[special_cols["review_coder"]]
    final_col = col_letter_by_idx[special_cols["final_coder"]]
    reviewer_value_col = col_letter_by_idx[pair.reviewer_col_idx]
    final_value_col = col_letter_by_idx[pair.final_col_idx]

    assignment_rows: list[dict[str, Any]] = []
    paired_rows: list[dict[str, Any]] = []
    paired_valid = 0
    reviewer_valid = 0
    final_valid = 0

    for row in rows:
        paper_id = display_raw(row.get(paper_col)).strip()
        if not paper_id:
            continue
        title = display_raw(row.get(title_col)).strip() if title_col else ""
        review_coder = anonymize_coder_id(row.get(review_col), coder_label_map)
        final_coder = anonymize_coder_id(row.get(final_col), coder_label_map)
        if pair.domain == "novelty_and_sensitivity" and novelty_final_override_coder_raw:
            final_coder = anonymize_coder_id(novelty_final_override_coder_raw, coder_label_map)

        raw_reviewer = row.get(reviewer_value_col)
        raw_final = row.get(final_value_col)
        norm_reviewer = normalize_response(raw_reviewer)
        norm_final = normalize_response(raw_final)
        embed_reviewer = text_for_embedding(raw_reviewer)
        embed_final = text_for_embedding(raw_final)
        reviewer_cat = norm_to_primary_category.get(norm_reviewer) if norm_reviewer is not None else None
        final_cat = norm_to_primary_category.get(norm_final) if norm_final is not None else None
        reviewer_valid += int(norm_reviewer is not None)
        final_valid += int(norm_final is not None)
        if norm_reviewer is not None and norm_final is not None:
            paired_valid += 1
            paired_rows.append(
                {
                    "paper_id": paper_id,
                    "reviewer_category": reviewer_cat,
                    "final_category": final_cat,
                    "reviewer_normalized": norm_reviewer,
                    "final_normalized": norm_final,
                }
            )

        for rating_role, coder_id, raw_value, norm_value, embed_text, category_id in [
            ("reviewer", review_coder, raw_reviewer, norm_reviewer, embed_reviewer, reviewer_cat),
            ("final_annotation", final_coder, raw_final, norm_final, embed_final, final_cat),
        ]:
            category_info = category_summary_lookup.get(category_id or "", {})
            assignment_rows.append(
                {
                    "paper_id": paper_id,
                    "title": title,
                    "domain": pair.domain,
                    "subgroup": pair.subgroup,
                    "variable_name": pair.variable_name,
                    "variable_label": pair.variable_label,
                    "rating_role": rating_role,
                    "coder_id": coder_id,
                    "raw_value": display_raw(raw_value),
                    "normalized_response": norm_value,
                    "embedding_text": embed_text,
                    "bert_category": category_id,
                    "category_description": category_info.get("category_description"),
                    "category_representative_response": category_info.get("representative_response"),
                    "is_valid_response": norm_value is not None,
                    "reviewer_col": pair.reviewer_excel_col,
                    "final_col": pair.final_excel_col,
                }
            )

    return assignment_rows, paired_rows, paired_valid, reviewer_valid, final_valid


def summarize_overall(variable_df: pd.DataFrame, run_notes: list[str]) -> pd.DataFrame:
    valid_alpha = pd.to_numeric(variable_df["krippendorff_alpha"], errors="coerce").dropna()
    eligible = variable_df[variable_df["eligible"] == True]
    summary_rows = [
        ("analysis_type", "BERT semantic-category ICR; not identical to predefined manual-category ICR"),
        ("eligible_variables", int(len(eligible))),
        ("variables_with_valid_alpha", int(len(valid_alpha))),
        ("median_krippendorff_alpha", float(valid_alpha.median()) if len(valid_alpha) else None),
        ("mean_krippendorff_alpha", float(valid_alpha.mean()) if len(valid_alpha) else None),
        ("min_krippendorff_alpha", float(valid_alpha.min()) if len(valid_alpha) else None),
        ("max_krippendorff_alpha", float(valid_alpha.max()) if len(valid_alpha) else None),
        ("iqr_krippendorff_alpha", float(valid_alpha.quantile(0.75) - valid_alpha.quantile(0.25)) if len(valid_alpha) else None),
        ("random_seed", RANDOM_SEED),
        ("primary_model", DEFAULT_MODEL_NAME),
        ("primary_clustering", f"agglomerative_average_cosine_distance_threshold_{PRIMARY_DISTANCE_THRESHOLD}"),
        ("missing_data_rule", "Variable-level alpha uses rows where both coders have valid non-missing responses."),
    ]
    for note in run_notes:
        summary_rows.append(("warning_or_note", note))
    return pd.DataFrame(summary_rows, columns=["metric", "value"])


def package_versions(model_name: str) -> dict[str, Any]:
    versions: dict[str, Any] = {
        "python": sys.version.replace(os.linesep, " "),
        "platform": platform.platform(),
        "numpy": getattr(np, "__version__", "unknown"),
        "pandas": getattr(pd, "__version__", "unknown"),
        "openpyxl": getattr(openpyxl, "__version__", "unknown"),
    }
    try:
        import sklearn

        versions["sklearn"] = getattr(sklearn, "__version__", "unknown")
    except Exception:
        versions["sklearn"] = "unavailable"
    try:
        import sentence_transformers

        versions["sentence_transformers"] = getattr(sentence_transformers, "__version__", "unknown")
    except Exception as exc:
        versions["sentence_transformers"] = f"unavailable: {exc!r}"
    try:
        import torch

        versions["torch"] = getattr(torch, "__version__", "unknown")
    except Exception as exc:
        versions["torch"] = f"unavailable: {exc!r}"
    versions["model_name"] = model_name
    return versions


def unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    parent = path.parent
    i = 1
    while True:
        candidate = parent / f"{stem}_{i}{suffix}"
        if not candidate.exists():
            return candidate
        i += 1


def write_csv(df: pd.DataFrame, path: Path) -> Path:
    out = unique_path(path)
    df.to_csv(out, index=False, quoting=csv.QUOTE_MINIMAL)
    return out


def run_analysis(input_workbook: Path, output_dir: Path, model_name: str = DEFAULT_MODEL_NAME) -> dict[str, Any]:
    random.seed(RANDOM_SEED)
    np.random.seed(RANDOM_SEED)
    os.environ.setdefault("TOKENIZERS_PARALLELISM", "false")

    output_dir.mkdir(parents=True, exist_ok=True)
    if SentenceTransformer is None:
        raise RuntimeError(
            "sentence-transformers is required for this BERT analysis. "
            f"Import error: {SENTENCE_TRANSFORMER_IMPORT_ERROR}. "
            "Install with: python3 -m pip install sentence-transformers"
        )

    rows, columns, pairs, special_cols, mapping_issues = read_workbook_model(input_workbook)
    required_special = {"paper_id", "review_coder", "final_coder", "notes"}
    missing_special = required_special - set(special_cols)
    if missing_special:
        raise ValueError(f"Missing required workbook columns: {sorted(missing_special)}")

    col_letter_by_idx = {col["col_idx"]: col["excel_col"] for col in columns}
    note_col = col_letter_by_idx[special_cols["notes"]]
    novelty_final_override_coder_raw = ""
    for row in rows:
        novelty_final_override_coder_raw = extract_coder_override_from_note(display_raw(row.get(note_col)))
        if novelty_final_override_coder_raw:
            break

    run_notes: list[str] = []
    if mapping_issues:
        run_notes.append(f"Mapping issues detected: {mapping_issues}")
    if novelty_final_override_coder_raw:
        run_notes.append("Notes exception applied: Novelty/Sensitivity final-annotation records assigned to a neutral coder ID for audit columns.")

    print("\nDETECTED CODER-COLUMN PAIRS")
    print("===========================")
    for pair in pairs:
        print(
            f"{pair.variable_name}: {pair.reviewer_excel_col} ({pair.reviewer_block}) "
            f"<-> {pair.final_excel_col} ({pair.final_block}); eligible={pair.free_text_eligible}; reason={pair.eligibility_reason}"
        )

    print(f"\nLoading BERT sentence embedding model: {model_name}")
    model = SentenceTransformer(model_name)

    variable_rows: list[dict[str, Any]] = []
    all_assignment_rows: list[dict[str, Any]] = []
    coder_label_map: dict[str, str] = {}
    coder_label_map: dict[str, str] = {}
    all_category_rows: list[dict[str, Any]] = []
    excluded_rows: list[dict[str, Any]] = []
    sensitivity_rows: list[dict[str, Any]] = []
    detected_pair_rows = [asdict(pair) for pair in pairs]

    paper_count = len(rows)
    for pair in pairs:
        reviewer_col = col_letter_by_idx[pair.reviewer_col_idx]
        final_col = col_letter_by_idx[pair.final_col_idx]

        pooled: dict[str, str] = {}
        reviewer_valid = 0
        final_valid = 0
        paired_valid_preview = 0
        for row in rows:
            raw_r = row.get(reviewer_col)
            raw_f = row.get(final_col)
            norm_r = normalize_response(raw_r)
            norm_f = normalize_response(raw_f)
            emb_r = text_for_embedding(raw_r)
            emb_f = text_for_embedding(raw_f)
            if norm_r is not None and emb_r is not None:
                pooled.setdefault(norm_r, emb_r)
                reviewer_valid += 1
            if norm_f is not None and emb_f is not None:
                pooled.setdefault(norm_f, emb_f)
                final_valid += 1
            paired_valid_preview += int(norm_r is not None and norm_f is not None)

        base_result = {
            "variable_name": pair.variable_name,
            "variable_label": pair.variable_label,
            "domain": pair.domain,
            "subgroup": pair.subgroup,
            "reviewer_col": pair.reviewer_excel_col,
            "final_col": pair.final_excel_col,
            "total_rows": paper_count,
            "paired_valid_rows": paired_valid_preview,
            "coder1_valid_rows": reviewer_valid,
            "coder2_valid_rows": final_valid,
            "unique_raw_responses": len(pooled),
            "eligible": pair.free_text_eligible,
            "eligibility_reason": pair.eligibility_reason,
            "n_generated_categories": None,
            "krippendorff_alpha": None,
            "alpha_status": "not_run",
            "alpha_reason": None,
            "exact_agreement": None,
            "primary_cluster_rule": None,
        }

        if not pair.free_text_eligible:
            row = {**base_result, "alpha_status": "excluded", "alpha_reason": pair.eligibility_reason}
            variable_rows.append(row)
            excluded_rows.append(row)
            continue

        if paired_valid_preview < 2:
            row = {**base_result, "alpha_status": "undefined", "alpha_reason": "too_few_paired_valid_rows"}
            variable_rows.append(row)
            excluded_rows.append(row)
            continue

        if len(pooled) == 0:
            row = {**base_result, "alpha_status": "undefined", "alpha_reason": "no_valid_responses"}
            variable_rows.append(row)
            excluded_rows.append(row)
            continue

        normalized_values = list(pooled.keys())
        display_values = [pooled[n] for n in normalized_values]
        embeddings = model.encode(display_values, normalize_embeddings=True, show_progress_bar=False)
        embeddings = np.asarray(embeddings)
        norm_to_category, category_summaries, cluster_rule = cluster_unique_responses(
            pooled,
            embeddings,
            distance_threshold=PRIMARY_DISTANCE_THRESHOLD,
            method_label="primary_agglomerative_average_cosine",
        )

        category_lookup = {row["category_id"]: row for row in category_summaries}
        assignment_rows, paired_rows, paired_valid, reviewer_valid, final_valid = build_assignments_for_variable(
            pair=pair,
            rows=rows,
            col_letter_by_idx=col_letter_by_idx,
            special_cols=special_cols,
            norm_to_primary_category=norm_to_category,
            category_summary_lookup=category_lookup,
            novelty_final_override_coder_raw=novelty_final_override_coder_raw,
            coder_label_map=coder_label_map,
        )

        alpha, status, exact_agreement = alpha_for_labels(
            paired_rows,
            "reviewer_category",
            "final_category",
            pair.variable_name,
        )

        category_counts = Counter()
        role_counts = defaultdict(Counter)
        for assignment in assignment_rows:
            cat = assignment.get("bert_category")
            if cat:
                category_counts[cat] += 1
                role_counts[cat][assignment["rating_role"]] += 1

        for summary in category_summaries:
            cat = summary["category_id"]
            all_category_rows.append(
                {
                    "variable_name": pair.variable_name,
                    "variable_label": pair.variable_label,
                    "domain": pair.domain,
                    "subgroup": pair.subgroup,
                    **summary,
                    "n_total_assignments": category_counts.get(cat, 0),
                    "n_reviewer_assignments": role_counts[cat].get("reviewer", 0),
                    "n_final_assignments": role_counts[cat].get("final_annotation", 0),
                }
            )

        all_assignment_rows.extend(assignment_rows)
        variable_rows.append(
            {
                **base_result,
                "paired_valid_rows": paired_valid,
                "coder1_valid_rows": reviewer_valid,
                "coder2_valid_rows": final_valid,
                "n_generated_categories": len(category_summaries),
                "krippendorff_alpha": alpha,
                "alpha_status": status,
                "alpha_reason": status if status != "valid" else "",
                "exact_agreement": exact_agreement,
                "primary_cluster_rule": cluster_rule,
            }
        )

        exact_paired_rows = []
        for paired in paired_rows:
            exact_paired_rows.append(
                {
                    **paired,
                    "reviewer_exact": paired["reviewer_normalized"],
                    "final_exact": paired["final_normalized"],
                }
            )
        exact_alpha, exact_status, exact_agreement = alpha_for_labels(
            exact_paired_rows,
            "reviewer_exact",
            "final_exact",
            f"{pair.variable_name}::exact",
        )
        sensitivity_rows.append(
            {
                "variable_name": pair.variable_name,
                "analysis": "exact_normalized_text",
                "parameter": "normalized_response",
                "krippendorff_alpha": exact_alpha,
                "alpha_status": exact_status,
                "exact_agreement": exact_agreement,
                "n_categories": len({x["reviewer_exact"] for x in exact_paired_rows} | {x["final_exact"] for x in exact_paired_rows}),
            }
        )

        for threshold in SENSITIVITY_THRESHOLDS:
            sens_map, _, _ = cluster_unique_responses(
                pooled,
                embeddings,
                distance_threshold=threshold,
                method_label="sensitivity_agglomerative_average_cosine",
            )
            sens_rows = []
            for paired in paired_rows:
                sens_rows.append(
                    {
                        **paired,
                        "reviewer_sens": sens_map.get(paired["reviewer_normalized"]),
                        "final_sens": sens_map.get(paired["final_normalized"]),
                    }
                )
            sens_alpha, sens_status, sens_agreement = alpha_for_labels(
                sens_rows,
                "reviewer_sens",
                "final_sens",
                f"{pair.variable_name}::threshold_{threshold}",
            )
            sensitivity_rows.append(
                {
                    "variable_name": pair.variable_name,
                    "analysis": "alternative_threshold",
                    "parameter": threshold,
                    "krippendorff_alpha": sens_alpha,
                    "alpha_status": sens_status,
                    "exact_agreement": sens_agreement,
                    "n_categories": len(set(sens_map.values())),
                }
            )

        rare_categories = {cat for cat, count in category_counts.items() if count <= RARE_CATEGORY_MAX_ASSIGNMENTS}
        if rare_categories:
            rare_rows = []
            for paired in paired_rows:
                rcat = paired["reviewer_category"]
                fcat = paired["final_category"]
                rare_rows.append(
                    {
                        **paired,
                        "reviewer_rare": "Rare_Category" if rcat in rare_categories else rcat,
                        "final_rare": "Rare_Category" if fcat in rare_categories else fcat,
                    }
                )
            rare_alpha, rare_status, rare_agreement = alpha_for_labels(
                rare_rows,
                "reviewer_rare",
                "final_rare",
                f"{pair.variable_name}::rare_collapsed",
            )
            rare_n_categories = len({x["reviewer_rare"] for x in rare_rows} | {x["final_rare"] for x in rare_rows})
            sensitivity_rows.append(
                {
                    "variable_name": pair.variable_name,
                    "analysis": "rare_categories_collapsed",
                    "parameter": f"assignment_count<={RARE_CATEGORY_MAX_ASSIGNMENTS}",
                    "krippendorff_alpha": rare_alpha,
                    "alpha_status": rare_status,
                    "exact_agreement": rare_agreement,
                    "n_categories": rare_n_categories,
                }
            )
        else:
            sensitivity_rows.append(
                {
                    "variable_name": pair.variable_name,
                    "analysis": "rare_categories_collapsed",
                    "parameter": "not_run_no_rare_categories",
                    "krippendorff_alpha": None,
                    "alpha_status": "not_run",
                    "exact_agreement": None,
                    "n_categories": len(category_summaries),
                }
            )

    variable_df = pd.DataFrame(variable_rows)
    category_df = pd.DataFrame(all_category_rows)
    assignment_df = pd.DataFrame(all_assignment_rows)
    excluded_df = pd.DataFrame(excluded_rows)
    sensitivity_df = pd.DataFrame(sensitivity_rows)
    detected_pairs_df = pd.DataFrame(detected_pair_rows)
    overall_df = summarize_overall(variable_df, run_notes)
    versions = package_versions(model_name)
    run_config_df = pd.DataFrame(
        [
            ("input_workbook", str(input_workbook)),
            ("worksheet_analyzed", "Sheet1"),
            ("bert_model_name", model_name),
            ("preprocessing_rules", "Raw text preserved; missing markers converted to missing; whitespace normalized; normalized duplicate key lowercased; no meaningful words removed."),
            ("clustering_method", "AgglomerativeClustering(linkage='average', metric='cosine') over pooled unique coder responses per variable."),
            ("primary_clustering_parameters", f"distance_threshold={PRIMARY_DISTANCE_THRESHOLD}; min_unique_for_clustering={MIN_UNIQUE_FOR_CLUSTERING}"),
            ("sensitivity_parameters", f"thresholds={SENSITIVITY_THRESHOLDS}; exact normalized text; rare categories collapsed when assignment_count<={RARE_CATEGORY_MAX_ASSIGNMENTS}"),
            ("random_seed", RANDOM_SEED),
            ("missing_data_rule", "Rows are used for a variable-level alpha only when both coders have valid non-missing responses."),
            ("krippendorff_alpha_measurement_level", "nominal"),
            ("package_versions", json.dumps(versions, sort_keys=True)),
            ("install_commands", "python3 -m pip install sentence-transformers"),
            ("methodological_note", "Semantic-category ICR from BERT-derived clusters is not identical to reliability on manually predefined Results categories."),
        ],
        columns=["setting", "value"],
    )

    outputs = {
        "overall_summary_csv": write_csv(overall_df, output_dir / "overall_summary.csv"),
        "variable_alpha_csv": write_csv(variable_df, output_dir / "variable_alpha.csv"),
        "category_summary_csv": write_csv(category_df, output_dir / "category_summary.csv"),
        "coder_assignments_csv": write_csv(assignment_df, output_dir / "coder_assignments.csv"),
        "excluded_variables_csv": write_csv(excluded_df, output_dir / "excluded_variables.csv"),
        "sensitivity_analysis_csv": write_csv(sensitivity_df, output_dir / "sensitivity_analysis.csv"),
        "run_configuration_csv": write_csv(run_config_df, output_dir / "run_configuration.csv"),
        "detected_pairs_csv": write_csv(detected_pairs_df, output_dir / "detected_pairs.csv"),
    }
    manifest = {
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "input_workbook": str(input_workbook),
        "output_dir": str(output_dir),
        "model_name": model_name,
        "random_seed": RANDOM_SEED,
        "outputs": {key: str(value) for key, value in outputs.items()},
        "overall": dict(zip(overall_df["metric"], overall_df["value"])),
        "run_notes": run_notes,
        "package_versions": versions,
    }
    manifest_path = unique_path(output_dir / "run_manifest.json")
    manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    outputs["run_manifest_json"] = manifest_path

    print("\nVARIABLE-LEVEL BERT SEMANTIC ICR ALPHA")
    print("======================================")
    printable_cols = [
        "variable_name",
        "paired_valid_rows",
        "unique_raw_responses",
        "n_generated_categories",
        "krippendorff_alpha",
        "alpha_status",
        "alpha_reason",
    ]
    with pd.option_context("display.max_rows", None, "display.max_columns", None, "display.width", 240):
        print(variable_df[printable_cols].to_string(index=False))

    print("\nOVERALL SUMMARY")
    print("===============")
    with pd.option_context("display.max_rows", None, "display.max_columns", None, "display.width", 180):
        print(overall_df.to_string(index=False))

    median_row = overall_df.loc[overall_df["metric"] == "median_krippendorff_alpha", "value"]
    median_alpha = median_row.iloc[0] if len(median_row) else None
    print("\nFINAL MEDIAN KRIPPENDORFF'S ALPHA:", median_alpha)
    print("Output directory:", output_dir)

    console_path = unique_path(output_dir / "bert_icr_console_results.txt")
    # Store compact key results; stdout remains the complete printed table.
    console_path.write_text(
        "BERT semantic-category ICR run completed.\n"
        f"Final median Krippendorff alpha: {median_alpha}\n"
        f"Output directory: {output_dir}\n\n"
        + variable_df[printable_cols].to_string(index=False)
        + "\n\n"
        + overall_df.to_string(index=False)
        + "\n",
        encoding="utf-8",
    )
    outputs["console_results_txt"] = console_path

    return {
        "overall": overall_df,
        "variable_alpha": variable_df,
        "category_summary": category_df,
        "coder_assignments": assignment_df,
        "excluded_variables": excluded_df,
        "sensitivity_analysis": sensitivity_df,
        "run_configuration": run_config_df,
        "detected_pairs": detected_pairs_df,
        "outputs": outputs,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run BERT-based semantic-category ICR analysis.")
    parser.add_argument("--input-workbook", type=Path, default=DEFAULT_INPUT_WORKBOOK)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--model-name", default=DEFAULT_MODEL_NAME)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    run_analysis(
        input_workbook=args.input_workbook.expanduser().resolve(),
        output_dir=args.output_dir.expanduser().resolve(),
        model_name=args.model_name,
    )


if __name__ == "__main__":
    main()